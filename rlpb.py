import tkinter as tk
from tkinter import ttk, messagebox, filedialog
try:
    import requests
except Exception:
    requests = None
import json
import datetime
import calendar
import os
import re
import threading
from tkcalendar import DateEntry, Calendar
from lunarcalendar import Converter, Solar, Lunar

# 系统托盘支持
try:
    import pystray
    from PIL import Image, ImageDraw
    TRAY_AVAILABLE = True
except ImportError:
    pystray = None
    Image = None
    ImageDraw = None
    TRAY_AVAILABLE = False

try:
    import pandas as pd
except ImportError:
    pd = None
except Exception:
    pd = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook, load_workbook, Font, PatternFill, Alignment = None, None, None, None
except Exception:
    Workbook, load_workbook, Font, PatternFill, Alignment = None, None, None, None

class _SimpleTooltip:
    """轻量级悬浮提示工具，避免引入额外依赖。
    使用 enter/leave 事件在控件附近显示说明文本。
    """
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None
        self.widget.bind("<Enter>", self._show)
        self.widget.bind("<Leave>", self._hide)

    def _show(self, _evt=None):
        if self.tip or not self.text:
            return
        try:
            x, y, cx, cy = self.widget.bbox("insert") if hasattr(self.widget, 'bbox') else (0, 0, 0, 0)
        except Exception:
            x, y, cx, cy = (0, 0, 0, 0)
        x += self.widget.winfo_rootx() + 20
        y += self.widget.winfo_rooty() + 20
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(self.tip, text=self.text, justify=tk.LEFT,
                       background="#FFFFE0", relief=tk.SOLID, borderwidth=1,
                       font=("Arial", 9))
        lbl.pack(ipadx=6, ipady=3)

    def _hide(self, _evt=None):
        if self.tip:
            try:
                self.tip.destroy()
            except Exception:
                pass
            self.tip = None

class DataValidator:
    """数据验证器类"""

    @staticmethod
    def validate_time_format(time_str):
        """验证时间格式是否为 HH:MM"""
        if not time_str:
            return False, "时间不能为空"

        try:
            hours, minutes = time_str.split(':')
            if not hours.isdigit() or not minutes.isdigit():
                return False, "时间必须为数字"

            hours = int(hours)
            minutes = int(minutes)

            if hours < 0 or hours > 23:
                return False, "小时必须在0-23之间"

            if minutes < 0 or minutes > 59:
                return False, "分钟必须在0-59之间"

            return True, ""
        except ValueError:
            return False, "时间格式错误，应为 HH:MM"

    @staticmethod
    def validate_date_format(date_str):
        """验证日期格式是否为 YYYY-MM-DD"""
        if not date_str:
            return False, "日期不能为空"

        try:
            year, month, day = map(int, date_str.split('-'))
            datetime.date(year, month, day)
            return True, ""
        except ValueError:
            return False, "日期格式错误，应为 YYYY-MM-DD"

    @staticmethod
    def validate_color_format(color_str):
        """验证颜色格式是否为有效的十六进制颜色"""
        if not color_str:
            return False, "颜色不能为空"

        if not color_str.startswith('#'):
            return False, "颜色必须以 # 开头"

        if len(color_str) != 7:
            return False, "颜色格式错误，应为 #RRGGBB"

        try:
            int(color_str[1:], 16)
            return True, ""
        except ValueError:
            return False, "颜色格式错误，应为有效的十六进制颜色"

    @staticmethod
    def validate_shift_name(name, existing_names):
        """验证班次名称"""
        if not name:
            return False, "班次名称不能为空"

        if len(name) > 20:
            return False, "班次名称不能超过20个字符"

        if name in existing_names:
            return False, "班次名称已存在"

        return True, ""

    @staticmethod
    def validate_person_name(name, existing_names):
        """验证人员姓名"""
        if not name:
            return False, "人员姓名不能为空"

        if len(name) > 50:
            return False, "人员姓名不能超过50个字符"

        if name in existing_names:
            return False, "人员姓名已存在"

        return True, ""

class ErrorHandler:
    """错误处理器类"""

    def __init__(self, parent):
        self.parent = parent

    def show_error(self, title, message, details=None):
        """显示错误对话框"""
        if details:
            full_message = f"{message}\n\n详细信息:\n{details}"
        else:
            full_message = message

        messagebox.showerror(title, full_message)

    def show_warning(self, title, message):
        """显示警告对话框"""
        messagebox.showwarning(title, message)

    def show_info(self, title, message):
        """显示信息对话框"""
        messagebox.showinfo(title, message)

    def ask_confirmation(self, title, message):
        """显示确认对话框"""
        return messagebox.askyesno(title, message)

    def handle_validation_errors(self, errors):
        """处理验证错误"""
        if not errors:
            return True

        error_message = "发现以下错误:\n\n" + "\n".join(f"• {error}" for error in errors)
        self.show_error("验证错误", error_message)
        return False

class ShiftScheduler:
    def __init__(self, root):
        self.root = root
        self.root.title("排班日历专业版 v3.0")

        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)

        # 动态计算初始窗口尺寸以适应日历显示
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        # 默认窗口大小：确保能显示完整日历
        init_width = min(1300, int(screen_width * 0.85))
        init_height = min(900, int(screen_height * 0.88))
        # 窗口居中
        x = (screen_width - init_width) // 2
        y = (screen_height - init_height) // 2
        self.root.geometry(f"{init_width}x{init_height}+{x}+{y}")
        self.root.minsize(1100, 750)

        # 字体设置（必须在setup_modern_styles之前初始化）
        self.font_family = tk.StringVar(value="Microsoft YaHei UI")
        self.font_size = tk.IntVar(value=10)
        self.multi_calendar_font_size = tk.IntVar(value=9)  # 多人日历格子字体大小

        # 主题设置
        self.theme_var = tk.StringVar(value="light")

        # 系统托盘设置
        self.minimize_to_tray = tk.BooleanVar(value=False)
        self.tray_icon = None  # 托盘图标对象

        # 设置现代化样式
        self.setup_modern_styles()

        # 初始化错误处理器
        self.error_handler = ErrorHandler(self.root)

        # 设置应用图标（如果有的话）
        try:
            self.root.iconbitmap("app_icon.ico")
        except:
            pass

        # 初始化数据结构
        self.shift_types = {
            "白班": {"start_time": "08:00", "end_time": "17:00", "color": "#FFE4B5"},  # 浅橙色背景
            "夜班": {"start_time": "20:00", "end_time": "08:00", "color": "#4B0082"},  # 靛蓝色背景
            "休息": {"start_time": "00:00", "end_time": "00:00", "color": "#CCFFCC"}
        }

        self.shift_schedules = {}
        self.swap_records = {}  # 调换班记录: {date_str: [{person_a, person_b, timestamp}]}
        self.current_schedule = None
        self.current_plan_name = None
        self.current_date = datetime.date.today()
        # 配额年份选择变量
        # 配额年份变量
        current_date = datetime.date.today()
        # 根据年休假规则设置默认年份：4-12月用当年，1-3月用去年
        if current_date.month >= 4:
            default_leave_year = current_date.year
        else:
            default_leave_year = current_date.year - 1
        self.quota_year_var = tk.StringVar(value=str(default_leave_year))
        # 请假数据
        self.leave_types = ["事假", "病假"]
        # 记录项: {"plan_name": str, "date": "YYYY-MM-DD", "type": str, "note": str}
        self.leave_records = []
        # 年度配额: {plan_name: {year: {type: quota_int}}}
        self.leave_quotas = {}
        # 节假日复制粘贴临时存储
        self._holidays_clipboard = None  # {"year": str, "data": {"MM-DD": "名称"}}

        # 节假日数据（可持久化覆盖），默认包含若干重要节假日
        self.holidays = {
            "2023": {"01-01":"元旦","01-22":"春节","05-01":"劳动节","10-01":"国庆"},
            "2024": {"01-01":"元旦","02-10":"春节","05-01":"劳动节","10-01":"国庆"},
            "2025": {"01-01":"元旦","05-01":"劳动节","10-01":"国庆"}
        }

        # 视图偏好设置：是否显示节假日与请假标识
        self.show_holidays = tk.BooleanVar(value=True)
        self.show_leaves = tk.BooleanVar(value=True)

        # 备份设置
        self.backup_enabled = tk.BooleanVar(value=True)
        self.backup_interval = tk.IntVar(value=1)  # 每天备份一次
        self.backup_count_limit = tk.IntVar(value=30)  # 保留30个备份
        self.backup_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
        self.last_backup_time = None

        # 创建备份目录
        if not os.path.exists(self.backup_directory):
            os.makedirs(self.backup_directory)

        # 延迟初始化：避免在构造函数中进行耗时操作
        self._data_loaded = False
        self.create_widgets()
        self.setup_shortcuts()
        self.root.after(100, self._delayed_init)  # 在GUI创建后异步加载数据

    def _delayed_init(self):
        """延迟初始化操作，避免阻塞GUI创建"""
        self.load_data()
        # 加载数据后应用字体设置
        self.setup_modern_styles()
        self._data_loaded = True

    def setup_modern_styles(self):
        """设置现代化样式 - 支持亮色/暗色主题"""
        style = ttk.Style()

        # 设置主题
        try:
            style.theme_use('clam')  # 使用clam主题作为基础
        except:
            pass

        # 根据选择的主题设置颜色方案
        current_theme = self.theme_var.get()

        if current_theme == "dark":
            # 暗色调主题颜色方案
            self.colors = {
                'primary': '#4A9EFF',        # 明亮的蓝色
                'primary_dark': '#3D8FE6',   # 深一点的蓝色（悬停）
                'primary_light': '#2D3748',  # 深蓝背景
                'secondary': '#9F7AEA',      # 浅紫色强调
                'success': '#48BB78',        # 清新绿色
                'success_light': '#2D3748',  # 深绿背景
                'danger': '#F56565',         # 明亮的红色
                'danger_light': '#2D3748',   # 深红背景
                'warning': '#ED8936',        # 明亮的橙色
                'warning_light': '#2D3748',  # 深橙背景
                'info': '#4299E1',           # 明亮的天蓝色
                'info_light': '#2D3748',     # 深天蓝背景
                'bg_main': '#1A202C',        # 主背景色（深灰）
                'bg_card': '#2D3748',        # 卡片背景（深蓝灰）
                'bg_hover': '#4A5568',       # 悬停背景
                'border': '#4A5568',         # 边框色
                'border_light': '#4A5568',   # 浅边框
                'text_primary': '#F7FAFC',   # 主文字色（白色）
                'text_secondary': '#CBD5E0', # 次要文字色（浅灰）
                'text_muted': '#718096',     # 弱化文字色（中灰）
                'white': '#F7FAFC',          # 白色文字
                'shadow': '#2D3748',         # 阴影色
                # 日历专用色（暗色主题）
                'today_bg': '#2D3748',       # 今日背景（深蓝灰）
                'today_border': '#4A9EFF',   # 今日边框（明亮蓝色）
                'weekend_bg': '#374151',     # 周末背景（更深）
                'calendar_header': '#4A5568', # 日历星期标题背景
            }
        else:
            # 亮色调主题颜色方案（默认）
            self.colors = {
                'primary': '#5B8DEF',        # 柔和的蓝色
                'primary_dark': '#4A7BD9',   # 深一点的蓝色（悬停）
                'primary_light': '#E8F0FE',  # 浅蓝背景
                'secondary': '#6C5CE7',      # 紫色强调
                'success': '#00B894',        # 清新绿色
                'success_light': '#E8F8F5',  # 浅绿背景
                'danger': '#E17055',         # 柔和的红色
                'danger_light': '#FDEAEA',   # 浅红背景
                'warning': '#FDCB6E',        # 温暖的黄色
                'warning_light': '#FEF9E7',  # 浅黄背景
                'info': '#74B9FF',           # 天蓝色
                'info_light': '#EBF5FB',     # 浅天蓝背景
                'bg_main': '#F8FAFC',        # 主背景色（极浅灰蓝）
                'bg_card': '#FFFFFF',        # 卡片背景
                'bg_hover': '#F1F5F9',       # 悬停背景
                'border': '#E2E8F0',         # 边框色
                'border_light': '#F1F5F9',   # 浅边框
                'text_primary': '#1E293B',   # 主文字色（深灰蓝）
                'text_secondary': '#64748B', # 次要文字色
                'text_muted': '#94A3B8',     # 弱化文字色
                'white': '#FFFFFF',
                'shadow': '#CBD5E1',         # 阴影色
                # 日历专用色
                'today_bg': '#EEF2FF',       # 今日背景（淡紫蓝）
                'today_border': '#818CF8',   # 今日边框（紫色）
                'weekend_bg': '#FEF7FF',     # 周末背景（淡粉紫）
                'calendar_header': '#F1F5F9', # 日历星期标题背景
            }

        # 获取当前字体设置
        font_family = self.font_family.get()
        font_size = self.font_size.get()

        # ============ 基础样式配置 ============
        # 主框架 - 使用浅灰蓝背景
        style.configure('TFrame', background=self.colors['bg_main'])

        # 标签样式
        style.configure('TLabel',
                        background=self.colors['bg_main'],
                        foreground=self.colors['text_primary'],
                        font=(font_family, font_size))

        # 标题标签样式
        style.configure('Title.TLabel',
                        background=self.colors['bg_main'],
                        foreground=self.colors['primary'],
                        font=(font_family, font_size + 4, 'bold'))

        # ============ 按钮样式配置 ============
        # 默认按钮 - 圆角感、柔和阴影
        style.configure('TButton',
                        font=(font_family, font_size),
                        padding=(12, 6),
                        background=self.colors['bg_card'],
                        foreground=self.colors['text_primary'],
                        borderwidth=1,
                        relief='flat')
        style.map('TButton',
                  background=[('active', self.colors['bg_hover']), ('pressed', self.colors['border'])],
                  foreground=[('active', self.colors['primary'])])

        # 主要按钮 - 蓝色填充
        style.configure('Primary.TButton',
                        font=(font_family, font_size, 'bold'),
                        padding=(12, 6),
                        background=self.colors['primary'],
                        foreground=self.colors['white'])
        style.map('Primary.TButton',
                  background=[('active', self.colors['primary_dark']), ('pressed', self.colors['primary_dark'])])

        # 成功按钮 - 绿色
        style.configure('Success.TButton',
                        font=(font_family, font_size, 'bold'),
                        padding=(12, 6),
                        background=self.colors['success'],
                        foreground=self.colors['white'])
        style.map('Success.TButton',
                  background=[('active', '#00A884'), ('pressed', '#009975')])

        # 危险按钮 - 红色
        style.configure('Danger.TButton',
                        font=(font_family, font_size, 'bold'),
                        padding=(12, 6),
                        background=self.colors['danger'],
                        foreground=self.colors['white'])
        style.map('Danger.TButton',
                  background=[('active', '#D63031'), ('pressed', '#C0392B')])

        # 信息按钮 - 天蓝色
        style.configure('Info.TButton',
                        font=(font_family, font_size, 'bold'),
                        padding=(12, 6),
                        background=self.colors['info'],
                        foreground=self.colors['white'])
        style.map('Info.TButton',
                  background=[('active', '#5DADE2'), ('pressed', '#3498DB')])

        # 小按钮样式
        style.configure('Small.TButton',
                        font=(font_family, max(8, font_size - 2)),
                        padding=(6, 3))

        # ============ 标签页样式配置 ============
        style.configure('TNotebook',
                        background=self.colors['bg_main'],
                        borderwidth=0)
        style.configure('TNotebook.Tab',
                        background=self.colors['bg_card'],
                        foreground=self.colors['text_secondary'],
                        padding=[16, 10],
                        font=(font_family, font_size, 'bold'),
                        borderwidth=0)
        style.map('TNotebook.Tab',
                  background=[('selected', self.colors['primary']), ('active', self.colors['primary_light'])],
                  foreground=[('selected', self.colors['white']), ('active', self.colors['primary'])])

        # ============ Treeview样式配置 ============
        style.configure('Treeview',
                        background=self.colors['bg_card'],
                        foreground=self.colors['text_primary'],
                        fieldbackground=self.colors['bg_card'],
                        font=(font_family, max(9, font_size - 1)),
                        rowheight=32,
                        borderwidth=0)
        style.configure('Treeview.Heading',
                        background=self.colors['primary'],
                        foreground=self.colors['white'],
                        font=(font_family, font_size, 'bold'),
                        padding=(8, 6))
        style.map('Treeview.Heading',
                  background=[('active', self.colors['primary_dark'])])
        style.map('Treeview',
                  background=[('selected', self.colors['primary_light'])],
                  foreground=[('selected', self.colors['primary'])])

        # ============ 输入框样式配置 ============
        style.configure('TEntry',
                        font=(font_family, font_size),
                        padding=10,
                        fieldbackground=self.colors['bg_card'],
                        borderwidth=1,
                        relief='solid')

        style.configure('TCombobox',
                        font=(font_family, font_size),
                        padding=8,
                        background=self.colors['bg_card'],
                        fieldbackground=self.colors['bg_card'],
                        arrowsize=14)
        style.map('TCombobox',
                  fieldbackground=[('readonly', self.colors['bg_card'])],
                  background=[('readonly', self.colors['bg_card'])])

        # ============ LabelFrame样式配置 ============
        style.configure('TLabelframe',
                        background=self.colors['bg_main'],
                        foreground=self.colors['text_primary'],
                        borderwidth=1,
                        relief='solid')
        style.configure('TLabelframe.Label',
                        background=self.colors['bg_main'],
                        foreground=self.colors['primary'],
                        font=(font_family, font_size, 'bold'))

        # ============ 滚动条样式配置 ============
        style.configure('TScrollbar',
                        background=self.colors['border'],
                        troughcolor=self.colors['bg_main'],
                        borderwidth=0,
                        arrowsize=14)
        style.map('TScrollbar',
                  background=[('active', self.colors['text_muted']), ('pressed', self.colors['text_secondary'])])

        # ============ 状态栏样式 ============
        style.configure('StatusBar.TLabel',
                        background=self.colors['border'],
                        foreground=self.colors['text_secondary'],
                        font=(font_family, max(8, font_size - 2)),
                        padding=(10, 5),
                        relief='flat')

        # ============ 卡片框架样式 ============
        style.configure('Card.TFrame',
                        background=self.colors['bg_card'],
                        borderwidth=1,
                        relief='solid')
    def create_widgets(self):
        """创建主界面布局 - 优化启动速度"""
        # 创建主容器
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 创建标题栏
        self.create_header(main_container)

        # 创建标签页容器
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # 状态栏 - 必须先初始化
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self.root, textvariable=self.status_var,
                             style='StatusBar.TLabel')
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

        # 标记标签页初始化状态
        self._tabs_initialized = {
            'shift_type': False,
            'schedule': False,
            'leave': False,
            'holiday': False,
            'calendar': False,
            'multi_calendar': False,
            'holiday_calendar': False,
            'swap_management': False
        }

        # 优先创建第一个标签页（首屏）
        self.setup_shift_type_tab()
        self._tabs_initialized['shift_type'] = True

        # 创建占位标签页（延迟加载内容）
        self._create_placeholder_tabs()

        # 绑定标签页切换事件 - 延迟加载
        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        self.update_status("系统启动中...")

        # 延迟初始化其他标签页
        self.root.after(50, self._delayed_ui_init)

    def _create_placeholder_tabs(self):
        """创建占位标签页框架"""
        # 排班计划管理
        self.schedule_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.schedule_frame, text="排班计划管理")

        # 请假管理
        self.leave_frame_placeholder = ttk.Frame(self.notebook)
        self.notebook.add(self.leave_frame_placeholder, text="🔥 请假管理")

        # 节假日管理
        self.holiday_frame_placeholder = ttk.Frame(self.notebook)
        self.notebook.add(self.holiday_frame_placeholder, text="节假日管理")

        # 调班管理
        self.swap_management_frame_placeholder = ttk.Frame(self.notebook)
        self.notebook.add(self.swap_management_frame_placeholder, text="调班管理")

        # 单人日历视图（隐藏，保留代码）
        self.calendar_frame_placeholder = ttk.Frame(self.notebook)
        # self.notebook.add(self.calendar_frame_placeholder, text="单人日历视图")

        # 多人日历视图
        self.multi_calendar_frame_placeholder = ttk.Frame(self.notebook)
        self.notebook.add(self.multi_calendar_frame_placeholder, text="多人日历视图")

        # 当月休假日历
        self.holiday_calendar_frame_placeholder = ttk.Frame(self.notebook)
        self.notebook.add(self.holiday_calendar_frame_placeholder, text="当月休假日历")

    def _on_tab_changed(self, event):
        """标签页切换事件 - 性能优化的延迟加载"""
        try:
            current_tab = self.notebook.index(self.notebook.select())
            tab_names = ['shift_type', 'schedule', 'leave', 'holiday', 'swap_management', 'multi_calendar', 'holiday_calendar']

            if current_tab < len(tab_names):
                tab_name = tab_names[current_tab]

                # 性能优化：对于已经初始化的标签页，不再重复加载
                if not self._tabs_initialized.get(tab_name, False):
                    # 使用延迟加载，避免界面卡顿
                    self.root.after(50, lambda: self._load_tab_content(tab_name, current_tab))
                    # 首次加载后延迟更长时间再调整窗口，确保内容渲染完成
                    if tab_name in ['calendar', 'multi_calendar', 'holiday_calendar']:
                        self.update_status(f"正在加载{self._get_tab_display_name(tab_name)}...")
                        self.root.after(500, self._auto_fit_calendar_display)
                        self.root.after(600, lambda: self.update_status(f"{self._get_tab_display_name(tab_name)}已加载"))
                else:
                    # 已初始化的标签页
                    if tab_name == 'multi_calendar':
                        # 多人日历视图：检查是否是首次显示
                        if hasattr(self, '_multi_calendar_first_show') and self._multi_calendar_first_show:
                            # 首次显示，调用初始化渲染（只渲染一次）
                            self._init_multi_calendar_delayed()
                            self._multi_calendar_first_show = False
                            self._multi_calendar_rendered = True
                        elif not self._multi_calendar_rendered:
                            # 非首次但需要刷新
                            self.root.after(100, self.update_multi_calendar)

                    # 日历相关标签页，调整窗口大小
                    if tab_name in ['calendar', 'multi_calendar', 'holiday_calendar']:
                        self.update_status(f"正在加载{self._get_tab_display_name(tab_name)}...")
                        self.root.after(150, self._auto_fit_calendar_display)
                        self.root.after(250, lambda: self.update_status(f"{self._get_tab_display_name(tab_name)}已加载"))
        except Exception as e:
            print(f"标签页切换出错：{e}")
            pass

    def _get_tab_display_name(self, tab_name):
        """获取标签页的显示名称"""
        name_map = {
            'calendar': '单人日历视图',
            'multi_calendar': '多人日历视图',
            'holiday_calendar': '当月休假日历',
            'swap_management': '调班管理'
        }
        return name_map.get(tab_name, tab_name)

    def _auto_fit_calendar_display(self):
        """自动调整窗口大小以完整显示日历所有格子"""
        try:
            # 获取屏幕尺寸
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()

            # 获取当前多人日历字体大小
            mc_font_size = self.multi_calendar_font_size.get()

            # 根据字体大小动态计算所需尺寸
            # 基础：字体9pt时，每格约115px宽，100px高
            # 字体每增加1pt，格子需要增加约8px
            base_cell_width = 115
            base_cell_height = 100
            font_scale = (mc_font_size - 9) * 8

            cell_width = base_cell_width + font_scale
            cell_height = base_cell_height + font_scale

            # 计算日历所需的最小尺寸
            # 7列 + 边距 + 右侧统计面板(约300px)
            min_calendar_width = 7 * cell_width + 40 + 300
            # 7行(标题+6周) + 控制栏 + 图例 + 边距
            min_calendar_height = 7 * cell_height + 200

            # 计算最佳窗口尺寸（在屏幕范围内，最大90%屏幕）
            optimal_width = min(max(int(min_calendar_width), 1200), int(screen_width * 0.92))
            optimal_height = min(max(int(min_calendar_height), 850), int(screen_height * 0.88))

            # 获取当前窗口尺寸
            current_width = self.root.winfo_width()
            current_height = self.root.winfo_height()

            # 判断是否需要调整
            need_resize = False
            new_width = current_width
            new_height = current_height

            if current_width < optimal_width:
                new_width = optimal_width
                need_resize = True
            if current_height < optimal_height:
                new_height = optimal_height
                need_resize = True

            if need_resize:
                # 居中显示
                x = max(0, (screen_width - new_width) // 2)
                y = max(0, (screen_height - new_height) // 2)

                self.root.geometry(f"{new_width}x{new_height}+{x}+{y}")
                self.root.update_idletasks()  # 强制更新界面
                self.update_status("窗口已自动调整以完整显示日历")
        except Exception as e:
            print(f"自动调整窗口失败: {e}")
            pass

    def _load_tab_content(self, tab_name, tab_index):
        """延迟加载标签页内容"""
        try:
            if tab_name == 'schedule' and not self._tabs_initialized['schedule']:
                # 删除占位框架，创建实际内容
                self.notebook.forget(tab_index)
                self.setup_schedule_tab()
                # 移动到正确位置
                self._reorder_tab(1)
                self._tabs_initialized['schedule'] = True

            elif tab_name == 'leave' and not self._tabs_initialized['leave']:
                self.notebook.forget(tab_index)
                self.setup_leave_tab()
                self._reorder_tab(2)
                self._tabs_initialized['leave'] = True

            elif tab_name == 'holiday' and not self._tabs_initialized['holiday']:
                self.notebook.forget(tab_index)
                self.setup_holiday_tab()
                self._reorder_tab(3)
                self._tabs_initialized['holiday'] = True

            elif tab_name == 'swap_management' and not self._tabs_initialized['swap_management']:
                self.notebook.forget(tab_index)
                self.setup_swap_management_tab()
                self._reorder_tab(4)
                self._tabs_initialized['swap_management'] = True

            # 单人日历视图已隐藏，保留代码但不加载
            # elif tab_name == 'calendar' and not self._tabs_initialized['calendar']:
            #     self.notebook.forget(tab_index)
            #     self.setup_calendar_tab()
            #     self._reorder_tab(4)
            #     self._tabs_initialized['calendar'] = True

            elif tab_name == 'multi_calendar' and not self._tabs_initialized['multi_calendar']:
                self.notebook.forget(tab_index)
                self.setup_multi_member_calendar_tab()
                self._reorder_tab(5)
                self._tabs_initialized['multi_calendar'] = True

            elif tab_name == 'holiday_calendar' and not self._tabs_initialized['holiday_calendar']:
                self.notebook.forget(tab_index)
                self.setup_holiday_calendar_tab()
                self._reorder_tab(6)
                self._tabs_initialized['holiday_calendar'] = True

            # 选择刚加载的标签页
            self.notebook.select(tab_index)

        except Exception as e:
            print(f"加载标签页失败: {e}")

    def _reorder_tab(self, target_index):
        """重新排序标签页到目标位置"""
        # 获取当前标签页数量
        tab_count = self.notebook.index('end')
        if tab_count > 0:
            # 将最后一个标签页移动到目标位置
            last_tab = self.notebook.tabs()[-1]
            self.notebook.insert(target_index, last_tab)

    def _delayed_ui_init(self):
        """延迟UI初始化操作，避免阻塞GUI创建"""
        # 更新状态
        self.update_status("系统已就绪")

        # 初始化配额年份选项（如果已存在）
        if hasattr(self, 'quota_year_combo'):
            self.update_quota_year_options()

        # 预加载常用标签页（在后台异步加载）
        self.root.after(100, self._preload_common_tabs)

    def _preload_common_tabs(self):
        """预加载常用标签页"""
        # 按优先级预加载：排班计划管理是第二常用的
        if not self._tabs_initialized.get('schedule', False):
            try:
                self.notebook.forget(1)  # 删除占位
                self.setup_schedule_tab()
                self._reorder_tab(1)
                self._tabs_initialized['schedule'] = True
            except Exception:
                pass

    def create_header(self, parent):
        """创建现代化标题栏 - 清新简约风格"""
        # 标题栏容器 - 使用卡片背景
        header_container = tk.Frame(parent, bg=self.colors['bg_card'])
        header_container.pack(fill=tk.X, pady=(0, 15))

        # 内部框架
        header_frame = tk.Frame(header_container, bg=self.colors['bg_card'])
        header_frame.pack(fill=tk.X, padx=15, pady=12)

        # 左侧：应用标题和副标题
        title_frame = tk.Frame(header_frame, bg=self.colors['bg_card'])
        title_frame.pack(side=tk.LEFT)

        # 主标题
        title_label = tk.Label(title_frame, text="排班日历",
                              font=('Microsoft YaHei UI', 18, 'bold'),
                              bg=self.colors['bg_card'],
                              fg=self.colors['primary'])
        title_label.pack(side=tk.LEFT)

        # 版本标签
        version_badge = tk.Frame(title_frame, bg=self.colors['primary_light'])
        version_badge.pack(side=tk.LEFT, padx=(10, 0))
        version_label = tk.Label(version_badge, text="v3.0",
                                font=('Microsoft YaHei UI', 9, 'bold'),
                                bg=self.colors['primary_light'],
                                fg=self.colors['primary'],
                                padx=8, pady=2)
        version_label.pack()

        # 副标题
        subtitle_label = tk.Label(title_frame, text="专业版",
                                 font=('Microsoft YaHei UI', 10),
                                 bg=self.colors['bg_card'],
                                 fg=self.colors['text_muted'],
                                 padx=8)
        subtitle_label.pack(side=tk.LEFT)

        # 右侧：工具栏
        toolbar_frame = tk.Frame(header_frame, bg=self.colors['bg_card'])
        toolbar_frame.pack(side=tk.RIGHT)

        # 工具栏按钮 - 使用更现代的样式
        toolbar_buttons = [
            ("导入", self.import_data_from_json, self.colors['primary']),
            ("统计", self.show_statistics, self.colors['primary']),
            ("报表", self.show_report_generator, self.colors['success']),
            ("备份", self.backup_data, self.colors['warning']),
            ("恢复", self.restore_data, self.colors['info']),
            ("搜索", self.show_search_dialog, self.colors['primary']),
            ("设置", self.show_settings, self.colors['text_secondary']),
        ]

        for btn_text, btn_command, btn_color in toolbar_buttons:
            btn_frame = tk.Frame(toolbar_frame, bg=btn_color)
            btn_frame.pack(side=tk.LEFT, padx=3)

            btn = tk.Label(btn_frame, text=btn_text,
                          font=('Microsoft YaHei UI', 9, 'bold'),
                          bg=btn_color, fg=self.colors['white'],
                          padx=12, pady=6, cursor='hand2')
            btn.pack()

            # 绑定点击事件
            btn.bind('<Button-1>', lambda e, cmd=btn_command: cmd())
            # 绑定悬停效果
            btn.bind('<Enter>', lambda e, f=btn_frame, c=btn_color: self._on_toolbar_btn_enter(f, c))
            btn.bind('<Leave>', lambda e, f=btn_frame, c=btn_color: self._on_toolbar_btn_leave(f, c))

    def _on_toolbar_btn_enter(self, frame, color):
        """工具栏按钮悬停进入"""
        # 稍微变暗
        frame.config(bg=self._darken_color(color, 0.1))
        for child in frame.winfo_children():
            child.config(bg=self._darken_color(color, 0.1))

    def _on_toolbar_btn_leave(self, frame, color):
        """工具栏按钮悬停离开"""
        frame.config(bg=color)
        for child in frame.winfo_children():
            child.config(bg=color)

    def _darken_color(self, hex_color, factor=0.1):
        """将颜色变暗"""
        hex_color = hex_color.lstrip('#')
        r = max(0, int(int(hex_color[0:2], 16) * (1 - factor)))
        g = max(0, int(int(hex_color[2:4], 16) * (1 - factor)))
        b = max(0, int(int(hex_color[4:6], 16) * (1 - factor)))
        return f'#{r:02x}{g:02x}{b:02x}'

    def apply_theme(self, theme):
        """应用主题切换"""
        try:
            # 更新主题变量
            self.theme_var.set(theme)

            # 重新配置样式
            self.setup_modern_styles()

            # 更新所有已创建的界面元素
            self.update_all_widgets_theme()

            # 保存主题设置
            self.save_data()

            # 更新状态栏
            theme_name = "亮色调" if theme == "light" else "暗色调"
            self.update_status(f"已切换到{theme_name}主题")

        except Exception as e:
            # 如果出错，恢复默认主题
            self.theme_var.set("light")
            self.setup_modern_styles()
            self.update_status("主题切换失败，已恢复默认主题")
            messagebox.showerror("错误", f"主题切换失败：{str(e)}")

    def update_all_widgets_theme(self):
        """更新所有界面元素的主题"""
        try:
            # 更新主窗口背景色
            if hasattr(self, 'root'):
                self.root.configure(bg=self.colors['bg_main'])

            # 更新标题栏
            if hasattr(self, 'header_container'):
                self.header_container.configure(bg=self.colors['bg_card'])
                self.header_frame.configure(bg=self.colors['bg_card'])
                self.title_frame.configure(bg=self.colors['bg_card'])
                self.toolbar_frame.configure(bg=self.colors['bg_card'])

                # 更新标题文字
                for widget in self.title_frame.winfo_children():
                    if isinstance(widget, tk.Label):
                        widget.configure(bg=self.colors['bg_card'], fg=self.colors.get('primary', self.colors['text_primary']))
                    elif isinstance(widget, tk.Frame):
                        widget.configure(bg=self.colors.get('primary_light', self.colors['bg_card']))
                        for child in widget.winfo_children():
                            if isinstance(child, tk.Label):
                                child.configure(bg=self.colors.get('primary_light', self.colors['bg_card']), fg=self.colors.get('primary', self.colors['text_primary']))

                # 更新工具栏按钮
                for btn_frame in self.toolbar_frame.winfo_children():
                    if isinstance(btn_frame, tk.Frame):
                        btn_color = btn_frame.cget('bg')
                        # 获取对应的主题颜色
                        if 'primary' in str(btn_color):
                            new_color = self.colors['primary']
                        elif 'success' in str(btn_color):
                            new_color = self.colors['success']
                        elif 'warning' in str(btn_color):
                            new_color = self.colors['warning']
                        elif 'info' in str(btn_color):
                            new_color = self.colors['info']
                        else:
                            new_color = self.colors['text_secondary']

                        btn_frame.configure(bg=new_color)
                        for child in btn_frame.winfo_children():
                            if isinstance(child, tk.Label):
                                child.configure(bg=new_color, fg=self.colors['white'])

            # 更新日历视图
            if hasattr(self, 'calendar_frame'):
                self.update_calendar_theme()

            # 更新状态栏
            if hasattr(self, 'status_var'):
                # 状态栏会在setup_modern_styles中自动更新
                pass

            # 强制刷新所有界面
            self.root.update_idletasks()

        except Exception as e:
            print(f"更新界面主题时出错：{str(e)}")

    def update_calendar_theme(self):
        """更新日历的主题颜色"""
        try:
            if hasattr(self, 'calendar_frame'):
                # 重新渲染日历以应用新主题
                self.update_calendar()
        except Exception as e:
            print(f"更新日历主题时出错：{str(e)}")

    def show_statistics(self):
        """显示统计分析仪表板"""
        stats_dialog = tk.Toplevel(self.root)
        stats_dialog.title("统计分析仪表板")
        stats_dialog.geometry("900x700")
        stats_dialog.resizable(True, True)

        # 设置对话框居中
        stats_dialog.transient(self.root)
        stats_dialog.grab_set()

        # 创建统计界面
        main_frame = ttk.Frame(stats_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = ttk.Label(main_frame, text="统计分析仪表板",
                             font=('Microsoft YaHei UI', 16, 'bold'),
                             foreground=self.colors['primary'])
        title_label.pack(pady=(0, 20))

        # 创建标签页
        stats_notebook = ttk.Notebook(main_frame)
        stats_notebook.pack(fill=tk.BOTH, expand=True)

        # 排班统计标签页
        self.create_shift_stats_tab(stats_notebook)

        # 请假统计标签页
        self.create_leave_stats_tab(stats_notebook)

        # 人员统计标签页
        self.create_person_stats_tab(stats_notebook)

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        ttk.Button(button_frame, text="刷新数据", command=lambda: self.refresh_statistics(stats_notebook),
                   style='Primary.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="导出报表", command=self.export_statistics_report,
                   style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="关闭", command=stats_dialog.destroy,
                   style='Danger.TButton').pack(side=tk.RIGHT)

        # 初始化数据
        self.refresh_statistics(stats_notebook)

    def create_shift_stats_tab(self, parent):
        """创建排班统计标签页"""
        frame = ttk.Frame(parent)
        parent.add(frame, text="排班统计")

        # 统计卡片区域
        cards_frame = ttk.Frame(frame)
        cards_frame.pack(fill=tk.X, pady=(0, 20))

        # 班次类型统计
        shift_card = self.create_stats_card(cards_frame, "班次类型统计", self.get_shift_type_stats())
        shift_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 人员统计
        person_card = self.create_stats_card(cards_frame, "人员统计", self.get_person_stats())
        person_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 排班分布统计
        distribution_frame = ttk.Frame(frame)
        distribution_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(distribution_frame, text="排班分布统计",
                  font=('Microsoft YaHei UI', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))

        # 创建排班分布树视图
        tree_frame = ttk.Frame(distribution_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.shift_distribution_tree = ttk.Treeview(tree_frame, columns=("person", "shift_type", "count", "percentage"),
                                                   show="headings", yscrollcommand=scrollbar.set)
        self.shift_distribution_tree.heading("person", text="人员")
        self.shift_distribution_tree.heading("shift_type", text="班次类型")
        self.shift_distribution_tree.heading("count", text="次数")
        self.shift_distribution_tree.heading("percentage", text="占比")

        self.shift_distribution_tree.column("person", width=120)
        self.shift_distribution_tree.column("shift_type", width=120)
        self.shift_distribution_tree.column("count", width=80)
        self.shift_distribution_tree.column("percentage", width=100)

        self.shift_distribution_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.shift_distribution_tree.yview)

    def create_leave_stats_tab(self, parent):
        """创建请假统计标签页"""
        frame = ttk.Frame(parent)
        parent.add(frame, text="请假统计")

        # 统计卡片区域
        cards_frame = ttk.Frame(frame)
        cards_frame.pack(fill=tk.X, pady=(0, 20))

        # 请假类型统计
        leave_type_card = self.create_stats_card(cards_frame, "请假类型统计", self.get_leave_type_stats())
        leave_type_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 月度请假统计
        monthly_card = self.create_stats_card(cards_frame, "月度请假统计", self.get_monthly_leave_stats())
        monthly_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 请假记录列表
        records_frame = ttk.Frame(frame)
        records_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(records_frame, text="最近请假记录",
                  font=('Microsoft YaHei UI', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))

        # 创建请假记录树视图
        tree_frame = ttk.Frame(records_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.leave_stats_tree = ttk.Treeview(tree_frame, columns=("person", "date", "type", "days_ago"),
                                             show="headings", yscrollcommand=scrollbar.set, height=10)
        self.leave_stats_tree.heading("person", text="人员")
        self.leave_stats_tree.heading("date", text="日期")
        self.leave_stats_tree.heading("type", text="类型")
        self.leave_stats_tree.heading("days_ago", text="天数前")

        self.leave_stats_tree.column("person", width=120)
        self.leave_stats_tree.column("date", width=100)
        self.leave_stats_tree.column("type", width=100)
        self.leave_stats_tree.column("days_ago", width=80)

        self.leave_stats_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.leave_stats_tree.yview)

    def create_person_stats_tab(self, parent):
        """创建人员统计标签页"""
        frame = ttk.Frame(parent)
        parent.add(frame, text="人员统计")

        # 统计卡片区域
        cards_frame = ttk.Frame(frame)
        cards_frame.pack(fill=tk.X, pady=(0, 20))

        # 出勤率统计
        attendance_card = self.create_stats_card(cards_frame, "出勤率统计", self.get_attendance_stats())
        attendance_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 请假率统计
        leave_rate_card = self.create_stats_card(cards_frame, "请假率统计", self.get_leave_rate_stats())
        leave_rate_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 人员详细统计
        details_frame = ttk.Frame(frame)
        details_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(details_frame, text="人员详细统计",
                  font=('Microsoft YaHei UI', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))

        # 创建人员统计树视图
        tree_frame = ttk.Frame(details_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.person_stats_tree = ttk.Treeview(tree_frame, columns=("person", "total_shifts", "leave_days", "attendance_rate"),
                                              show="headings", yscrollcommand=scrollbar.set)
        self.person_stats_tree.heading("person", text="人员")
        self.person_stats_tree.heading("total_shifts", text="总班次")
        self.person_stats_tree.heading("leave_days", text="请假天数")
        self.person_stats_tree.heading("attendance_rate", text="出勤率")

        self.person_stats_tree.column("person", width=120)
        self.person_stats_tree.column("total_shifts", width=100)
        self.person_stats_tree.column("leave_days", width=100)
        self.person_stats_tree.column("attendance_rate", width=100)

        self.person_stats_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.person_stats_tree.yview)

    def create_stats_card(self, parent, title, data):
        """创建统计卡片"""
        card_frame = ttk.Frame(parent, relief=tk.RAISED, borderwidth=1)
        card_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 卡片标题
        title_label = ttk.Label(card_frame, text=title,
                             font=('Microsoft YaHei UI', 12, 'bold'),
                             foreground=self.colors['primary'])
        title_label.pack(pady=(10, 5))

        # 卡片内容
        content_frame = ttk.Frame(card_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # 添加统计数据
        for key, value in data.items():
            item_frame = ttk.Frame(content_frame)
            item_frame.pack(fill=tk.X, pady=2)

            ttk.Label(item_frame, text=f"{key}:",
                      font=('Microsoft YaHei UI', 10)).pack(side=tk.LEFT)
            ttk.Label(item_frame, text=str(value),
                      font=('Microsoft YaHei UI', 10, 'bold'),
                      foreground=self.colors['dark']).pack(side=tk.RIGHT)

        return card_frame

    def get_shift_type_stats(self):
        """获取班次类型统计"""
        stats = {}
        stats["班次类型数"] = len(self.shift_types)
        stats["活跃班次"] = len([s for s in self.shift_types.values() if s['start_time'] != "00:00"])
        stats["休息班次"] = len([s for s in self.shift_types.values() if s['start_time'] == "00:00"])
        return stats

    def get_person_stats(self):
        """获取人员统计"""
        stats = {}
        stats["总人数"] = len(self.shift_schedules)
        stats["有排班人数"] = len([s for s in self.shift_schedules.values() if s.get('shifts')])
        stats["无排班人数"] = stats["总人数"] - stats["有排班人数"]
        return stats

    def get_leave_type_stats(self):
        """获取请假类型统计"""
        stats = {}
        stats["请假类型数"] = len(self.leave_types)
        stats["总请假记录"] = len(self.leave_records)
        stats["本月请假"] = len([r for r in self.leave_records if r['date'].startswith(datetime.date.today().strftime("%Y-%m"))])
        return stats

    def get_monthly_leave_stats(self):
        """获取月度请假统计

        按照年休假规则处理：当年4月到次年3月为一个统计周期
        """
        stats = {}
        current_date = datetime.date.today()

        # 确定当前年休假年度
        if current_date.month >= 4:
            # 4-12月：属于当前年份的年休假年度
            leave_year = current_date.year
        else:
            # 1-3月：属于上一年的年休假年度
            leave_year = current_date.year - 1

        for month in range(1, 13):
            month_key = f"{month}月"
            count = 0

            for r in self.leave_records:
                date_str = r.get('date', '')
                if not date_str:
                    continue

                try:
                    parts = date_str.split('-')
                    record_year = int(parts[0])
                    record_month = int(parts[1])

                    # 检查是否属于当年年休假周期
                    if record_month >= 4 and record_year == leave_year:
                        # 当年4-12月属于当年年休假周期
                        if record_month == month:
                            count += 1
                    elif record_month <= 3 and record_year == leave_year + 1:
                        # 次年1-3月属于当年年休假周期
                        if record_month == month:
                            count += 1
                    elif record_year == current_date.year and record_month == month:
                        # 非年休假的其他假期，按自然年统计
                        if not self._is_annual_leave(r.get('type', '')):
                            count += 1
                except Exception:
                    continue

            stats[month_key] = count

        return stats

    def get_attendance_stats(self):
        """获取出勤率统计"""
        stats = {}
        total_people = len(self.shift_schedules)
        if total_people == 0:
            return {"总人数": 0, "平均出勤率": "0%"}

        # 计算每个人的出勤率
        attendance_rates = []
        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                total_days = len(schedule['shifts'])
                if total_days > 0:
                    # 计算非休息日天数
                    work_days = len([d for d, shift in schedule['shifts'].items()
                                   if shift != '休息'])
                    attendance_rate = (work_days / total_days) * 100
                    attendance_rates.append(attendance_rate)

        if attendance_rates:
            avg_attendance = sum(attendance_rates) / len(attendance_rates)
            stats["平均出勤率"] = f"{avg_attendance:.1f}%"
        else:
            stats["平均出勤率"] = "0%"

        stats["总人数"] = total_people
        return stats

    def get_leave_rate_stats(self):
        """获取请假率统计"""
        stats = {}
        total_people = len(self.shift_schedules)
        if total_people == 0:
            return {"总人数": 0, "平均请假率": "0%"}

        # 计算每个人的请假率
        leave_rates = []
        for person_name in self.shift_schedules.keys():
            person_leaves = len([r for r in self.leave_records if r['plan_name'] == person_name])
            if person_leaves > 0:
                leave_rates.append(person_leaves)

        if leave_rates:
            avg_leave = sum(leave_rates) / len(leave_rates)
            stats["平均请假率"] = f"{avg_leave:.1f}天"
        else:
            stats["平均请假率"] = "0天"

        stats["总人数"] = total_people
        return stats

    def refresh_statistics(self, notebook):
        """刷新统计数据"""
        try:
            # 刷新所有统计树视图
            if hasattr(self, 'shift_distribution_tree'):
                self.refresh_shift_distribution()

            if hasattr(self, 'leave_stats_tree'):
                self.refresh_leave_stats()

            if hasattr(self, 'person_stats_tree'):
                self.refresh_person_stats()

            self.update_status("统计数据已刷新")

        except Exception as e:
            self.error_handler.show_error("刷新失败", f"刷新统计数据时发生错误", str(e))

    def refresh_shift_distribution(self):
        """刷新排班分布统计"""
        # 清空现有数据
        for item in self.shift_distribution_tree.get_children():
            self.shift_distribution_tree.delete(item)

        # 统计每个人员的班次分布
        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                shift_counts = {}
                total_shifts = len(schedule['shifts'])

                for shift_type in schedule['shifts'].values():
                    shift_counts[shift_type] = shift_counts.get(shift_type, 0) + 1

                # 添加到树视图
                for shift_type, count in shift_counts.items():
                    percentage = (count / total_shifts) * 100
                    self.shift_distribution_tree.insert("", tk.END, values=(
                        person_name, shift_type, count, f"{percentage:.1f}%"
                    ))

    def refresh_leave_stats(self):
        """刷新请假统计"""
        # 清空现有数据
        for item in self.leave_stats_tree.get_children():
            self.leave_stats_tree.delete(item)

        # 获取最近30天的请假记录
        today = datetime.date.today()
        thirty_days_ago = today - datetime.timedelta(days=30)

        recent_leaves = [r for r in self.leave_records
                         if datetime.datetime.strptime(r['date'], '%Y-%m-%d').date() >= thirty_days_ago]

        # 按日期排序
        recent_leaves.sort(key=lambda x: x['date'], reverse=True)

        # 添加到树视图
        for leave in recent_leaves[:20]:  # 只显示最近20条
            leave_date = datetime.datetime.strptime(leave['date'], '%Y-%m-%d').date()
            days_ago = (today - leave_date).days

            self.leave_stats_tree.insert("", tk.END, values=(
                leave['plan_name'], leave['date'], leave['type'], f"{days_ago}天前"
            ))

    def refresh_person_stats(self):
        """刷新人员统计"""
        # 清空现有数据
        for item in self.person_stats_tree.get_children():
            self.person_stats_tree.delete(item)

        # 统计每个人的详细数据
        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                total_shifts = len(schedule['shifts'])
                leave_days = len([r for r in self.leave_records if r['plan_name'] == person_name])

                # 计算出勤率
                work_days = len([d for d, shift in schedule['shifts'].items() if shift != '休息'])
                attendance_rate = (work_days / total_shifts) * 100 if total_shifts > 0 else 0

                self.person_stats_tree.insert("", tk.END, values=(
                    person_name, total_shifts, leave_days, f"{attendance_rate:.1f}%"
                ))

    def export_statistics_report(self):
        """导出统计报表"""
        try:
            # 生成报表数据
            report_data = {
                "生成时间": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "班次类型统计": self.get_shift_type_stats(),
                "人员统计": self.get_person_stats(),
                "请假类型统计": self.get_leave_type_stats(),
                "月度请假统计": self.get_monthly_leave_stats(),
                "出勤率统计": self.get_attendance_stats(),
                "请假率统计": self.get_leave_rate_stats()
            }

            # 选择保存位置
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile=f"statistics_report_{datetime.date.today().strftime('%Y%m%d')}.json"
            )

            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, ensure_ascii=False, indent=2)

                self.error_handler.show_info("导出成功", f"统计报表已保存到:\n{filename}")
                self.update_status(f"统计报表导出成功: {os.path.basename(filename)}")

        except Exception as e:
            self.error_handler.show_error("导出失败", f"导出统计报表时发生错误", str(e))

    def show_report_generator(self):
        """显示高级报表生成器"""
        report_dialog = tk.Toplevel(self.root)
        report_dialog.title("高级报表生成器")
        report_dialog.geometry("700x500")
        report_dialog.resizable(False, False)

        # 设置对话框居中
        report_dialog.transient(self.root)
        report_dialog.grab_set()

        # 创建报表界面
        main_frame = ttk.Frame(report_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = ttk.Label(main_frame, text="高级报表生成器",
                             font=('Microsoft YaHei UI', 16, 'bold'),
                             foreground=self.colors['primary'])
        title_label.pack(pady=(0, 20))

        # 报表类型选择
        type_frame = ttk.Frame(main_frame)
        type_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(type_frame, text="报表类型:", font=('Microsoft YaHei UI', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))

        self.report_type_var = tk.StringVar(value="monthly_schedule")
        report_types = [
            ("月度排班表", "monthly_schedule"),
            ("人员考勤汇总", "person_attendance"),
            ("部门排班统计", "department_stats"),
            ("请假分析报告", "leave_analysis"),
            ("年度统计报告", "annual_report"),
            ("自定义报表", "custom_report")
        ]

        for display_name, value in report_types:
            ttk.Radiobutton(type_frame, text=display_name, variable=self.report_type_var,
                           value=value).pack(anchor=tk.W, pady=2)

        # 报表参数设置
        params_frame = ttk.Frame(main_frame)
        params_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(params_frame, text="报表参数:", font=('Microsoft YaHei UI', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))

        # 时间范围选择
        time_frame = ttk.Frame(params_frame)
        time_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(time_frame, text="时间范围:").pack(side=tk.LEFT, padx=(0, 10))
        self.time_range_var = tk.StringVar(value="current_month")
        time_combo = ttk.Combobox(time_frame, textvariable=self.time_range_var,
                                values=["current_month", "last_month", "current_quarter", "current_year", "custom_range"],
                                state="readonly", width=15)
        time_combo.pack(side=tk.LEFT)

        # 自定义时间范围
        custom_frame = ttk.Frame(params_frame)
        custom_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(custom_frame, text="开始日期:").pack(side=tk.LEFT, padx=(0, 5))
        self.start_date_var = tk.StringVar()
        start_entry = ttk.Entry(custom_frame, textvariable=self.start_date_var, width=12)
        start_entry.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(custom_frame, text="结束日期:").pack(side=tk.LEFT, padx=(0, 5))
        self.end_date_var = tk.StringVar()
        end_entry = ttk.Entry(custom_frame, textvariable=self.end_date_var, width=12)
        end_entry.pack(side=tk.LEFT)

        # 输出格式选择
        format_frame = ttk.Frame(params_frame)
        format_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(format_frame, text="输出格式:").pack(side=tk.LEFT, padx=(0, 10))
        self.output_format_var = tk.StringVar(value="excel")
        format_combo = ttk.Combobox(format_frame, textvariable=self.output_format_var,
                                   values=["excel", "pdf", "html", "csv"], state="readonly", width=10)
        format_combo.pack(side=tk.LEFT)

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        def generate_report():
            try:
                report_type = self.report_type_var.get()
                time_range = self.time_range_var.get()
                output_format = self.output_format_var.get()

                # 验证参数
                if time_range == "custom_range":
                    start_date = self.start_date_var.get().strip()
                    end_date = self.end_date_var.get().strip()

                    if not start_date or not end_date:
                        self.error_handler.show_warning("参数错误", "请选择自定义时间范围")
                        return

                    is_valid, error_msg = DataValidator.validate_date_format(start_date)
                    if not is_valid:
                        self.error_handler.show_warning("参数错误", f"开始日期: {error_msg}")
                        return

                    is_valid, error_msg = DataValidator.validate_date_format(end_date)
                    if not is_valid:
                        self.error_handler.show_warning("参数错误", f"结束日期: {error_msg}")
                        return

                # 生成报表
                self.generate_advanced_report(report_type, time_range, output_format)
                report_dialog.destroy()

            except Exception as e:
                self.error_handler.show_error("生成失败", f"生成报表时发生错误", str(e))

        ttk.Button(button_frame, text="生成报表", command=generate_report,
                   style='Primary.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="预览", command=lambda: self.preview_report(),
                   style='Info.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="取消", command=report_dialog.destroy,
                   style='Danger.TButton').pack(side=tk.RIGHT)

    def generate_advanced_report(self, report_type, time_range, output_format):
        """生成高级报表"""
        try:
            # 计算时间范围
            start_date, end_date = self.calculate_date_range(time_range)

            # 根据报表类型生成数据
            if report_type == "monthly_schedule":
                report_data = self.generate_monthly_schedule_report(start_date, end_date)
                filename = f"月度排班表_{start_date.strftime('%Y%m%d')}"
            elif report_type == "person_attendance":
                report_data = self.generate_person_attendance_report(start_date, end_date)
                filename = f"人员考勤汇总_{start_date.strftime('%Y%m%d')}"
            elif report_type == "department_stats":
                report_data = self.generate_department_stats_report(start_date, end_date)
                filename = f"部门排班统计_{start_date.strftime('%Y%m%d')}"
            elif report_type == "leave_analysis":
                report_data = self.generate_leave_analysis_report(start_date, end_date)
                filename = f"请假分析报告_{start_date.strftime('%Y%m%d')}"
            elif report_type == "annual_report":
                report_data = self.generate_annual_report(start_date, end_date)
                filename = f"年度统计报告_{start_date.strftime('%Y%m%d')}"
            else:
                report_data = self.generate_custom_report(start_date, end_date)
                filename = f"自定义报表_{start_date.strftime('%Y%m%d')}"

            # 根据输出格式保存报表
            if output_format == "excel":
                self.save_excel_report(report_data, filename)
            elif output_format == "pdf":
                self.save_pdf_report(report_data, filename)
            elif output_format == "html":
                self.save_html_report(report_data, filename)
            else:  # csv
                self.save_csv_report(report_data, filename)

        except Exception as e:
            self.error_handler.show_error("报表生成失败", f"生成报表时发生错误", str(e))

    def calculate_date_range(self, time_range):
        """计算时间范围"""
        today = datetime.date.today()

        if time_range == "current_month":
            start_date = today.replace(day=1)
            end_date = today
        elif time_range == "last_month":
            if today.month == 1:
                start_date = today.replace(year=today.year-1, month=12, day=1)
                end_date = today.replace(year=today.year-1, month=12, day=31)
            else:
                start_date = today.replace(month=today.month-1, day=1)
                end_date = today.replace(day=1) - datetime.timedelta(days=1)
        elif time_range == "current_quarter":
            quarter = (today.month - 1) // 3 + 1
            start_date = today.replace(month=(quarter-1)*3+1, day=1)
            end_date = today
        elif time_range == "current_year":
            start_date = today.replace(month=1, day=1)
            end_date = today
        else:  # custom_range
            start_date = datetime.datetime.strptime(self.start_date_var.get(), '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(self.end_date_var.get(), '%Y-%m-%d').date()

        return start_date, end_date

    def generate_monthly_schedule_report(self, start_date, end_date):
        """生成月度排班表报表"""
        report_data = {
            "title": f"月度排班表 ({start_date.strftime('%Y年%m月')})",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "data": []
        }

        # 为每个人员生成排班数据
        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                person_data = {
                    "name": person_name,
                    "schedules": []
                }

                # 获取时间范围内的排班数据
                current_date = start_date
                while current_date <= end_date:
                    date_str = current_date.strftime('%Y-%m-%d')
                    shift_type = schedule['shifts'].get(date_str, '无')

                    person_data["schedules"].append({
                        "date": date_str,
                        "weekday": current_date.strftime('%A'),
                        "shift": shift_type
                    })

                    current_date += datetime.timedelta(days=1)

                report_data["data"].append(person_data)

        return report_data

    def generate_person_attendance_report(self, start_date, end_date):
        """生成人员考勤汇总报表"""
        report_data = {
            "title": f"人员考勤汇总 ({start_date.strftime('%Y年%m月')})",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "data": []
        }

        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                # 计算统计数据
                total_days = len([d for d, s in schedule['shifts'].items()
                               if start_date <= datetime.datetime.strptime(d, '%Y-%m-%d').date() <= end_date])
                work_days = len([d for d, s in schedule['shifts'].items()
                               if start_date <= datetime.datetime.strptime(d, '%Y-%m-%d').date() <= end_date and s != '休息'])
                leave_days = len([r for r in self.leave_records
                               if r['plan_name'] == person_name and
                               start_date <= datetime.datetime.strptime(r['date'], '%Y-%m-%d').date() <= end_date])

                attendance_rate = (work_days / total_days * 100) if total_days > 0 else 0

                person_data = {
                    "name": person_name,
                    "total_days": total_days,
                    "work_days": work_days,
                    "leave_days": leave_days,
                    "attendance_rate": f"{attendance_rate:.1f}%"
                }

                report_data["data"].append(person_data)

        return report_data

    def generate_department_stats_report(self, start_date, end_date):
        """生成部门排班统计报表"""
        report_data = {
            "title": f"部门排班统计 ({start_date.strftime('%Y年%m月')})",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "summary": {},
            "data": []
        }

        # 统计班次类型分布
        shift_distribution = {}
        total_shifts = 0

        for schedule in self.shift_schedules.values():
            if schedule.get('shifts'):
                for date, shift_type in schedule['shifts'].items():
                    if start_date <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= end_date:
                        shift_distribution[shift_type] = shift_distribution.get(shift_type, 0) + 1
                        total_shifts += 1

        # 生成汇总数据
        for shift_type, count in shift_distribution.items():
            percentage = (count / total_shifts * 100) if total_shifts > 0 else 0
            report_data["data"].append({
                "shift_type": shift_type,
                "count": count,
                "percentage": f"{percentage:.1f}%"
            })

        report_data["summary"] = {
            "total_shifts": total_shifts,
            "shift_types": len(shift_distribution),
            "total_people": len(self.shift_schedules)
        }

        return report_data

    def generate_leave_analysis_report(self, start_date, end_date):
        """生成请假分析报表"""
        report_data = {
            "title": f"请假分析报告 ({start_date.strftime('%Y年%m月')})",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "summary": {},
            "data": []
        }

        # 统计请假类型分布
        leave_distribution = {}
        total_leaves = 0

        for leave_record in self.leave_records:
            leave_date = datetime.datetime.strptime(leave_record['date'], '%Y-%m-%d').date()
            if start_date <= leave_date <= end_date:
                leave_type = leave_record['type']
                leave_distribution[leave_type] = leave_distribution.get(leave_type, 0) + 1
                total_leaves += 1

        # 生成汇总数据
        for leave_type, count in leave_distribution.items():
            percentage = (count / total_leaves * 100) if total_leaves > 0 else 0
            report_data["data"].append({
                "leave_type": leave_type,
                "count": count,
                "percentage": f"{percentage:.1f}%"
            })

        report_data["summary"] = {
            "total_leaves": total_leaves,
            "leave_types": len(leave_distribution),
            "avg_leaves_per_person": f"{total_leaves / len(self.shift_schedules):.1f}" if self.shift_schedules else "0"
        }

        return report_data

    def generate_annual_report(self, start_date, end_date):
        """生成年度统计报告"""
        report_data = {
            "title": f"年度统计报告 ({start_date.year}年)",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "sections": {}
        }

        # 添加各个统计部分
        report_data["sections"]["shift_types"] = self.get_shift_type_stats()
        report_data["sections"]["person_stats"] = self.get_person_stats()
        report_data["sections"]["leave_stats"] = self.get_leave_type_stats()
        report_data["sections"]["attendance_stats"] = self.get_attendance_stats()

        return report_data

    def generate_custom_report(self, start_date, end_date):
        """生成自定义报表"""
        report_data = {
            "title": f"自定义报表 ({start_date.strftime('%Y年%m月')})",
            "period": f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
            "generated_at": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "shift_schedules": [],
            "leave_records": []
        }

        # 添加排班数据
        for person_name, schedule in self.shift_schedules.items():
            if schedule.get('shifts'):
                for date, shift_type in schedule['shifts'].items():
                    schedule_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                    if start_date <= schedule_date <= end_date:
                        report_data["shift_schedules"].append({
                            "person": person_name,
                            "date": date,
                            "shift": shift_type
                        })

        # 添加请假数据
        for leave_record in self.leave_records:
            leave_date = datetime.datetime.strptime(leave_record['date'], '%Y-%m-%d').date()
            if start_date <= leave_date <= end_date:
                report_data["leave_records"].append(leave_record)

        return report_data

    def save_excel_report(self, report_data, filename):
        """保存Excel格式报表"""
        try:
            if pd is None:
                self.error_handler.show_warning("功能不可用", "Excel报表生成需要安装pandas库")
                return

            # 选择保存位置
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"{filename}.xlsx"
            )

            if filepath:
                # 这里简化处理，实际应该根据不同报表类型生成不同的Excel格式
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write("报表数据 (JSON格式):\n\n")
                    json.dump(report_data, f, ensure_ascii=False, indent=2)

                self.error_handler.show_info("保存成功", f"Excel报表已保存到:\n{filepath}")
                self.update_status(f"Excel报表保存成功: {os.path.basename(filepath)}")

        except Exception as e:
            self.error_handler.show_error("保存失败", f"保存Excel报表时发生错误", str(e))

    def save_pdf_report(self, report_data, filename):
        """保存PDF格式报表"""
        try:
            # 选择保存位置
            filepath = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"{filename}.pdf"
            )

            if filepath:
                # 这里简化处理，实际应该使用PDF生成库
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(f"{report_data['title']}\n\n")
                    f.write(f"时间范围: {report_data['period']}\n")
                    f.write(f"生成时间: {report_data['generated_at']}\n\n")
                    f.write("注意: PDF格式需要安装相应的PDF生成库\n")

                self.error_handler.show_info("保存成功", f"PDF报表已保存到:\n{filepath}")
                self.update_status(f"PDF报表保存成功: {os.path.basename(filepath)}")

        except Exception as e:
            self.error_handler.show_error("保存失败", f"保存PDF报表时发生错误", str(e))

    def save_html_report(self, report_data, filename):
        """保存HTML格式报表"""
        try:
            # 选择保存位置
            filepath = filedialog.asksaveasfilename(
                defaultextension=".html",
                filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
                initialfile=f"{filename}.html"
            )

            if filepath:
                # 生成HTML内容
                html_content = self.generate_html_content(report_data)
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(html_content)

                self.error_handler.show_info("保存成功", f"HTML报表已保存到:\n{filepath}")
                self.update_status(f"HTML报表保存成功: {os.path.basename(filepath)}")

        except Exception as e:
            self.error_handler.show_error("保存失败", f"保存HTML报表时发生错误", str(e))

    def save_csv_report(self, report_data, filename):
        """保存CSV格式报表"""
        try:
            # 选择保存位置
            filepath = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"{filename}.csv"
            )

            if filepath:
                # 生成CSV内容
                csv_content = self.generate_csv_content(report_data)
                with open(filepath, 'w', encoding='utf-8-sig') as f:
                    f.write(csv_content)

                self.error_handler.show_info("保存成功", f"CSV报表已保存到:\n{filepath}")
                self.update_status(f"CSV报表保存成功: {os.path.basename(filepath)}")

        except Exception as e:
            self.error_handler.show_error("保存失败", f"保存CSV报表时发生错误", str(e))

    def generate_html_content(self, report_data):
        """生成HTML内容"""
        html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_data['title']}</title>
    <style>
        body {{ font-family: 'Microsoft YaHei UI', Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #2196F3; border-bottom: 2px solid #2196F3; padding-bottom: 10px; }}
        .info {{ background-color: #f5f5f5; padding: 10px; border-radius: 5px; margin: 20px 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #2196F3; color: white; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
    </style>
</head>
<body>
    <h1>{report_data['title']}</h1>
    <div class="info">
        <strong>时间范围:</strong> {report_data['period']}<br>
        <strong>生成时间:</strong> {report_data['generated_at']}
    </div>
    <div class="content">
        <p>报表数据已生成，包含详细的统计信息。</p>
        <p>数据格式: JSON</p>
    </div>
</body>
</html>"""
        return html_template

    def generate_csv_content(self, report_data):
        """生成CSV内容"""
        csv_lines = [f"{report_data['title']}"]
        csv_lines.append(f"时间范围,{report_data['period']}")
        csv_lines.append(f"生成时间,{report_data['generated_at']}")
        csv_lines.append("")

        # 根据报表类型添加数据
        if "data" in report_data:
            csv_lines.append("类型,数值")
            for item in report_data["data"]:
                if isinstance(item, dict):
                    for key, value in item.items():
                        csv_lines.append(f"{key},{value}")
                else:
                    csv_lines.append(f"{item}")

        return "\n".join(csv_lines)

    def preview_report(self):
        """预览报表"""
        # 临时实现，显示预览对话框
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title("报表预览")
        preview_dialog.geometry("600x400")

        preview_dialog.transient(self.root)
        preview_dialog.grab_set()

        main_frame = ttk.Frame(preview_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="报表预览功能正在开发中...",
                  font=('Microsoft YaHei UI', 12)).pack(expand=True)

        ttk.Button(main_frame, text="关闭", command=preview_dialog.destroy,
                   style='Danger.TButton').pack(pady=(20, 0))

    def backup_data(self):
        """备份数据"""
        try:
            # 生成备份文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"shift_data_backup_{timestamp}.json"
            backup_path = os.path.join(self.backup_directory, backup_filename)

            # 读取当前数据
            data = {
                "shift_types": self.shift_types,
                "schedules": self.shift_schedules,
                "leave_types": self.leave_types,
                "leave_records": self.leave_records,
                "leave_quotas": self.leave_quotas,
                "holidays": self.holidays,
                "backup_timestamp": timestamp,
                "backup_version": "2.0"
            }

            # 写入备份文件
            with open(backup_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            # 清理旧备份
            self.cleanup_old_backups()

            # 更新最后备份时间
            self.last_backup_time = datetime.datetime.now()

            self.error_handler.show_info("备份成功", f"数据已备份到:\n{backup_path}")
            self.update_status(f"备份成功: {backup_filename}")

        except Exception as e:
            self.error_handler.show_error("备份失败", f"备份数据时发生错误", str(e))

    def cleanup_old_backups(self):
        """清理旧备份文件"""
        try:
            # 获取所有备份文件
            backup_files = []
            for filename in os.listdir(self.backup_directory):
                if filename.startswith("shift_data_backup_") and filename.endswith(".json"):
                    filepath = os.path.join(self.backup_directory, filename)
                    backup_files.append((filepath, os.path.getmtime(filepath)))

            # 按修改时间排序
            backup_files.sort(key=lambda x: x[1], reverse=True)

            # 删除超过限制的备份文件
            limit = self.backup_count_limit.get()
            for filepath, _ in backup_files[limit:]:
                try:
                    os.remove(filepath)
                    self.update_status(f"已删除旧备份: {os.path.basename(filepath)}")
                except Exception as e:
                    self.error_handler.show_error("删除备份失败", f"删除备份文件失败", str(e))

        except Exception as e:
            self.error_handler.show_error("清理备份失败", f"清理旧备份时发生错误", str(e))

    def check_auto_backup(self):
        """检查是否需要自动备份"""
        if not self.backup_enabled.get():
            return

        now = datetime.datetime.now()
        
        # 延迟首次备份检查，避免启动时立即执行
        if not hasattr(self, '_first_backup_check_performed'):
            self._first_backup_check_performed = True
            # 使用定时器延迟首次备份检查，避免影响启动速度
            self.root.after(5000, self._perform_backup_check)  # 5秒后执行备份检查
            return

    def _perform_backup_check(self):
        """执行备份检查的内部方法"""
        now = datetime.datetime.now()
        interval_hours = self.backup_interval.get() * 24  # 转换为小时

        if self.last_backup_time is None:
            # 首次备份
            self.backup_data()
        else:
            # 检查是否达到备份间隔
            time_diff = now - self.last_backup_time
            if time_diff.total_seconds() >= interval_hours * 3600:
                self.backup_data()

    def show_backup_settings(self):
        """显示备份设置对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("备份设置")
        dialog.geometry("400x300")
        dialog.resizable(False, False)

        # 设置对话框居中
        dialog.transient(self.root)
        dialog.grab_set()

        # 创建设置界面
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 自动备份开关
        auto_frame = ttk.Frame(main_frame)
        auto_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Checkbutton(auto_frame, text="启用自动备份", variable=self.backup_enabled).pack(side=tk.LEFT)

        # 备份间隔设置
        interval_frame = ttk.Frame(main_frame)
        interval_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(interval_frame, text="备份间隔:").pack(side=tk.LEFT, padx=(0, 10))
        interval_combo = ttk.Combobox(interval_frame, textvariable=self.backup_interval,
                                    values=[1, 3, 7, 30], state="readonly", width=10)
        interval_combo.pack(side=tk.LEFT)
        ttk.Label(interval_frame, text="天").pack(side=tk.LEFT, padx=(5, 0))

        # 备份数量限制
        limit_frame = ttk.Frame(main_frame)
        limit_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(limit_frame, text="保留备份数量:").pack(side=tk.LEFT, padx=(0, 10))
        limit_spinbox = ttk.Spinbox(limit_frame, from_=1, to=100, textvariable=self.backup_count_limit, width=10)
        limit_spinbox.pack(side=tk.LEFT)
        ttk.Label(limit_frame, text="个").pack(side=tk.LEFT, padx=(5, 0))

        # 备份目录信息
        dir_frame = ttk.Frame(main_frame)
        dir_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(dir_frame, text="备份目录:").pack(anchor=tk.W)
        dir_label = ttk.Label(dir_frame, text=self.backup_directory, font=('Microsoft YaHei UI', 8))
        dir_label.pack(anchor=tk.W, pady=(5, 0))

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="立即备份", command=lambda: [self.backup_data(), dialog.destroy()],
                   style='Primary.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="打开备份目录", command=self.open_backup_directory,
                   style='Info.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="取消", command=dialog.destroy,
                   style='Danger.TButton').pack(side=tk.RIGHT)

    def open_backup_directory(self):
        """打开备份目录"""
        try:
            if os.path.exists(self.backup_directory):
                if os.name == 'nt':  # Windows
                    os.startfile(self.backup_directory)
                elif os.name == 'posix':  # macOS/Linux
                    import subprocess
                    subprocess.run(['open', self.backup_directory] if sys.platform == 'darwin' else ['xdg-open', self.backup_directory])
        except Exception as e:
            self.error_handler.show_error("打开目录失败", f"无法打开备份目录", str(e))

    def restore_data(self):
        """恢复数据"""
        try:
            # 获取备份文件列表
            backup_files = []
            for filename in os.listdir(self.backup_directory):
                if filename.startswith("shift_data_backup_") and filename.endswith(".json"):
                    filepath = os.path.join(self.backup_directory, filename)
                    backup_files.append((filepath, os.path.getmtime(filepath)))

            if not backup_files:
                self.error_handler.show_warning("没有备份文件", "没有找到可用的备份文件")
                return

            # 按修改时间排序（最新的在前）
            backup_files.sort(key=lambda x: x[1], reverse=True)

            # 创建恢复对话框
            restore_dialog = tk.Toplevel(self.root)
            restore_dialog.title("数据恢复")
            restore_dialog.geometry("600x500")
            restore_dialog.resizable(False, False)

            # 设置对话框居中
            restore_dialog.transient(self.root)
            restore_dialog.grab_set()

            # 创建恢复界面
            main_frame = ttk.Frame(restore_dialog, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # 说明文字
            info_label = ttk.Label(main_frame, text="选择要恢复的备份文件：", font=('Microsoft YaHei UI', 12, 'bold'))
            info_label.pack(anchor=tk.W, pady=(0, 10))

            warning_label = ttk.Label(main_frame, text="注意：恢复数据将覆盖当前所有数据，请谨慎操作！",
                                  font=('Microsoft YaHei UI', 10), foreground=self.colors['danger'])
            warning_label.pack(anchor=tk.W, pady=(0, 10))

            # 备份文件列表
            list_frame = ttk.Frame(main_frame)
            list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

            # 创建滚动条
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # 备份列表
            backup_tree = ttk.Treeview(list_frame, columns=("filename", "date", "size"),
                                       show="headings", yscrollcommand=scrollbar.set, height=15)
            backup_tree.heading("filename", text="文件名")
            backup_tree.heading("date", text="备份时间")
            backup_tree.heading("size", text="文件大小")

            backup_tree.column("filename", width=200)
            backup_tree.column("date", width=150)
            backup_tree.column("size", width=100)

            backup_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=backup_tree.yview)

            # 添加备份文件到列表
            for filepath, mtime in backup_files:
                filename = os.path.basename(filepath)
                backup_date = datetime.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                file_size = f"{os.path.getsize(filepath) / 1024:.1f} KB"

                backup_tree.insert("", tk.END, values=(filename, backup_date, file_size))

            # 绑定双击事件
            backup_tree.bind('<Double-1>', lambda e: self.restore_from_file(backup_tree, restore_dialog))

            # 按钮区域
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X)

            def do_restore():
                selected = backup_tree.selection()
                if selected:
                    if self.error_handler.ask_confirmation("确认恢复", "恢复数据将覆盖当前所有数据，是否继续？"):
                        self.restore_from_file(backup_tree, restore_dialog)
                else:
                    self.error_handler.show_warning("提示", "请先选择要恢复的备份文件")

            ttk.Button(button_frame, text="恢复", command=do_restore,
                       style='Primary.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="打开备份目录", command=lambda: [self.open_backup_directory(), restore_dialog.destroy()],
                       style='Info.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="取消", command=restore_dialog.destroy,
                       style='Danger.TButton').pack(side=tk.RIGHT)

            # 默认选择最新的备份文件
            if backup_tree.get_children():
                backup_tree.selection_set(backup_tree.get_children()[0])
                backup_tree.focus(backup_tree.get_children()[0])

        except Exception as e:
            self.error_handler.show_error("恢复失败", f"准备恢复时发生错误", str(e))

    def restore_from_file(self, backup_tree, dialog):
        """从指定文件恢复数据"""
        try:
            selected = backup_tree.selection()
            if not selected:
                return

            item = selected[0]
            filename = backup_tree.item(item)['values'][0]
            backup_path = os.path.join(self.backup_directory, filename)

            # 读取备份文件
            with open(backup_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)

            # 验证备份数据格式
            required_keys = ['shift_types', 'schedules', 'leave_types', 'leave_records', 'leave_quotas', 'holidays']
            missing_keys = [key for key in required_keys if key not in backup_data]

            if missing_keys:
                self.error_handler.show_error("备份格式错误", f"备份文件缺少必要字段: {', '.join(missing_keys)}")
                return

            # 备份当前数据（安全措施）
            self.backup_data()

            # 恢复数据
            self.shift_types = backup_data['shift_types']
            self.shift_schedules = backup_data['schedules']
            self.leave_types = backup_data['leave_types']
            self.leave_records = backup_data['leave_records']
            self.leave_quotas = backup_data['leave_quotas']
            self.holidays = backup_data['holidays']

            # 保存恢复后的数据
            self.save_data()

            # 更新界面
            self.update_shift_type_tree()
            self.update_schedule_tree()
            self.update_leave_tree()
            self.update_holiday_tree()
            self.update_calendar()

            # 关闭对话框
            dialog.destroy()

            # 显示成功消息
            backup_time = backup_data.get('backup_timestamp', '未知')
            self.error_handler.show_info("恢复成功", f"数据已从备份恢复\n备份时间: {backup_time}")
            self.update_status(f"数据恢复成功: {filename}")

        except json.JSONDecodeError:
            self.error_handler.show_error("恢复失败", "备份文件格式错误，无法解析")
        except Exception as e:
            self.error_handler.show_error("恢复失败", f"恢复数据时发生错误", str(e))

    def import_data_from_json(self):
        """从JSON文件导入数据"""
        try:
            # 打开文件选择对话框
            file_path = filedialog.askopenfilename(
                title="选择要导入的JSON文件",
                initialdir=os.path.dirname(os.path.abspath(__file__)),
                filetypes=[
                    ("JSON文件", "*.json"),
                    ("所有文件", "*.*")
                ],
                parent=self.root
            )

            if not file_path:
                return

            # 确认导入操作
            confirm = messagebox.askyesno(
                "确认导入",
                f"准备从文件导入数据：\n{os.path.basename(file_path)}\n\n"
                "此操作将合并现有数据，是否继续？\n"
                "（建议先备份当前数据）",
                icon=messagebox.WARNING
            )

            if not confirm:
                return

            # 读取选定的JSON文件
            with open(file_path, 'r', encoding='utf-8') as f:
                import_data = json.load(f)

            # 验证导入数据格式
            if not isinstance(import_data, dict):
                raise ValueError("JSON文件格式错误：根对象必须是字典")

            # 创建导入预览对话框
            preview_dialog = tk.Toplevel(self.root)
            preview_dialog.title("导入预览")
            preview_dialog.geometry("500x400")
            preview_dialog.transient(self.root)
            preview_dialog.grab_set()

            # 预览框架
            preview_frame = ttk.Frame(preview_dialog, padding=15)
            preview_frame.pack(fill=tk.BOTH, expand=True)

            # 显示文件信息
            ttk.Label(preview_frame, text=f"文件路径：{file_path}",
                     font=('Microsoft YaHei UI', 9)).pack(anchor=tk.W, pady=(0, 10))

            # 分析导入内容
            import_summary = []

            # 检查班次类型
            if 'shift_types' in import_data:
                shift_count = len(import_data['shift_types'])
                import_summary.append(f"班次类型：{shift_count} 个")

            # 检查排班计划
            if 'schedules' in import_data:
                schedule_count = len(import_data['schedules'])
                import_summary.append(f"排班计划：{schedule_count} 个")

            # 检查请假类型
            if 'leave_types' in import_data:
                leave_types_count = len(import_data['leave_types'])
                import_summary.append(f"请假类型：{leave_types_count} 个")

            # 检查请假记录
            if 'leave_records' in import_data:
                leave_records_count = len(import_data['leave_records'])
                import_summary.append(f"请假记录：{leave_records_count} 条")

            # 检查配额信息
            if 'leave_quotas' in import_data:
                quota_count = len(import_data['leave_quotas'])
                import_summary.append(f"年度配额：{quota_count} 人员")

            # 检查节假日
            if 'holidays' in import_data:
                holiday_count = sum(len(h) for h in import_data['holidays'].values())
                import_summary.append(f"节假日：{holiday_count} 个")

            # 显示导入摘要
            ttk.Label(preview_frame, text="导入内容摘要：",
                     font=('Microsoft YaHei UI', 10, 'bold')).pack(anchor=tk.W, pady=(10, 5))

            summary_text = "\n".join(f"  • {item}" for item in import_summary)
            ttk.Label(preview_frame, text=summary_text,
                     font=('Microsoft YaHei UI', 9)).pack(anchor=tk.W, padx=(20, 0))

            # 导入选项
            options_frame = ttk.LabelFrame(preview_frame, text="导入选项", padding=10)
            options_frame.pack(fill=tk.X, pady=(15, 0))

            # 导入模式选择
            import_mode = tk.StringVar(value="merge")
            ttk.Radiobutton(options_frame, text="合并导入（保留现有数据，添加新数据）",
                          variable=import_mode, value="merge").pack(anchor=tk.W, pady=2)
            ttk.Radiobutton(options_frame, text="替换导入（删除现有数据，使用导入数据）",
                          variable=import_mode, value="replace").pack(anchor=tk.W, pady=2)

            # 数据类型选择
            data_types_frame = ttk.LabelFrame(preview_frame, text="选择要导入的数据类型", padding=10)
            data_types_frame.pack(fill=tk.X, pady=(10, 0))

            # 创建复选框变量
            include_vars = {}
            available_data = {
                'shift_types': '班次类型',
                'schedules': '排班计划',
                'leave_types': '请假类型',
                'leave_records': '请假记录',
                'leave_quotas': '年度配额',
                'holidays': '节假日'
            }

            for key, label in available_data.items():
                if key in import_data:
                    var = tk.BooleanVar(value=True)
                    include_vars[key] = var
                    ttk.Checkbutton(data_types_frame, text=label, variable=var).pack(anchor=tk.W, pady=1)

            # 按钮框架
            button_frame = ttk.Frame(preview_frame)
            button_frame.pack(fill=tk.X, pady=(15, 0))

            def do_import():
                try:
                    # 备份当前数据
                    self.backup_data()

                    # 根据导入模式处理数据
                    mode = import_mode.get()

                    # 合并或替换班次类型
                    if include_vars.get('shift_types', tk.BooleanVar()).get() and 'shift_types' in import_data:
                        if mode == "replace":
                            self.shift_types = import_data['shift_types']
                        else:  # merge
                            self.shift_types.update(import_data['shift_types'])

                    # 合并或替换排班计划
                    if include_vars.get('schedules', tk.BooleanVar()).get() and 'schedules' in import_data:
                        if mode == "replace":
                            self.shift_schedules = import_data['schedules']
                        else:  # merge
                            self.shift_schedules.update(import_data['schedules'])

                    # 合并或替换请假类型
                    if include_vars.get('leave_types', tk.BooleanVar()).get() and 'leave_types' in import_data:
                        if mode == "replace":
                            self.leave_types = import_data['leave_types']
                        else:  # merge
                            for item in import_data['leave_types']:
                                if item not in self.leave_types:
                                    self.leave_types.append(item)

                    # 合并或替换请假记录
                    if include_vars.get('leave_records', tk.BooleanVar()).get() and 'leave_records' in import_data:
                        if mode == "replace":
                            self.leave_records = import_data['leave_records']
                        else:  # merge
                            self.leave_records.extend(import_data['leave_records'])

                    # 合并或替换年度配额
                    if include_vars.get('leave_quotas', tk.BooleanVar()).get() and 'leave_quotas' in import_data:
                        if mode == "replace":
                            self.leave_quotas = import_data['leave_quotas']
                        else:  # merge
                            for person, quotas in import_data['leave_quotas'].items():
                                if person not in self.leave_quotas:
                                    self.leave_quotas[person] = {}
                                self.leave_quotas[person].update(quotas)

                    # 合并或替换节假日
                    if include_vars.get('holidays', tk.BooleanVar()).get() and 'holidays' in import_data:
                        if mode == "replace":
                            self.holidays = import_data['holidays']
                        else:  # merge
                            for year, holidays in import_data['holidays'].items():
                                if year not in self.holidays:
                                    self.holidays[year] = {}
                                self.holidays[year].update(holidays)

                    # 保存导入后的数据
                    self.save_data()

                    # 更新界面
                    self.update_shift_type_tree()
                    self.update_schedule_tree()
                    self.update_leave_tree()
                    self.update_holiday_tree()
                    self.update_calendar()

                    # 关闭预览对话框
                    preview_dialog.destroy()

                    # 显示成功消息
                    messagebox.showinfo("导入成功",
                                      f"数据已成功导入！\n文件：{os.path.basename(file_path)}\n模式：{'合并' if mode == 'merge' else '替换'}")
                    self.update_status(f"数据导入成功: {os.path.basename(file_path)}")

                except Exception as e:
                    messagebox.showerror("导入失败", f"导入数据时发生错误：{str(e)}")

            ttk.Button(button_frame, text="确认导入", command=do_import).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="取消", command=preview_dialog.destroy).pack(side=tk.RIGHT)

        except json.JSONDecodeError as e:
            self.error_handler.show_error("文件格式错误", "JSON文件格式不正确，请检查文件是否完整", str(e))
        except Exception as e:
            self.error_handler.show_error("导入失败", "导入数据时发生错误", str(e))

    def show_search_dialog(self):
        """显示全局搜索对话框"""
        search_dialog = tk.Toplevel(self.root)
        search_dialog.title("全局搜索")
        search_dialog.geometry("600x400")
        search_dialog.resizable(False, False)

        # 设置对话框居中
        search_dialog.transient(self.root)
        search_dialog.grab_set()

        # 创建搜索界面
        main_frame = ttk.Frame(search_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 搜索区域
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(search_frame, text="搜索内容:", font=('Microsoft YaHei UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 10))
        self.global_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.global_search_var, font=('Microsoft YaHei UI', 10))
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        search_entry.bind('<Return>', lambda e: self.perform_global_search())
        search_entry.focus()

        ttk.Button(search_frame, text="🔍 搜索", command=self.perform_global_search,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=(10, 0))

        # 搜索选项
        options_frame = ttk.Frame(main_frame)
        options_frame.pack(fill=tk.X, pady=(0, 20))

        self.search_in_shift_types = tk.BooleanVar(value=True)
        self.search_in_schedules = tk.BooleanVar(value=True)
        self.search_in_leaves = tk.BooleanVar(value=True)
        self.search_in_holidays = tk.BooleanVar(value=True)

        ttk.Checkbutton(options_frame, text="班次类型", variable=self.search_in_shift_types).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(options_frame, text="排班计划", variable=self.search_in_schedules).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(options_frame, text="请假记录", variable=self.search_in_leaves).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(options_frame, text="节假日", variable=self.search_in_holidays).pack(side=tk.LEFT, padx=(0, 10))

        # 搜索结果区域
        results_frame = ttk.Frame(main_frame)
        results_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(results_frame, text="搜索结果:", font=('Microsoft YaHei UI', 10, 'bold')).pack(anchor=tk.W, pady=(0, 5))

        # 创建结果列表
        result_scroll = ttk.Scrollbar(results_frame)
        result_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.search_results_tree = ttk.Treeview(results_frame, columns=("type", "name", "details"),
                                              show="headings", yscrollcommand=result_scroll.set, height=10)
        self.search_results_tree.heading("type", text="类型")
        self.search_results_tree.heading("name", text="名称")
        self.search_results_tree.heading("details", text="详细信息")

        self.search_results_tree.column("type", width=80)
        self.search_results_tree.column("name", width=150)
        self.search_results_tree.column("details", width=300)

        self.search_results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        result_scroll.config(command=self.search_results_tree.yview)

        # 绑定双击事件
        self.search_results_tree.bind('<Double-1>', self.on_search_result_double_click)

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        ttk.Button(button_frame, text="关闭", command=search_dialog.destroy,
                   style='Danger.TButton').pack(side=tk.RIGHT)

    def perform_global_search(self):
        """执行全局搜索"""
        search_term = self.global_search_var.get().lower().strip()
        if not search_term:
            messagebox.showwarning("提示", "请输入搜索内容")
            return

        # 清空之前的结果
        for item in self.search_results_tree.get_children():
            self.search_results_tree.delete(item)

        results_count = 0

        # 搜索班次类型
        if self.search_in_shift_types.get():
            for shift_name, shift_data in self.shift_types.items():
                if search_term in shift_name.lower():
                    self.search_results_tree.insert("", tk.END, values=(
                        "班次类型",
                        shift_name,
                        f"{shift_data['start_time']} - {shift_data['end_time']}"
                    ))
                    results_count += 1

        # 搜索排班计划
        if self.search_in_schedules.get():
            for person_name, schedule_data in self.shift_schedules.items():
                if search_term in person_name.lower():
                    pattern = " → ".join(schedule_data['shift_pattern'])
                    self.search_results_tree.insert("", tk.END, values=(
                        "排班计划",
                        person_name,
                        f"模式: {pattern}, 开始: {schedule_data['start_date']}"
                    ))
                    results_count += 1

        # 搜索请假记录
        if self.search_in_leaves.get():
            for leave_record in self.leave_records:
                if (search_term in leave_record['plan_name'].lower() or
                    search_term in leave_record['type'].lower() or
                    (leave_record.get('note', '') and search_term in leave_record['note'].lower())):
                    self.search_results_tree.insert("", tk.END, values=(
                        "请假记录",
                        leave_record['plan_name'],
                        f"{leave_record['date']} {leave_record['type']} - {leave_record.get('note', '')}"
                    ))
                    results_count += 1

        # 搜索节假日
        if self.search_in_holidays.get():
            for year, holidays in self.holidays.items():
                for date, name in holidays.items():
                    if search_term in name.lower():
                        self.search_results_tree.insert("", tk.END, values=(
                            "节假日",
                            name,
                            f"{year}-{date}"
                        ))
                        results_count += 1

        self.update_status(f"搜索完成，找到 {results_count} 个结果")

    def on_search_result_double_click(self, event):
        """双击搜索结果时的处理"""
        selected = self.search_results_tree.selection()
        if selected:
            item = selected[0]
            values = self.search_results_tree.item(item)['values']
            result_type = values[0]

            # 切换到对应的标签页
            if result_type == "班次类型":
                self.notebook.select(0)
            elif result_type == "排班计划":
                self.notebook.select(1)
            elif result_type == "请假记录":
                self.notebook.select(2)
            elif result_type == "节假日":
                self.notebook.select(3)

            # 关闭搜索对话框
            for widget in self.root.winfo_children():
                if isinstance(widget, tk.Toplevel) and widget.title() == "全局搜索":
                    widget.destroy()
                    break

    def show_settings(self):
        """显示设置对话框"""
        settings_dialog = tk.Toplevel(self.root)
        settings_dialog.title("系统设置")
        settings_dialog.geometry("550x580")
        settings_dialog.resizable(False, False)

        # 设置对话框居中
        settings_dialog.transient(self.root)
        settings_dialog.grab_set()

        # 创建设置界面
        main_frame = ttk.Frame(settings_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建标签页
        settings_notebook = ttk.Notebook(main_frame)
        settings_notebook.pack(fill=tk.BOTH, expand=True)

        # 常规设置标签页
        general_frame = ttk.Frame(settings_notebook)
        settings_notebook.add(general_frame, text="常规")

        ttk.Label(general_frame, text="常规设置", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(10, 20))

        # 自动备份设置
        backup_frame = ttk.Frame(general_frame)
        backup_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Checkbutton(backup_frame, text="启动时检查自动备份", variable=self.backup_enabled).pack(anchor=tk.W)

        # 数据保存设置
        save_frame = ttk.Frame(general_frame)
        save_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(save_frame, text="数据保存位置:").pack(anchor=tk.W)
        data_path_label = ttk.Label(save_frame, text="shift_data.json", font=('Microsoft YaHei UI', 8))
        data_path_label.pack(anchor=tk.W, pady=(5, 0))

        # 系统托盘设置
        tray_frame = ttk.Frame(general_frame)
        tray_frame.pack(fill=tk.X, pady=(10, 10))

        tray_check = ttk.Checkbutton(tray_frame, text="点击关闭后最小化到系统托盘",
                                     variable=self.minimize_to_tray,
                                     command=lambda: self._on_tray_setting_changed())
        tray_check.pack(anchor=tk.W)

        # 托盘功能说明
        if not TRAY_AVAILABLE:
            tray_note = ttk.Label(tray_frame,
                                 text="⚠ 需要安装 pystray 和 Pillow 库才能使用此功能",
                                 font=('Microsoft YaHei UI', 8), foreground='orange')
            tray_note.pack(anchor=tk.W, pady=(5, 0))
            tray_check.config(state='disabled')
        else:
            tray_note = ttk.Label(tray_frame,
                                 text="勾选后点击关闭按钮将最小化到托盘，右键托盘图标可恢复或退出",
                                 font=('Microsoft YaHei UI', 8), foreground='gray')
            tray_note.pack(anchor=tk.W, pady=(5, 0))

        # 备份设置标签页
        backup_tab_frame = ttk.Frame(settings_notebook)
        settings_notebook.add(backup_tab_frame, text="备份")

        ttk.Label(backup_tab_frame, text="备份设置", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(10, 20))

        # 备份信息
        info_frame = ttk.Frame(backup_tab_frame)
        info_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(info_frame, text="自动备份可以保护您的数据安全", font=('Microsoft YaHei UI', 10)).pack(anchor=tk.W)
        ttk.Label(info_frame, text="建议定期备份重要数据", font=('Microsoft YaHei UI', 10)).pack(anchor=tk.W, pady=(5, 0))

        # 外观设置标签页
        appearance_frame = ttk.Frame(settings_notebook)
        settings_notebook.add(appearance_frame, text="外观")

        ttk.Label(appearance_frame, text="主题设置", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(10, 20), anchor=tk.W, padx=20)

        # 主题选择
        theme_frame = ttk.Frame(appearance_frame)
        theme_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        ttk.Label(theme_frame, text="界面主题:").pack(anchor=tk.W, pady=(0, 10))

        # 创建主题选择的单选按钮
        theme_var_local = tk.StringVar(value=self.theme_var.get())

        light_radio = ttk.Radiobutton(theme_frame, text="☀️ 亮色调（默认）",
                                    variable=theme_var_local, value="light",
                                    command=lambda: self.apply_theme(theme_var_local.get()))
        light_radio.pack(anchor=tk.W, pady=(5, 0))

        dark_radio = ttk.Radiobutton(theme_frame, text="🌙 暗色调（护眼）",
                                   variable=theme_var_local, value="dark",
                                   command=lambda: self.apply_theme(theme_var_local.get()))
        dark_radio.pack(anchor=tk.W, pady=(5, 0))

        # 主题说明
        theme_info_frame = ttk.Frame(appearance_frame)
        theme_info_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        theme_info_label = ttk.Label(theme_info_frame,
                                   text="切换主题会立即应用到整个界面，选择最适合您工作环境的配色方案。",
                                   font=('Microsoft YaHei UI', 9),
                                   foreground=self.colors['text_secondary'])
        theme_info_label.pack(anchor=tk.W, pady=(5, 0))

        ttk.Label(appearance_frame, text="字体设置", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(20, 15), anchor=tk.W, padx=20)

        # 字体选择
        font_frame = ttk.Frame(appearance_frame)
        font_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        ttk.Label(font_frame, text="字体:").pack(side=tk.LEFT, padx=(0, 10))

        # 获取系统可用字体列表
        available_fonts = sorted(set(tk.font.families()))
        font_combo = ttk.Combobox(font_frame, textvariable=self.font_family, values=available_fonts,
                                   state="readonly", width=30)
        font_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 字体大小选择
        size_frame = ttk.Frame(appearance_frame)
        size_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

        ttk.Label(size_frame, text="字体大小:").pack(side=tk.LEFT, padx=(0, 10))
        size_spin = ttk.Spinbox(size_frame, from_=8, to=20, textvariable=self.font_size, width=10)
        size_spin.pack(side=tk.LEFT)
        ttk.Label(size_frame, text="pt").pack(side=tk.LEFT, padx=(5, 0))

        # 预览标签
        preview_frame = ttk.LabelFrame(appearance_frame, text="预览", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))

        preview_label = ttk.Label(preview_frame, text="这是字体预览示例文本\nThe quick brown fox jumps over the lazy dog")
        preview_label.pack(pady=10)

        def update_preview():
            """实时更新预览"""
            font_name = self.font_family.get()
            font_sz = self.font_size.get()
            preview_font = tk.font.Font(family=font_name, size=font_sz)
            preview_label.config(font=preview_font)

        # 为字体选择绑定事件
        font_combo.bind('<<ComboboxSelected>>', lambda e: update_preview())
        size_spin.bind('<KeyRelease>', lambda e: update_preview())

        # 立即更新预览
        update_preview()

        # 应用按钮
        apply_frame = ttk.Frame(appearance_frame)
        apply_frame.pack(fill=tk.X, padx=20, pady=(0, 10))

        def apply_font_settings():
            """应用字体设置"""
            self.save_data()  # 保存字体设置
            self.setup_modern_styles()  # 重新配置样式
            self.update_status("字体设置已应用并保存")
            messagebox.showinfo("成功", "字体设置已应用并保存！\n部分界面元素需重启应用才能完全生效")

        def reset_font_settings():
            """恢复默认字体设置"""
            if messagebox.askyesno("确认", "确定要恢复为默认字体设置吗？\n（Microsoft YaHei UI, 10pt）"):
                # 恢复值
                self.font_family.set("Microsoft YaHei UI")
                self.font_size.set(10)

                # 更新 Combobox 显示（使用索引，因为是readonly）
                try:
                    default_font_index = available_fonts.index("Microsoft YaHei UI")
                    font_combo.current(default_font_index)
                except (ValueError, tk.TclError):
                    # 如果找不到该字体，就设置为第一个
                    font_combo.current(0)

                # 更新 Spinbox 显示
                size_spin.delete(0, tk.END)
                size_spin.insert(0, "10")

                # 更新预览
                update_preview()

                # 直接保存和应用字体设置（不需要用户再点应用按钮）
                self.save_data()  # 保存字体设置
                self.setup_modern_styles()  # 重新配置样式

                # 反馈
                self.update_status("已恢复为默认字体设置并应用")
                messagebox.showinfo("成功", "已恢复为默认字体设置并自动应用！\n部分界面元素需重启应用才能完全生效")

        ttk.Button(apply_frame, text="应用字体设置", command=apply_font_settings).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(apply_frame, text="恢复默认", command=reset_font_settings).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(apply_frame, text="修改后需重启应用才能完全生效", font=('Microsoft YaHei UI', 8), foreground='gray').pack(side=tk.LEFT)

        # 多人日历标签页
        multi_cal_frame = ttk.Frame(settings_notebook)
        settings_notebook.add(multi_cal_frame, text="多人日历")

        ttk.Label(multi_cal_frame, text="多人日历显示设置", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(10, 20), anchor=tk.W, padx=20)

        # 说明文字
        ttk.Label(multi_cal_frame, text="调整多人日历视图中日期格子内的文字显示大小",
                 font=('Microsoft YaHei UI', 9)).pack(anchor=tk.W, padx=20, pady=(0, 15))

        # 字体大小设置
        mc_size_frame = ttk.Frame(multi_cal_frame)
        mc_size_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        ttk.Label(mc_size_frame, text="格子内字体大小:").pack(side=tk.LEFT, padx=(0, 10))
        mc_size_spin = ttk.Spinbox(mc_size_frame, from_=6, to=16, textvariable=self.multi_calendar_font_size, width=10)
        mc_size_spin.pack(side=tk.LEFT)
        ttk.Label(mc_size_frame, text="pt（默认9pt）").pack(side=tk.LEFT, padx=(5, 0))

        # 预览说明
        preview_info_frame = ttk.LabelFrame(multi_cal_frame, text="设置说明", padding=15)
        preview_info_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        preview_info_text = """此设置影响多人日历视图中以下元素的字体大小：
• 班次标签（如：白班、夜班）
• 成员名称
• 休假标签
• 日期数字
• "更多"提示文字

提示：字体越大，单个格子能显示的信息越少。
建议根据您的屏幕分辨率和排班人数调整。"""

        ttk.Label(preview_info_frame, text=preview_info_text,
                 font=('Microsoft YaHei UI', 9),
                 justify=tk.LEFT).pack(anchor=tk.W)

        # 应用按钮区域
        mc_apply_frame = ttk.Frame(multi_cal_frame)
        mc_apply_frame.pack(fill=tk.X, padx=20, pady=(10, 10))

        def apply_multi_cal_settings():
            """应用多人日历设置"""
            self.save_data()
            # 清除多人日历缓存，强制重新渲染
            if hasattr(self, '_multi_calendar_cell_cache'):
                self._multi_calendar_cell_cache.clear()
            # 清除标题相关标志，确保标题能够重新创建
            if hasattr(self, '_header_widgets'):
                delattr(self, '_header_widgets')
            if hasattr(self, '_multi_calendar_headers_created'):
                delattr(self, '_multi_calendar_headers_created')
            # 如果多人日历已初始化，立即刷新
            if hasattr(self, 'multi_calendar_container') and self._tabs_initialized.get('multi_calendar', False):
                self.update_multi_calendar()
                # 自动调整窗口大小以适应新字体
                self.root.after(100, self._auto_fit_calendar_display)
            self.update_status("多人日历设置已应用并保存")
            messagebox.showinfo("成功", "多人日历设置已应用！")

        def reset_multi_cal_settings():
            """恢复默认多人日历设置"""
            if messagebox.askyesno("确认", "确定要恢复为默认设置吗？\n（字体大小：9pt）"):
                self.multi_calendar_font_size.set(9)
                mc_size_spin.delete(0, tk.END)
                mc_size_spin.insert(0, "9")
                self.save_data()
                # 清除多人日历缓存，强制重新渲染
                if hasattr(self, '_multi_calendar_cell_cache'):
                    self._multi_calendar_cell_cache.clear()
                # 清除标题相关标志，确保标题能够重新创建
                if hasattr(self, '_header_widgets'):
                    delattr(self, '_header_widgets')
                if hasattr(self, '_multi_calendar_headers_created'):
                    delattr(self, '_multi_calendar_headers_created')
                if hasattr(self, 'multi_calendar_container') and self._tabs_initialized.get('multi_calendar', False):
                    self.update_multi_calendar()
                    # 自动调整窗口大小
                    self.root.after(100, self._auto_fit_calendar_display)
                self.update_status("已恢复为默认多人日历设置")
                messagebox.showinfo("成功", "已恢复为默认设置！")

        ttk.Button(mc_apply_frame, text="应用设置", command=apply_multi_cal_settings, style='Primary.TButton').pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(mc_apply_frame, text="恢复默认", command=reset_multi_cal_settings).pack(side=tk.LEFT)

        # 快捷键标签页
        shortcuts_frame = ttk.Frame(settings_notebook)
        settings_notebook.add(shortcuts_frame, text="快捷键")

        ttk.Label(shortcuts_frame, text="快捷键参考", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(10, 20))

        # 快捷键列表
        shortcuts_text = """全局快捷键:
Ctrl+N    - 新建班次
Ctrl+O    - 编辑班次
Ctrl+D    - 删除班次
Ctrl+S    - 保存数据
Ctrl+F    - 全局搜索
Ctrl+B    - 备份数据
F1        - 显示帮助
F5        - 刷新数据
Esc       - 关闭对话框

标签页切换:
Ctrl+1~5  - 切换到对应标签页

日历导航:
← →      - 上个月/下个月
↑         - 当前月份
Home      - 跳转到今天

通用操作:
Delete    - 删除选中项目
Enter     - 编辑选中项目"""

        shortcuts_label = ttk.Label(shortcuts_frame, text=shortcuts_text, font=('Microsoft YaHei UI', 8))
        shortcuts_label.pack(anchor=tk.W, padx=(20, 0))

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        ttk.Button(button_frame, text="备份设置", command=lambda: [self.show_backup_settings(), settings_dialog.destroy()],
                   style='Primary.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="查看帮助", command=self.show_help,
                   style='Info.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="关闭", command=settings_dialog.destroy,
                   style='Danger.TButton').pack(side=tk.RIGHT)

    # ==================== 系统托盘功能 ====================

    def _on_tray_setting_changed(self):
        """托盘设置更改时的回调"""
        self.save_data()
        if self.minimize_to_tray.get():
            self.update_status("已启用最小化到托盘功能")
        else:
            self.update_status("已禁用最小化到托盘功能")

    def _create_tray_image(self):
        """创建托盘图标图像"""
        if not TRAY_AVAILABLE:
            return None
        # 创建一个简单的图标（蓝色圆形带白色日历图案）
        size = 64
        image = Image.new('RGBA', (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)
        # 绘制蓝色圆形背景
        draw.ellipse([2, 2, size-2, size-2], fill='#2196F3')
        # 绘制简单的日历图案
        margin = 14
        draw.rectangle([margin, margin+6, size-margin, size-margin], fill='white')
        draw.rectangle([margin, margin, size-margin, margin+8], fill='#1976D2')
        # 绘制日历格子
        for i in range(3):
            for j in range(3):
                x = margin + 6 + i * 12
                y = margin + 16 + j * 10
                draw.rectangle([x, y, x+8, y+6], fill='#E3F2FD')
        return image

    def _create_tray_icon(self):
        """创建系统托盘图标"""
        if not TRAY_AVAILABLE or self.tray_icon is not None:
            return

        image = self._create_tray_image()
        if image is None:
            return

        # 创建托盘菜单
        menu = pystray.Menu(
            pystray.MenuItem("显示窗口", self._show_window_from_tray, default=True),
            pystray.MenuItem("退出程序", self._quit_app_from_tray)
        )

        # 创建托盘图标
        self.tray_icon = pystray.Icon(
            "排班助手",
            image,
            "排班助手 - 运行中",
            menu
        )

        # 在新线程中运行托盘图标
        tray_thread = threading.Thread(target=self.tray_icon.run, daemon=True)
        tray_thread.start()

    def _show_window_from_tray(self, icon=None, item=None):
        """从托盘恢复窗口"""
        self.root.after(0, self._show_window)

    def _show_window(self):
        """显示主窗口"""
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()
        self.update_status("窗口已恢复")

    def _quit_app_from_tray(self, icon=None, item=None):
        """从托盘退出应用"""
        self.root.after(0, self._quit_app)

    def _quit_app(self):
        """完全退出应用"""
        # 停止托盘图标
        if self.tray_icon is not None:
            self.tray_icon.stop()
            self.tray_icon = None
        # 保存数据
        self.save_data()
        # 销毁窗口
        self.root.destroy()

    def _minimize_to_tray(self):
        """最小化到系统托盘"""
        if not TRAY_AVAILABLE:
            return False

        # 创建托盘图标（如果还没有）
        self._create_tray_icon()

        # 隐藏主窗口
        self.root.withdraw()
        self.update_status("已最小化到系统托盘")
        return True

    def _on_window_close(self):
        """窗口关闭事件处理"""
        if self.minimize_to_tray.get() and TRAY_AVAILABLE:
            # 最小化到托盘
            self._minimize_to_tray()
        else:
            # 直接退出
            self._quit_app()

    def setup_shortcuts(self):
        """设置快捷键"""
        # 全局快捷键
        self.root.bind('<Control-n>', lambda e: self.add_shift_type())
        self.root.bind('<Control-o>', lambda e: self.edit_shift_type())
        self.root.bind('<Control-d>', lambda e: self.delete_shift_type())
        self.root.bind('<Control-s>', lambda e: self.save_data())
        self.root.bind('<Control-f>', lambda e: self.show_search_dialog())
        self.root.bind('<Control-b>', lambda e: self.backup_data())
        self.root.bind('<F1>', lambda e: self.show_help())
        self.root.bind('<F5>', lambda e: self.refresh_all_data())
        self.root.bind('<Escape>', lambda e: self.close_current_dialog())

        # 标签页切换快捷键
        self.root.bind('<Control-1>', lambda e: self.switch_to_tab(0))
        self.root.bind('<Control-2>', lambda e: self.switch_to_tab(1))
        self.root.bind('<Control-3>', lambda e: self.switch_to_tab(2))
        self.root.bind('<Control-4>', lambda e: self.switch_to_tab(3))
        self.root.bind('<Control-5>', lambda e: self.switch_to_tab(4))

        # 日历导航快捷键
        self.root.bind('<Left>', lambda e: self.prev_month() if self.notebook.index(self.notebook.select()) == 4 else None)
        self.root.bind('<Right>', lambda e: self.next_month() if self.notebook.index(self.notebook.select()) == 4 else None)
        self.root.bind('<Up>', lambda e: self.show_current_month() if self.notebook.index(self.notebook.select()) == 4 else None)
        self.root.bind('<Home>', lambda e: self.go_to_today())

        # 删除键和回车键
        self.root.bind('<Delete>', lambda e: self.delete_selected_item())
        self.root.bind('<Return>', lambda e: self.edit_selected_item())

    def switch_to_tab(self, index):
        """切换到指定标签页"""
        try:
            self.notebook.select(index)
        except:
            pass

    def show_help(self):
        """显示帮助对话框"""
        help_text = """排班日历专业版 v3.0 快捷键帮助

全局快捷键:
Ctrl+N    - 新建班次
Ctrl+O    - 编辑班次
Ctrl+D    - 删除班次
Ctrl+S    - 保存数据
Ctrl+F    - 搜索
Ctrl+B    - 备份数据
F1        - 显示帮助
F5        - 刷新数据
Esc       - 关闭当前对话框

标签页切换:
Ctrl+1    - 班次类型管理
Ctrl+2    - 排班计划管理
Ctrl+3    - 请假管理
Ctrl+4    - 节假日管理
Ctrl+5    - 日历视图

日历导航:
← →      - 上个月/下个月 (仅在日历视图)
↑         - 当前月份 (仅在日历视图)
Home      - 跳转到今天

通用操作:
Delete    - 删除选中项目
Enter     - 编辑选中项目

右键菜单:
在列表中右键点击可显示上下文菜单"""
        messagebox.showinfo("快捷键帮助", help_text)

    def refresh_all_data(self):
        """刷新所有数据"""
        self.update_shift_type_tree()
        self.update_schedule_tree()
        self.update_leave_tree()
        self.update_holiday_tree()
        self.update_calendar()
        self.update_status("所有数据已刷新")

    def close_current_dialog(self):
        """关闭当前对话框"""
        # 查找并关闭当前活动的顶级窗口
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Toplevel) and widget.winfo_viewable():
                widget.destroy()
                break

    def delete_selected_item(self):
        """删除选中项目"""
        current_tab = self.notebook.index(self.notebook.select())
        if current_tab == 0:  # 班次类型管理
            self.delete_shift_type()
        elif current_tab == 1:  # 排班计划管理
            self.delete_schedule()
        elif current_tab == 2:  # 请假管理
            self.delete_leave_record()
        elif current_tab == 3:  # 节假日管理
            self.delete_selected_holiday()

    def edit_selected_item(self):
        """编辑选中项目"""
        current_tab = self.notebook.index(self.notebook.select())
        if current_tab == 0:  # 班次类型管理
            self.edit_shift_type()
        elif current_tab == 1:  # 排班计划管理
            self.edit_schedule()
        elif current_tab == 2:  # 请假管理
            self.edit_leave_record()
        elif current_tab == 3:  # 节假日管理
            self.edit_holiday_dialog()

    def setup_shift_type_tab(self):
        """班次类型管理标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="班次类型管理")

        # 创建标题栏
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

        title_label = ttk.Label(title_frame, text="班次类型管理",
                               font=('Microsoft YaHei UI', 14, 'bold'),
                               foreground=self.colors['primary'])
        title_label.pack(side=tk.LEFT)

        # 创建操作按钮区域
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(button_frame, text="➕ 添加班次", command=self.add_shift_type,
                   style='Success.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="✏️ 编辑班次", command=self.edit_shift_type,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="🗑️ 删除班次", command=self.delete_shift_type,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=2)

        # 创建搜索和筛选区域
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT, padx=(0, 5))
        self.shift_type_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.shift_type_search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        search_entry.bind('<KeyRelease>', lambda e: self.filter_shift_types())

        # 创建班次类型列表
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 创建滚动条
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 班次类型列表
        self.shift_type_tree = ttk.Treeview(tree_frame, columns=("name", "start", "end", "color"),
                                          show="headings", yscrollcommand=scrollbar.set)
        self.shift_type_tree.heading("name", text="班次名称")
        self.shift_type_tree.heading("start", text="开始时间")
        self.shift_type_tree.heading("end", text="结束时间")
        self.shift_type_tree.heading("color", text="颜色")

        # 设置列宽
        self.shift_type_tree.column("name", width=150)
        self.shift_type_tree.column("start", width=120)
        self.shift_type_tree.column("end", width=120)
        self.shift_type_tree.column("color", width=100)

        self.shift_type_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.shift_type_tree.yview)

        # 绑定双击事件
        self.shift_type_tree.bind('<Double-1>', lambda e: self.edit_shift_type())

        # 添加右键菜单
        self.create_shift_type_context_menu()

        # 更新树视图
        self.update_shift_type_tree()

        # 配置网格权重
        frame.grid_rowconfigure(4, weight=1)
        frame.grid_columnconfigure(0, weight=1)

    def create_shift_type_context_menu(self):
        """创建班次类型右键菜单"""
        self.shift_type_context_menu = tk.Menu(self.root, tearoff=0)
        self.shift_type_context_menu.add_command(label="编辑班次", command=self.edit_shift_type)
        self.shift_type_context_menu.add_command(label="删除班次", command=self.delete_shift_type)
        self.shift_type_context_menu.add_separator()
        self.shift_type_context_menu.add_command(label="复制班次", command=self.copy_shift_type)
        self.shift_type_context_menu.add_command(label="粘贴班次", command=self.paste_shift_type)

        # 绑定右键菜单
        self.shift_type_tree.bind('<Button-3>', self.show_shift_type_context_menu)

    def show_shift_type_context_menu(self, event):
        """显示班次类型右键菜单"""
        item = self.shift_type_tree.identify('item', event.x, event.y)
        if item:
            self.shift_type_tree.selection_set(item)
            self.shift_type_context_menu.post(event.x_root, event.y_root)

    def copy_shift_type(self):
        """复制班次类型"""
        selected = self.shift_type_tree.selection()
        if selected:
            item = selected[0]
            values = self.shift_type_tree.item(item)['values']
            self.copied_shift_type = {
                'name': values[0],
                'start_time': values[1],
                'end_time': values[2],
                'color': values[3]
            }
            self.update_status(f"已复制班次: {values[0]}")

    def paste_shift_type(self):
        """粘贴班次类型"""
        if hasattr(self, 'copied_shift_type'):
            # 创建新的班次名称
            base_name = self.copied_shift_type['name']
            new_name = f"{base_name}_副本"
            counter = 1
            while new_name in self.shift_types:
                new_name = f"{base_name}_副本{counter}"
                counter += 1

            # 添加新的班次类型
            self.shift_types[new_name] = {
                'start_time': self.copied_shift_type['start_time'],
                'end_time': self.copied_shift_type['end_time'],
                'color': self.copied_shift_type['color']
            }

            self.update_shift_type_tree()
            self.save_data()
            self.update_status(f"已粘贴班次: {new_name}")
        else:
            messagebox.showwarning("提示", "请先复制一个班次类型")

    def filter_shift_types(self):
        """过滤班次类型"""
        search_term = self.shift_type_search_var.get().lower()
        for item in self.shift_type_tree.get_children():
            values = self.shift_type_tree.item(item)['values']
            if search_term in values[0].lower():
                self.shift_type_tree.item(item, tags=())
            else:
                self.shift_type_tree.item(item, tags=('hidden',))

        # 隐藏匹配的项
        self.shift_type_tree.tag_configure('hidden', hide=True)
    
    def setup_schedule_tab(self):
        """排班计划管理标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="排班计划管理")

        # 创建标题栏
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

        title_label = ttk.Label(title_frame, text="排班计划管理",
                               font=('Microsoft YaHei UI', 14, 'bold'),
                               foreground=self.colors['primary'])
        title_label.pack(side=tk.LEFT)

        # 创建操作按钮区域
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(button_frame, text="➕ 新建计划", command=self.create_schedule,
                   style='Success.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="✏️ 编辑计划", command=self.edit_schedule,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="🗑️ 删除计划", command=self.delete_schedule,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="🔄 生成排班", command=self.generate_schedule,
                   style='Warning.TButton').pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="📋 批量操作", command=self.show_batch_operations,
                   style='Info.TButton').pack(side=tk.LEFT, padx=2)

        # 创建搜索和筛选区域
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(search_frame, text="搜索人员:").pack(side=tk.LEFT, padx=(0, 5))
        self.schedule_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.schedule_search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        search_entry.bind('<KeyRelease>', lambda e: self.filter_schedules())

        # 创建排班计划列表
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 创建滚动条
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 排班计划列表
        self.schedule_tree = ttk.Treeview(tree_frame, columns=("name", "pattern", "start_date"),
                                        show="headings", yscrollcommand=scrollbar.set)
        self.schedule_tree.heading("name", text="人员名称")
        self.schedule_tree.heading("pattern", text="轮班模式")
        self.schedule_tree.heading("start_date", text="开始日期")

        # 设置列宽
        self.schedule_tree.column("name", width=150)
        self.schedule_tree.column("pattern", width=300)
        self.schedule_tree.column("start_date", width=120)

        self.schedule_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.schedule_tree.yview)

        # 绑定双击事件，允许用户通过双击选择当前人员
        self.schedule_tree.bind("<Double-1>", self.select_current_person)

        # 添加右键菜单
        self.create_schedule_context_menu()

        # 更新树视图
        self.update_schedule_tree()

        # 配置网格权重
        frame.grid_rowconfigure(4, weight=1)
        frame.grid_columnconfigure(0, weight=1)

    def create_schedule_context_menu(self):
        """创建排班计划右键菜单"""
        self.schedule_context_menu = tk.Menu(self.root, tearoff=0)
        self.schedule_context_menu.add_command(label="编辑计划", command=self.edit_schedule)
        self.schedule_context_menu.add_command(label="删除计划", command=self.delete_schedule)
        self.schedule_context_menu.add_separator()
        self.schedule_context_menu.add_command(label="复制计划", command=self.copy_schedule)
        self.schedule_context_menu.add_command(label="粘贴计划", command=self.paste_schedule)
        self.schedule_context_menu.add_separator()
        self.schedule_context_menu.add_command(label="设为当前人员", command=self.select_current_person)

        # 绑定右键菜单
        self.schedule_tree.bind('<Button-3>', self.show_schedule_context_menu)

    def show_schedule_context_menu(self, event):
        """显示排班计划右键菜单"""
        item = self.schedule_tree.identify('item', event.x, event.y)
        if item:
            self.schedule_tree.selection_set(item)
            self.schedule_context_menu.post(event.x_root, event.y_root)

    def copy_schedule(self):
        """复制排班计划"""
        selected = self.schedule_tree.selection()
        if selected:
            item = selected[0]
            values = self.schedule_tree.item(item)['values']
            self.copied_schedule = {
                'name': values[0],
                'pattern': values[1],
                'start_date': values[2]
            }
            self.update_status(f"已复制排班计划: {values[0]}")

    def paste_schedule(self):
        """粘贴排班计划"""
        if hasattr(self, 'copied_schedule'):
            # 创建新的计划名称
            base_name = self.copied_schedule['name']
            new_name = f"{base_name}_副本"
            counter = 1
            while new_name in self.shift_schedules:
                new_name = f"{base_name}_副本{counter}"
                counter += 1

            # 添加新的排班计划
            self.shift_schedules[new_name] = {
                'shift_pattern': self.copied_schedule['pattern'].split(' → '),
                'start_date': self.copied_schedule['start_date'],
                'shifts': {}
            }

            self.update_schedule_tree()
            self.save_data()
            self.update_status(f"已粘贴排班计划: {new_name}")
        else:
            messagebox.showwarning("提示", "请先复制一个排班计划")

    def filter_schedules(self):
        """过滤排班计划"""
        search_term = self.schedule_search_var.get().lower()
        for item in self.schedule_tree.get_children():
            values = self.schedule_tree.item(item)['values']
            if search_term in values[0].lower():
                self.schedule_tree.item(item, tags=())
            else:
                self.schedule_tree.item(item, tags=('hidden',))

        # 隐藏匹配的项
        self.schedule_tree.tag_configure('hidden', hide=True)

    def show_batch_operations(self):
        """显示批量操作对话框"""
        # 临时实现，后续会完善
        messagebox.showinfo("批量操作", "批量操作功能正在开发中...")
    
    def setup_calendar_tab(self):
        """单人日历视图标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="单人日历视图")
        
        # 控制栏
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(control_frame, text="上个月", command=self.prev_month).pack(side=tk.LEFT)
        ttk.Button(control_frame, text="今天", command=self.show_current_month).pack(side=tk.LEFT, padx=10)
        ttk.Button(control_frame, text="下个月", command=self.next_month).pack(side=tk.LEFT)
        
        
        self.month_year_var = tk.StringVar()
        ttk.Label(control_frame, textvariable=self.month_year_var).pack(side=tk.LEFT, expand=True)

        # 年份选择下拉
        ttk.Label(control_frame, text="年份").pack(side=tk.LEFT, padx=(0, 6))
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(control_frame, textvariable=self.year_var, width=6, state="readonly")
        self.year_combo.pack(side=tk.LEFT)
        self.year_combo.bind('<<ComboboxSelected>>', lambda e: self.on_year_selected())
        self.update_year_options()

        # 月份选择下拉
        ttk.Label(control_frame, text="月份").pack(side=tk.LEFT, padx=(12, 6))
        self.month_var = tk.StringVar()
        self.month_combo = ttk.Combobox(control_frame, textvariable=self.month_var, width=4, state="readonly")
        self.month_combo.pack(side=tk.LEFT)
        self.month_combo.bind('<<ComboboxSelected>>', lambda e: self.on_month_selected())
        self.update_month_options()

        # 视图开关：节假日与请假
        ttk.Checkbutton(control_frame, text="显示节假日", variable=self.show_holidays,
                        command=self.update_calendar).pack(side=tk.LEFT, padx=(16, 6))
        ttk.Checkbutton(control_frame, text="显示请假", variable=self.show_leaves,
                        command=self.update_calendar).pack(side=tk.LEFT)
        
        # 日历显示区域
        self.calendar_container = ttk.Frame(frame)
        self.calendar_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        # 让7列、8行单元格自适应拉伸（增加人员信息行）
        for c in range(7):
            self.calendar_container.columnconfigure(c, weight=1, minsize=110)
        for r in range(8):  # 包含人员信息行、标题行和6行日期
            self.calendar_container.rowconfigure(r, weight=1, minsize=90)
        
        # 图例区：展示班次颜色、节假日与请假标识
        self.legend_frame = ttk.Frame(frame)
        self.legend_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        self._render_legend()

        self.update_calendar()

    def setup_holiday_calendar_tab(self):
        """当月休假日历标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="当月休假日历")

        # 创建标题栏
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        title_label = ttk.Label(title_frame, text="当月休假日历",
                               font=('Microsoft YaHei UI', 14, 'bold'),
                               foreground=self.colors['primary'])
        title_label.pack(side=tk.LEFT)

        # 控制面板框架
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        # 月份导航按钮
        ttk.Button(control_frame, text="上个月", command=self.holiday_calendar_prev_month).pack(side=tk.LEFT)
        ttk.Button(control_frame, text="今天", command=self.holiday_calendar_show_current_month).pack(side=tk.LEFT, padx=10)
        ttk.Button(control_frame, text="下个月", command=self.holiday_calendar_next_month).pack(side=tk.LEFT)

        # 月份年份显示
        self.holiday_month_year_var = tk.StringVar()
        ttk.Label(control_frame, textvariable=self.holiday_month_year_var,
                 font=('Microsoft YaHei UI', 12, 'bold')).pack(side=tk.LEFT, padx=20)

        # 年份选择
        ttk.Label(control_frame, text="年份:").pack(side=tk.LEFT, padx=(20, 5))
        self.holiday_year_var = tk.StringVar(value=str(self.current_date.year))
        self.holiday_year_combo = ttk.Combobox(control_frame, textvariable=self.holiday_year_var,
                                              width=6, state="readonly")
        self.holiday_year_combo.pack(side=tk.LEFT)
        self.holiday_year_combo.bind('<<ComboboxSelected>>', lambda e: self.holiday_calendar_on_year_selected())

        # 月份选择
        ttk.Label(control_frame, text="月份:").pack(side=tk.LEFT, padx=(12, 5))
        self.holiday_month_var = tk.StringVar(value=str(self.current_date.month))
        self.holiday_month_combo = ttk.Combobox(control_frame, textvariable=self.holiday_month_var,
                                               width=4, state="readonly")
        self.holiday_month_combo.pack(side=tk.LEFT)
        self.holiday_month_combo.bind('<<ComboboxSelected>>', lambda e: self.holiday_calendar_on_month_selected())

        # 筛选区域
        filter_frame = ttk.Frame(frame)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # 成员筛选
        ttk.Label(filter_frame, text="成员:").pack(side=tk.LEFT, padx=(0, 5))
        self.holiday_member_var = tk.StringVar(value="全部成员")
        self.holiday_member_combo = ttk.Combobox(filter_frame, textvariable=self.holiday_member_var,
                                               width=12, state="readonly")
        self.holiday_member_combo.pack(side=tk.LEFT)
        self.holiday_member_combo.bind('<<ComboboxSelected>>', lambda e: self.update_holiday_calendar())

        # 请假类型筛选
        ttk.Label(filter_frame, text="请假类型:").pack(side=tk.LEFT, padx=(20, 5))
        self.holiday_leave_type_var = tk.StringVar(value="全部类型")
        self.holiday_leave_type_combo = ttk.Combobox(filter_frame, textvariable=self.holiday_leave_type_var,
                                                    width=12, state="readonly")
        self.holiday_leave_type_combo.pack(side=tk.LEFT)
        self.holiday_leave_type_combo.bind('<<ComboboxSelected>>', lambda e: self.update_holiday_calendar())

        # 刷新按钮
        ttk.Button(filter_frame, text="刷新", command=self.update_holiday_calendar,
                  style='Info.TButton').pack(side=tk.LEFT, padx=(20, 0))

        # 主要内容区域：左右分栏
        main_content_frame = ttk.Frame(frame)
        main_content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 配置主内容区域的权重分配
        main_content_frame.columnconfigure(0, weight=3)  # 左侧日历区域占更多空间
        main_content_frame.columnconfigure(1, weight=1)  # 右侧统计区域占较少空间
        main_content_frame.rowconfigure(0, weight=1)

        # 左侧：日历视图（带滚动条）
        left_frame = ttk.LabelFrame(main_content_frame, text="日历视图", padding=10)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5))
        left_frame.rowconfigure(0, weight=1)
        left_frame.columnconfigure(0, weight=1)

        # 创建日历滚动区域
        canvas_frame = tk.Frame(left_frame)
        canvas_frame.grid(row=0, column=0, sticky='nsew')
        canvas_frame.rowconfigure(0, weight=1)
        canvas_frame.columnconfigure(0, weight=1)

        # 创建Canvas和滚动条
        self.holiday_calendar_canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        self.holiday_calendar_canvas.grid(row=0, column=0, sticky='nsew')

        # 垂直滚动条
        holiday_v_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL,
                                          command=self.holiday_calendar_canvas.yview)
        holiday_v_scrollbar.grid(row=0, column=1, sticky='ns')

        # 水平滚动条
        holiday_h_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL,
                                          command=self.holiday_calendar_canvas.xview)
        holiday_h_scrollbar.grid(row=1, column=0, sticky='ew')

        # 配置Canvas滚动
        self.holiday_calendar_canvas.configure(yscrollcommand=holiday_v_scrollbar.set,
                                             xscrollcommand=holiday_h_scrollbar.set)

        # 绑定鼠标滚轮事件到Canvas
        self._bind_mousewheel(self.holiday_calendar_canvas)

        # 创建可滚动的日历容器
        self.holiday_calendar_container = ttk.Frame(self.holiday_calendar_canvas)
        self.holiday_calendar_canvas_window = self.holiday_calendar_canvas.create_window(
            (0, 0), window=self.holiday_calendar_container, anchor='nw'
        )

        # 绑定Canvas配置事件以更新滚动区域
        self.holiday_calendar_container.bind('<Configure>', self._on_holiday_calendar_configure)
        self.holiday_calendar_canvas.bind('<Configure>', self._on_canvas_configure)

        # 右侧：统计信息
        right_frame = ttk.LabelFrame(main_content_frame, text="休假统计", padding=10)
        right_frame.grid(row=0, column=1, sticky='nsew', padx=(5, 0))
        right_frame.rowconfigure(0, weight=1)
        right_frame.columnconfigure(0, weight=1)

        # 统计信息显示区域
        self.holiday_stats_text = tk.Text(right_frame, wrap=tk.WORD,
                                         font=('Microsoft YaHei UI', 9))
        self.holiday_stats_text.grid(row=0, column=0, sticky='nsew')

        # 统计信息滚动条
        stats_scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL,
                                       command=self.holiday_stats_text.yview)
        stats_scrollbar.grid(row=0, column=1, sticky='ns')
        self.holiday_stats_text.config(yscrollcommand=stats_scrollbar.set)

        # 绑定鼠标滚轮事件到统计文本区域
        self._bind_mousewheel(self.holiday_stats_text)

        # 设置统计区域的最小宽度（通过设置minsize）
        right_frame.grid_propagate(False)
        self.root.after(100, lambda: right_frame.config(width=280))  # 延迟设置宽度

        # 底部：图例
        legend_frame = ttk.Frame(frame)
        legend_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        ttk.Label(legend_frame, text="请假类型图例:",
                 font=('Microsoft YaHei UI', 10, 'bold')).pack(side=tk.LEFT)
        self.holiday_legend_frame = ttk.Frame(legend_frame)
        self.holiday_legend_frame.pack(side=tk.LEFT, padx=(10, 0))

        # 初始化控件
        self._init_holiday_calendar_controls()

        # 初始化日历
        self.update_holiday_calendar()

        # 延迟优化显示尺寸（等待UI完全加载）
        self.root.after(500, self._optimize_holiday_calendar_display)

    def setup_swap_management_tab(self):
        """调班管理标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="调班管理")

        # 标题栏
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        ttk.Label(title_frame, text="调班管理",
                 font=('Microsoft YaHei UI', 14, 'bold'),
                 foreground=self.colors['primary']).pack(side=tk.LEFT)

        # 工具栏
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(toolbar, text="新增调班", command=self.add_swap_record,
                  style='Success.TButton').pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar, text="刷新", command=self.refresh_swap_list).pack(side=tk.LEFT)

        # 调班记录列表
        list_frame = ttk.LabelFrame(frame, text="调班记录", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 创建Treeview
        columns = ("swap_id", "person_a", "date_a", "shift_a", "person_b", "date_b", "shift_b", "timestamp")
        self.swap_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)

        # 设置列标题
        self.swap_tree.heading("swap_id", text="调班ID")
        self.swap_tree.heading("person_a", text="人员A")
        self.swap_tree.heading("date_a", text="日期A")
        self.swap_tree.heading("shift_a", text="班次A")
        self.swap_tree.heading("person_b", text="人员B")
        self.swap_tree.heading("date_b", text="日期B")
        self.swap_tree.heading("shift_b", text="班次B")
        self.swap_tree.heading("timestamp", text="调班时间")

        # 设置列宽
        self.swap_tree.column("swap_id", width=0, stretch=False)
        self.swap_tree.column("person_a", width=80)
        self.swap_tree.column("date_a", width=100)
        self.swap_tree.column("shift_a", width=80)
        self.swap_tree.column("person_b", width=80)
        self.swap_tree.column("date_b", width=100)
        self.swap_tree.column("shift_b", width=80)
        self.swap_tree.column("timestamp", width=150)

        # 滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.swap_tree.yview)
        self.swap_tree.configure(yscrollcommand=scrollbar.set)

        self.swap_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 右键菜单
        self.swap_tree.bind("<Button-3>", self.show_swap_context_menu)

        # 加载数据
        self.refresh_swap_list()

    def setup_multi_member_calendar_tab(self):
        """多人日历视图标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="多人日历视图")

        # 创建标题栏
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        title_label = ttk.Label(title_frame, text="多人日历视图",
                               font=('Microsoft YaHei UI', 14, 'bold'),
                               foreground=self.colors['primary'])
        title_label.pack(side=tk.LEFT)

        # 控制面板框架
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        # 月份导航按钮
        ttk.Button(control_frame, text="上个月", command=self.multi_calendar_prev_month).pack(side=tk.LEFT)
        ttk.Button(control_frame, text="今天", command=self.multi_calendar_show_current_month).pack(side=tk.LEFT, padx=10)
        ttk.Button(control_frame, text="下个月", command=self.multi_calendar_next_month).pack(side=tk.LEFT)

        # 月份年份显示
        self.multi_month_year_var = tk.StringVar()
        ttk.Label(control_frame, textvariable=self.multi_month_year_var,
                 font=('Microsoft YaHei UI', 12, 'bold')).pack(side=tk.LEFT, padx=20)

        # 年份选择
        ttk.Label(control_frame, text="年份:").pack(side=tk.LEFT, padx=(20, 5))
        self.multi_year_var = tk.StringVar(value=str(self.current_date.year))
        self.multi_year_combo = ttk.Combobox(control_frame, textvariable=self.multi_year_var,
                                            width=6, state="readonly")
        self.multi_year_combo.pack(side=tk.LEFT)
        self.multi_year_combo.bind('<<ComboboxSelected>>', lambda e: self.multi_calendar_on_year_selected())

        # 月份选择
        ttk.Label(control_frame, text="月份:").pack(side=tk.LEFT, padx=(12, 5))
        self.multi_month_var = tk.StringVar(value=f"{self.current_date.month:02d}")
        self.multi_month_combo = ttk.Combobox(control_frame, textvariable=self.multi_month_var,
                                             width=4, state="readonly")
        self.multi_month_combo.pack(side=tk.LEFT)
        self.multi_month_combo.bind('<<ComboboxSelected>>', lambda e: self.multi_calendar_on_month_selected())

        # 筛选区域
        filter_frame = ttk.Frame(frame)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # 班次筛选
        ttk.Label(filter_frame, text="班次:").pack(side=tk.LEFT, padx=(0, 5))
        self.multi_shift_var = tk.StringVar(value="全部班次")
        self.multi_shift_combo = ttk.Combobox(filter_frame, textvariable=self.multi_shift_var,
                                            width=12, state="readonly")
        self.multi_shift_combo.pack(side=tk.LEFT)
        self.multi_shift_combo.bind('<<ComboboxSelected>>', lambda e: self.update_multi_calendar())

        # 刷新按钮
        ttk.Button(filter_frame, text="刷新", command=self.update_multi_calendar,
                  style='Info.TButton').pack(side=tk.LEFT, padx=(20, 0))

        # 导出当月考勤按钮
        ttk.Button(filter_frame, text="导出当月考勤", command=self.export_monthly_attendance,
                  style='Primary.TButton').pack(side=tk.LEFT, padx=(10, 0))

        # 主要内容区域：日历视图（占满整个宽度）
        main_content_frame = ttk.Frame(frame)
        main_content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 日历视图容器
        self.multi_calendar_container = ttk.Frame(main_content_frame)
        self.multi_calendar_container.pack(fill=tk.BOTH, expand=True)

        # 底部：班次图例
        legend_frame = ttk.Frame(frame)
        legend_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        ttk.Label(legend_frame, text="班次图例:",
                 font=('Microsoft YaHei UI', 10, 'bold')).pack(side=tk.LEFT)
        self.multi_legend_frame = ttk.Frame(legend_frame)
        self.multi_legend_frame.pack(side=tk.LEFT, padx=(10, 0))

        # 初始化控件
        self._init_multi_calendar_controls()

        # 标记多人日历是否已初始化渲染（用于避免重复刷新）
        self._multi_calendar_rendered = False
        self._multi_calendar_first_show = True  # 首次显示标志

        # 初始化日历 - 不再延迟渲染，等首次显示时再渲染
        # self.root.after(100, self._init_multi_calendar_delayed)  # 注释掉延迟初始化

    def _init_multi_calendar_controls(self):
        """初始化多人日历控件选项"""
        # 更新年份选项（当前年份前后5年）
        current_year = self.current_date.year
        years = [str(year) for year in range(current_year - 5, current_year + 6)]
        self.multi_year_combo['values'] = years

        # 更新月份选项
        months = [f"{month:02d}" for month in range(1, 13)]
        self.multi_month_combo['values'] = months

        # 更新班次选项（排除休息和常日班，并按优先级排序）
        exclude_shifts = ["休息", "常日班"]
        all_shift_types = [shift for shift in self.shift_types.keys() if shift not in exclude_shifts]

        def sort_shift_options(shift_type):
            if shift_type == '白班':
                return (0, shift_type)
            elif shift_type == '夜班':
                return (1, shift_type)
            else:
                return (2, shift_type)

        sorted_shift_types = sorted(all_shift_types, key=sort_shift_options)
        shift_types = ["全部班次"] + sorted_shift_types
        self.multi_shift_combo['values'] = shift_types

    def multi_calendar_prev_month(self):
        """多人日历显示上个月"""
        try:
            year = int(self.multi_year_var.get())
            month = int(self.multi_month_var.get())
        except ValueError:
            year = self.current_date.year
            month = self.current_date.month

        # 计算上个月
        if month == 1:
            year -= 1
            month = 12
        else:
            month -= 1

        self.multi_year_var.set(str(year))
        self.multi_month_var.set(f"{month:02d}")
        self.update_multi_calendar()

    def multi_calendar_next_month(self):
        """多人日历显示下个月"""
        try:
            year = int(self.multi_year_var.get())
            month = int(self.multi_month_var.get())
        except ValueError:
            year = self.current_date.year
            month = self.current_date.month

        # 计算下个月
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1

        self.multi_year_var.set(str(year))
        self.multi_month_var.set(f"{month:02d}")
        self.update_multi_calendar()

    def multi_calendar_show_current_month(self):
        """多人日历显示当前月份"""
        today = datetime.date.today()
        self.multi_year_var.set(str(today.year))
        self.multi_month_var.set(f"{today.month:02d}")
        self.update_multi_calendar()

    def multi_calendar_on_year_selected(self):
        """年份选择事件处理"""
        self.update_multi_calendar()

    def multi_calendar_on_month_selected(self):
        """月份选择事件处理"""
        self.update_multi_calendar()

    def update_multi_calendar(self):
        """更新多人日历显示"""
        # 清除缓存以确保显示最新数据
        if hasattr(self, '_calendar_data_cache'):
            self._calendar_data_cache.clear()
        if hasattr(self, '_multi_calendar_cell_cache'):
            self._multi_calendar_cell_cache.clear()

        try:
            # 获取当前选择的年月
            year = int(self.multi_year_var.get())
            month = int(self.multi_month_var.get())
        except ValueError:
            today = datetime.date.today()
            year = today.year
            month = today.month
            self.multi_year_var.set(str(year))
            self.multi_month_var.set(f"{month:02d}")

        # 获取筛选条件
        shift_filter = self.multi_shift_var.get()

        # 更新月份年份显示
        self.multi_month_year_var.set(f"{year}年{month}月")

        # 渲染多人日历
        self._render_multi_calendar_grid(year, month, shift_filter)

        # 渲染班次图例
        self._render_multi_legend()

        # 更新状态栏
        self.update_status(f"多人日历已更新: {year}年{month}月")

    def _render_multi_calendar_grid(self, year, month, shift_filter):
        """渲染多人日历网格视图 - 性能优化版本"""
        # 性能优化：使用批量更新减少界面刷新
        try:
            # 暂时禁用界面更新，避免渲染过程中的闪烁
            self.multi_calendar_container.config(state='disabled')
        except:
            pass

        try:
            # 缓存容器尺寸，避免重复计算
            if not hasattr(self, '_cached_container_width'):
                self._cached_container_width = 0

            container_width = self.multi_calendar_container.winfo_width()
            if container_width < 100:
                container_width = 1000  # 默认宽度

            # 只有当宽度变化超过50px时才重新计算布局
            if abs(container_width - self._cached_container_width) > 50:
                self._cached_container_width = container_width
                cell_min_width = max(110, (container_width - 20) // 7)

                # 重新设置网格布局权重
                for col in range(7):
                    self.multi_calendar_container.columnconfigure(col, weight=1, minsize=cell_min_width)
        except Exception:
            cell_min_width = 120

        try:
            # 只在第一次渲染时设置行配置
            if not hasattr(self, '_multi_calendar_grid_initialized'):
                self._multi_calendar_grid_initialized = True
                # 星期标题行高度设置为40px，日期行高度设置为115px
                self.multi_calendar_container.rowconfigure(0, weight=0, minsize=40)  # 星期标题行
                for row in range(1, 7):  # 日期行 (第1-6行)
                    self.multi_calendar_container.rowconfigure(row, weight=1, minsize=115)

            # 只在第一次渲染时创建星期标题
            if not hasattr(self, '_multi_calendar_headers_created'):
                self._multi_calendar_headers_created = True
                # 星期标题 - 现代简约风格
                week_days = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
                week_colors = [self.colors['text_secondary']] * 5 + [self.colors['danger'], self.colors['danger']]

                for col, (day_name, day_color) in enumerate(zip(week_days, week_colors)):
                    header_frame = tk.Frame(self.multi_calendar_container,
                                           bg=self.colors['calendar_header'],
                                           relief='flat', borderwidth=0)
                    header_frame.grid(row=0, column=col, padx=1, pady=(0, 2), sticky='ew')

                    label = tk.Label(header_frame, text=day_name,
                                    font=('Microsoft YaHei UI', 10, 'bold'),
                                    bg=self.colors['calendar_header'],
                                    fg=day_color,
                                    anchor='center')
                    label.pack(fill=tk.BOTH, padx=4, pady=6)

            # 获取月份第一天和最后一天
            first_day = datetime.date(year, month, 1)
            days_in_month = calendar.monthrange(year, month)[1]

            # 计算第一周的起始位置 (周一为0)
            start_weekday = first_day.weekday()

            # 获取所有成员的排班数据（优化：只获取必要的数据）
            calendar_data = self.get_multi_member_calendar_data(year, month)

            # 应用筛选
            filtered_data = self._filter_multi_calendar_data(calendar_data, shift_filter)

            # 性能优化：缓存和复用日期格子控件
            if not hasattr(self, '_multi_calendar_cell_cache'):
                self._multi_calendar_cell_cache = {}

            # 生成缓存键（包含字体大小，确保字体变化时重新渲染）
            font_size = self.multi_calendar_font_size.get()
            cache_key = f"{year}_{month}_{shift_filter}_{font_size}"

            # 注意：移除了渲染缓存跳过逻辑，确保每次切换月份都能正确渲染
            # 数据缓存仍然保留（在get_multi_member_calendar_data中），性能影响很小

            # 缓存星期标题控件（必须在清空操作之前）
            if not hasattr(self, '_header_widgets'):
                self._header_widgets = []
                for widget in self.multi_calendar_container.winfo_children():
                    grid_info = widget.grid_info()
                    if grid_info and grid_info.get('row') == 0:  # 星期标题行
                        self._header_widgets.append(widget)

            # 清空现有日期格子（保留星期标题）
            for widget in self.multi_calendar_container.winfo_children():
                if isinstance(widget, tk.Frame) and widget.winfo_children() and widget not in self._header_widgets:
                    widget.destroy()

            # 渲染日期格子（批量更新）
            day_counter = 1
            today = datetime.date.today()
            cells_to_create = []

            for week in range(6):
                for weekday in range(7):
                    row = week + 1  # 日期行从第1行开始（星期标题在第0行）
                    col = weekday

                    if week == 0 and weekday < start_weekday:
                        # 空格子 - 更柔和的样式
                        cells_to_create.append(('empty', row, col, None, None, None, None))
                        continue

                    if day_counter > days_in_month:
                        # 空格子
                        cells_to_create.append(('empty', row, col, None, None, None, None))
                        continue

                    # 创建日期格子数据
                    cells_to_create.append(('cell', row, col, year, month, day_counter,
                                          filtered_data.get(day_counter, [])))
                    day_counter += 1

            # 批量创建格子以减少界面刷新次数
            # 注意：TFrame没有config方法，所以移除这行代码
            for cell_type, row, col, year, month, day, data in cells_to_create:
                if cell_type == 'empty':
                    empty_frame = tk.Frame(self.multi_calendar_container,
                                         bg=self.colors['bg_main'], relief='flat', borderwidth=0)
                    empty_frame.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                else:
                    self._create_multi_calendar_cell(
                        self.multi_calendar_container, row, col,
                        year, month, day, data, today
                    )
        finally:
            # 重新启用界面更新并强制刷新（无论是否出错）
            try:
                self.multi_calendar_container.config(state='normal')
            except:
                pass

            # 强制一次性更新所有控件
            self.multi_calendar_container.update_idletasks()

        # 注意：移除了渲染缓存的更新逻辑，因为已经不再使用渲染缓存

    def _get_data_hash(self, data):
        """计算数据的哈希值用于缓存比较"""
        import hashlib
        data_str = str(sorted(data.items()))
        return hashlib.md5(data_str.encode()).hexdigest()[:8]

    def get_multi_member_calendar_data(self, year, month):
        """获取指定年月的多人排班数据 - 性能优化版本"""
        # 性能优化：添加数据缓存
        if not hasattr(self, '_calendar_data_cache'):
            self._calendar_data_cache = {}

        cache_key = f"multi_{year}_{month}"

        # 检查缓存是否存在且数据未变化
        if cache_key in self._calendar_data_cache:
            cached_item = self._calendar_data_cache[cache_key]
            # 检查数据是否有更新
            current_data_hash = self._get_schedules_hash()
            if cached_item['data_hash'] == current_data_hash:
                return cached_item['data']

        calendar_data = {}

        # 获取指定月份的天数
        days_in_month = calendar.monthrange(year, month)[1]

        # 初始化每一天的数据
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            calendar_data[day] = []

        # 性能优化：预先过滤有效成员，避免遍历所有成员
        valid_members = {
            name: data for name, data in self.shift_schedules.items()
            if 'shifts' in data and data['shifts']
        }

        # 填入排班记录（排除休息）
        for member_name, member_data in valid_members.items():
            for date_str, shift_value in member_data['shifts'].items():
                try:
                    # 支持单个班次（字符串）或多个班次（列表）
                    shift_types = shift_value if isinstance(shift_value, list) else [shift_value]

                    for shift_type in shift_types:
                        # 只处理非休息和非常日班的班次
                        if shift_type in ["休息", "常日班"]:
                            continue

                        record_date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                        if record_date.year == year and record_date.month == month:
                            day = record_date.day

                            # 获取班次颜色信息
                            shift_color = "#E1E8ED"  # 默认颜色
                            if shift_type in self.shift_types:
                                shift_color = self.shift_types[shift_type].get('color', "#E1E8ED")

                            # 添加记录
                            calendar_data[day].append({
                                'member': member_name,
                                'shift': shift_type,
                                'color': shift_color
                            })
                except (ValueError, KeyError):
                    continue

        # 缓存结果
        self._calendar_data_cache[cache_key] = {
            'data': calendar_data,
            'data_hash': self._get_schedules_hash(),
            'timestamp': datetime.datetime.now()
        }

        # 清理旧缓存（保留最近10个）
        if len(self._calendar_data_cache) > 10:
            oldest_keys = sorted(self._calendar_data_cache.keys(),
                               key=lambda k: self._calendar_data_cache[k]['timestamp'])[:-10]
            for key in oldest_keys:
                del self._calendar_data_cache[key]

        return calendar_data

    def _get_schedules_hash(self):
        """计算排班数据的哈希值用于检测数据变化"""
        import hashlib
        # 只计算关键数据的哈希，避免计算整个数据结构
        hash_data = {
            'members_count': len(self.shift_schedules),
            'shifts_count': sum(len(data.get('shifts', {})) for data in self.shift_schedules.values()),
            'types_count': len(self.shift_types)
        }
        data_str = str(sorted(hash_data.items()))
        return hashlib.md5(data_str.encode()).hexdigest()[:8]

    def _filter_multi_calendar_data(self, calendar_data, shift_filter):
        """筛选多人日历数据"""
        if shift_filter == "全部班次":
            return calendar_data

        filtered_data = {}
        for day, records in calendar_data.items():
            filtered_records = []
            for record in records:
                if record['shift'] == shift_filter:
                    filtered_records.append(record)
            if filtered_records:
                filtered_data[day] = filtered_records

        return filtered_data

    def _create_tooltip(self, widget, text):
        """为控件创建工具提示

        Args:
            widget: 目标控件
            text (str): 提示文本
        """
        def on_enter(event):
            # 显示提示
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)  # 无边框窗口
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")  # 位置在鼠标右下角

            label = tk.Label(tooltip, text=text, justify=tk.LEFT,
                           background="#FFFFE0", relief=tk.SOLID, borderwidth=1,
                           font=('Microsoft YaHei UI', 9))
            label.pack()

            # 将tooltip对象存储到widget的引用中
            widget.tooltip = tooltip

        def on_leave(event):
            # 隐藏提示
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        # 绑定鼠标事件
        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def _get_leave_info(self, member_name, date_str):
        """获取指定成员在指定日期的休假信息

        Args:
            member_name (str): 成员名称
            date_str (str): 日期字符串 (YYYY-MM-DD)

        Returns:
            dict or None: 休假信息，如果没有休假则返回None
        """
        for record in self.leave_records:
            if (record.get('plan_name') == member_name and
                record.get('date') == date_str):
                return {
                    'type': record.get('type', '未知'),
                    'note': record.get('note', '')
                }
        return None

    def _get_contrast_color(self, bg_color):
        """根据背景色选择对比色（黑色或白色）"""
        # 移除#号并转换为RGB
        bg_color = bg_color.lstrip('#')
        if len(bg_color) == 3:
            # 处理简写格式如 #FFF
            bg_color = ''.join([c*2 for c in bg_color])

        try:
            r = int(bg_color[0:2], 16)
            g = int(bg_color[2:4], 16)
            b = int(bg_color[4:6], 16)
        except ValueError:
            return '#000000'  # 默认黑色

        # 计算亮度
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255

        # 如果背景色较亮，使用黑色文字；否则使用白色文字
        return '#000000' if luminance > 0.5 else '#FFFFFF'

    def _create_multi_calendar_cell(self, parent, row, col, year, month, day, shift_records, today):
        """创建多人日历的单个日期格子 - 现代卡片风格"""
        # 获取多人日历格子字体大小设置
        mc_font_size = self.multi_calendar_font_size.get()

        # 判断是否为今天
        is_today = (today.year == year and today.month == month and today.day == day)

        # 判断是否为周末
        current_date = datetime.date(year, month, day)
        is_weekend = current_date.weekday() >= 5  # 周六、周日

        # 设置背景色和边框色
        if is_today:
            bg_color = self.colors['today_bg']
            border_color = self.colors['today_border']
            border_width = 2
        elif is_weekend:
            bg_color = self.colors['weekend_bg']
            border_color = self.colors['border_light']
            border_width = 1
        else:
            bg_color = self.colors['bg_card']
            border_color = self.colors['border_light']
            border_width = 1

        # 外层边框框架
        border_frame = tk.Frame(parent, bg=border_color)
        border_frame.grid(row=row, column=col, padx=2, pady=2, sticky='nsew')

        # 格子框架 - 卡片风格
        cell_frame = tk.Frame(border_frame, bg=bg_color, relief='flat', borderwidth=0)
        cell_frame.pack(fill=tk.BOTH, expand=True, padx=border_width, pady=border_width)

        # 配置格子内部权重
        cell_frame.rowconfigure(0, weight=0)  # 日期头部
        cell_frame.rowconfigure(1, weight=1)  # 排班信息
        cell_frame.columnconfigure(0, weight=1)

        # ============ 日期头部区域 ============
        date_header = tk.Frame(cell_frame, bg=bg_color)
        date_header.grid(row=0, column=0, padx=6, pady=(6, 2), sticky='ew')
        date_header.columnconfigure(0, weight=0)
        date_header.columnconfigure(1, weight=1)

        # 今日标签（如果是今天）
        if is_today:
            today_badge = tk.Frame(date_header, bg=self.colors['secondary'])
            today_badge.grid(row=0, column=0, sticky='w', padx=(0, 6))
            today_label = tk.Label(today_badge, text="今日",
                                  bg=self.colors['secondary'], fg=self.colors['white'],
                                  font=('Microsoft YaHei UI', max(6, mc_font_size - 1), 'bold'),
                                  padx=6, pady=1)
            today_label.pack()

        # 日期数字标签
        if is_today:
            date_color = self.colors['secondary']
            date_font = ('Microsoft YaHei UI', mc_font_size + 5, 'bold')
        elif is_weekend:
            date_color = self.colors['danger']
            date_font = ('Microsoft YaHei UI', mc_font_size + 3, 'normal')
        else:
            date_color = self.colors['text_primary']
            date_font = ('Microsoft YaHei UI', mc_font_size + 3, 'normal')

        # 创建日期和节假日容器
        date_container = tk.Frame(date_header, bg=bg_color)
        date_container.grid(row=0, column=1, sticky='e')

        date_label = tk.Label(date_container, text=str(day),
                             bg=bg_color, fg=date_color,
                             font=date_font, anchor='e')
        date_label.pack(side=tk.LEFT)

        # 检查是否是节假日
        year_str = str(year)
        month_day_str = f"{month:02d}-{day:02d}"
        if year_str in self.holidays and month_day_str in self.holidays[year_str]:
            holiday_name = self.holidays[year_str][month_day_str]
            holiday_badge = tk.Frame(date_container, bg=self.colors['danger'])
            holiday_badge.pack(side=tk.LEFT, padx=(4, 0))
            holiday_label = tk.Label(holiday_badge, text=holiday_name[:2],
                                    bg=self.colors['danger'], fg=self.colors['white'],
                                    font=('Microsoft YaHei UI', max(6, mc_font_size - 2), 'bold'),
                                    padx=3, pady=0)
            holiday_label.pack()

        # ============ 排班信息区域 ============
        shift_frame = tk.Frame(cell_frame, bg=bg_color)
        shift_frame.grid(row=1, column=0, padx=4, pady=(0, 4), sticky='nsew')

        # 对排班记录进行排序：白班在前，夜班在后，其他班次按字母顺序
        def sort_shifts(record):
            shift_type = record['shift']
            if shift_type == '白班':
                return (0, shift_type)
            elif shift_type == '夜班':
                return (1, shift_type)
            else:
                return (2, shift_type)

        sorted_shift_records = sorted(shift_records, key=sort_shifts)

        # 显示排班记录（最多显示4条）
        display_records = sorted_shift_records[:4]
        for i, record in enumerate(display_records):
            # 检查该成员在这一天是否有休假
            current_date_str = f"{year}-{month:02d}-{day:02d}"

            # 创建单条排班记录的容器
            record_frame = tk.Frame(shift_frame, bg=bg_color)
            record_frame.pack(fill=tk.X, pady=(2, 0))

            leave_info = self._get_leave_info(record['member'], current_date_str)

            # 班次类型标签 - 药丸形状风格
            text_color = self._get_contrast_color(record['color'])
            shift_badge = tk.Frame(record_frame, bg=record['color'])
            shift_badge.pack(side=tk.LEFT, padx=(0, 4))

            shift_text = record['shift']
            if len(shift_text) > 3:
                shift_text = shift_text[:2] + '…'

            shift_label = tk.Label(shift_badge, text=shift_text,
                                 bg=record['color'], fg=text_color,
                                 font=('Microsoft YaHei UI', max(6, mc_font_size - 1), 'bold'),
                                 padx=5, pady=1)
            shift_label.pack()

            # 成员名称
            member_text = record['member']
            if len(member_text) > 4:
                member_text = member_text[:3] + '…'

            name_label = tk.Label(record_frame, text=member_text,
                                 bg=bg_color, fg=self.colors['text_secondary'],
                                 font=('Microsoft YaHei UI', mc_font_size),
                                 anchor='w')
            name_label.pack(side=tk.LEFT)

            # 如果有调换班记录，显示"调"字标签（只检查该班次是否被调换）
            if self.check_swap_record(record['member'], current_date_str, record['shift']):
                swap_badge = tk.Frame(record_frame, bg=self.colors['warning'])
                swap_badge.pack(side=tk.LEFT, padx=(4, 0))

                swap_label = tk.Label(swap_badge, text="调",
                                     bg=self.colors['warning'], fg=self.colors['white'],
                                     font=('Microsoft YaHei UI', max(6, mc_font_size - 2), 'bold'),
                                     padx=4, pady=0)
                swap_label.pack()

            # 如果有休假，显示休假标签
            if leave_info:
                leave_badge = tk.Frame(record_frame, bg=self.colors['danger'])
                leave_badge.pack(side=tk.LEFT, padx=(4, 0))

                leave_label = tk.Label(leave_badge, text="假",
                                     bg=self.colors['danger'], fg=self.colors['white'],
                                     font=('Microsoft YaHei UI', max(6, mc_font_size - 2), 'bold'),
                                     padx=4, pady=0)
                leave_label.pack()

                # 绑定鼠标悬停提示
                tooltip_text = f"休假类型: {leave_info['type']}"
                if leave_info.get('note'):
                    tooltip_text += f"\n备注: {leave_info['note']}"
                self._create_tooltip(leave_badge, tooltip_text)

            # 为 record_frame 及其所有子控件绑定右键菜单
            member_name = record['member']
            self._bind_context_menu_recursive(record_frame, member_name, current_date_str)

        # 如果还有更多记录，显示省略号
        if len(sorted_shift_records) > 4:
            more_frame = tk.Frame(shift_frame, bg=bg_color)
            more_frame.pack(fill=tk.X, pady=(2, 0))

            more_label = tk.Label(more_frame, text=f"+{len(sorted_shift_records)-4} 更多",
                                 bg=bg_color, fg=self.colors['text_muted'],
                                 font=('Microsoft YaHei UI', max(6, mc_font_size - 1)),
                                 anchor='w')
            more_label.pack(side=tk.LEFT)

    def _render_multi_legend(self):
        """渲染班次图例"""
        # 清空现有图例
        for widget in self.multi_legend_frame.winfo_children():
            widget.destroy()

        # 按班次优先级排序（排除休息和常日班）：白班 -> 夜班 -> 其他
        exclude_shifts = ["休息", "常日班"]
        all_shifts = [item for item in self.shift_types.items() if item[0] not in exclude_shifts]

        def sort_legend_items(item):
            shift_type = item[0]
            if shift_type == '白班':
                return (0, shift_type)
            elif shift_type == '夜班':
                return (1, shift_type)
            else:
                return (2, shift_type)

        sorted_shifts = sorted(all_shifts, key=sort_legend_items)

        # 创建图例项
        for shift_type, shift_info in sorted_shifts:
            legend_item = tk.Frame(self.multi_legend_frame)
            legend_item.pack(side=tk.LEFT, padx=(0, 15))

            # 颜色块
            color_canvas = tk.Canvas(legend_item, width=16, height=12,
                                   highlightthickness=1, highlightbackground="#CCCCCC")
            color_canvas.create_rectangle(1, 1, 15, 11, fill=shift_info.get('color', '#FFFFFF'), outline="")
            color_canvas.pack(side=tk.LEFT, padx=(0, 5))

            # 类型名称
            type_label = ttk.Label(legend_item, text=shift_type,
                                 font=('Microsoft YaHei UI', 9))
            type_label.pack(side=tk.LEFT)

    def _format_member_name(self, name):
        """格式化成员姓名，为2字姓名添加空格，调整3字姓名的对齐"""
        if not isinstance(name, str):
            return name

        name = name.strip()
        name_length = len(name)

        if name_length == 2:
            # 2字姓名：中间加2个空格，总长度达到4
            return f"{name[0]}  {name[1]}"
        elif name_length == 3:
            # 3字姓名：保持原样，但在显示时会减少与冒号间的空格
            return name
        else:
            # 其他长度：保持原样
            return name
        self.holiday_year_combo['values'] = years

        # 更新月份选项
        months = [f"{month:02d}" for month in range(1, 13)]
        self.holiday_month_combo['values'] = months

        # 更新成员选项
        members = ["全部成员"] + self.get_all_members_for_holiday_calendar()
        self.holiday_member_combo['values'] = members

        # 更新请假类型选项
        leave_types = ["全部类型"] + self.get_leave_types_for_holiday_calendar()
        self.holiday_leave_type_combo['values'] = leave_types

    # ==================== 当月休假日历控制方法 ====================

    def holiday_calendar_prev_month(self):
        """当月休假日历显示上个月"""
        try:
            year = int(self.holiday_year_var.get())
            month = int(self.holiday_month_var.get())
        except ValueError:
            year = self.current_date.year
            month = self.current_date.month

        # 计算上个月
        if month == 1:
            year -= 1
            month = 12
        else:
            month -= 1

        self.holiday_year_var.set(str(year))
        self.holiday_month_var.set(f"{month:02d}")
        self.update_holiday_calendar()

    def holiday_calendar_next_month(self):
        """当月休假日历显示下个月"""
        try:
            year = int(self.holiday_year_var.get())
            month = int(self.holiday_month_var.get())
        except ValueError:
            year = self.current_date.year
            month = self.current_date.month

        # 计算下个月
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1

        self.holiday_year_var.set(str(year))
        self.holiday_month_var.set(f"{month:02d}")
        self.update_holiday_calendar()

    def holiday_calendar_show_current_month(self):
        """当月休假日历显示当前月份"""
        today = datetime.date.today()
        self.holiday_year_var.set(str(today.year))
        self.holiday_month_var.set(f"{today.month:02d}")
        self.update_holiday_calendar()

    def _init_holiday_calendar_controls(self):
        """初始化当月休假日历控件选项"""
        # 更新年份选项（当前年份前后5年）
        current_year = self.current_date.year
        years = [str(year) for year in range(current_year - 5, current_year + 6)]
        self.holiday_year_combo['values'] = years

        # 更新月份选项
        months = [f"{month:02d}" for month in range(1, 13)]
        self.holiday_month_combo['values'] = months

        # 更新成员选项
        members = ["全部成员"] + self.get_all_members_for_holiday_calendar()
        self.holiday_member_combo['values'] = members

        # 更新请假类型选项
        leave_types = ["全部类型"] + self.get_leave_types_for_holiday_calendar()
        self.holiday_leave_type_combo['values'] = leave_types

    def _init_multi_calendar_delayed(self):
        """延迟初始化多人日历，避免在创建时调用update_status"""
        try:
            # 更新月份年份显示
            today = datetime.date.today()
            self.multi_month_year_var.set(f"{today.year}年{today.month:02d}月")

            # 初始化日历（但不调用update_status）
            self._render_multi_calendar_grid(today.year, today.month, "全部班次")
            self._render_multi_legend()

            # 注意：不再在这里设置 _multi_calendar_rendered，由调用者设置
        except Exception as e:
            # 如果出错，至少确保标签页能创建
            pass

    def export_monthly_attendance(self):
        """导出当月考勤到Excel"""
        try:
            import xlwt
            import xlrd
            from tkinter import filedialog
            from xlutils.copy import copy as xl_copy

            # 获取当前选择的年月
            year = int(self.multi_year_var.get())
            month = int(self.multi_month_var.get())

            # 获取月份天数
            days_in_month = calendar.monthrange(year, month)[1]

            # 读取模板文件
            template_path = os.path.join(os.path.dirname(__file__), '运行一部外协员工2025年11月考勤.xls')
            if not os.path.exists(template_path):
                messagebox.showerror("错误", "找不到模板文件：运行一部外协员工2025年11月考勤.xls")
                return

            # 打开模板文件
            template_workbook = xlrd.open_workbook(template_path, formatting_info=True)
            template_sheet = template_workbook.sheet_by_index(0)

            # 提取员工名单（从第4行开始，第2列是姓名）
            employee_list = []
            for row_idx in range(4, template_sheet.nrows):
                name_cell = template_sheet.cell(row_idx, 2)
                if name_cell.value and isinstance(name_cell.value, str) and name_cell.value.strip():
                    employee_list.append(name_cell.value.strip())
                else:
                    break  # 遇到空行停止

            if not employee_list:
                messagebox.showwarning("提示", "模板中没有找到员工名单")
                return

            # 获取所有员工的考勤数据
            raw_calendar_data = self.get_multi_member_calendar_data(year, month)

            # 转换数据结构：从 {day: [records]} 转换为 {member: {day: shift}}
            calendar_data = {}
            for day, records in raw_calendar_data.items():
                for record in records:
                    member_name = record['member']
                    shift_type = record['shift']
                    if member_name not in calendar_data:
                        calendar_data[member_name] = {}
                    calendar_data[member_name][day] = shift_type

            # 构建请假记录快速查找字典：{员工名: {日期: True}}
            leave_lookup = {}
            for record in self.leave_records:
                try:
                    leave_date = datetime.datetime.strptime(record['date'], "%Y-%m-%d")
                    if leave_date.year == year and leave_date.month == month:
                        member_name = record['plan_name']
                        day = leave_date.day
                        if member_name not in leave_lookup:
                            leave_lookup[member_name] = {}
                        leave_lookup[member_name][day] = True
                except (ValueError, KeyError):
                    continue

            # 同时获取所有成员（包括休息和常日班的）
            for member_name, member_data in self.shift_schedules.items():
                if member_name not in calendar_data:
                    calendar_data[member_name] = {}

                # 填充该成员在当月的所有排班
                if 'shifts' in member_data:
                    for date_str, shift_type in member_data['shifts'].items():
                        try:
                            record_date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                            if record_date.year == year and record_date.month == month:
                                day = record_date.day
                                if day not in calendar_data[member_name]:
                                    calendar_data[member_name][day] = shift_type
                        except (ValueError, KeyError):
                            continue

            # 复制模板工作簿
            workbook = xl_copy(template_workbook)
            sheet = workbook.get_sheet(0)

            # 定义数据样式（居中对齐）
            data_style = xlwt.XFStyle()
            data_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
            data_style.alignment.vert = xlwt.Alignment.VERT_CENTER
            borders = xlwt.Borders()
            borders.left = xlwt.Borders.THIN
            borders.right = xlwt.Borders.THIN
            borders.top = xlwt.Borders.THIN
            borders.bottom = xlwt.Borders.THIN
            data_style.borders = borders

            # 更新标题中的年月
            title_style = xlwt.XFStyle()
            title_font = xlwt.Font()
            title_font.name = '宋体'
            title_font.height = 180  # 9号字体
            title_style.font = title_font
            title_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
            sheet.write_merge(0, 0, 0, 36, f'{year}年{month}月外协员工考勤卡', title_style)

            # 更新表头中的日期（第2行，从第6列开始）
            header_style = xlwt.XFStyle()
            header_font = xlwt.Font()
            header_font.name = '宋体'
            header_font.height = 180  # 9号字体
            header_style.font = header_font
            header_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
            header_style.alignment.vert = xlwt.Alignment.VERT_CENTER
            header_style.borders = borders

            for day in range(1, days_in_month + 1):
                sheet.write(2, 5 + day, day, header_style)

            # 按模板顺序更新员工考勤数据
            for idx, member_name in enumerate(employee_list):
                row = 4 + idx

                # 获取该员工的排班数据
                shifts = calendar_data.get(member_name, {})

                # 只更新考勤数据列（从第6列开始，即1号）
                for day in range(1, days_in_month + 1):
                    shift = shifts.get(day, '')
                    # 检查是否有请假记录
                    has_leave = leave_lookup.get(member_name, {}).get(day, False)
                    if has_leave and shift in ['白班', '夜班']:
                        # 有请假记录且是白班或夜班，替换为"□"
                        shift = '□'
                    else:
                        # 简化班次名称
                        if shift == '休息':
                            shift = ''
                        elif shift == '常日班':
                            shift = '常'
                        elif shift == '白班':
                            shift = '白'
                        elif shift == '夜班':
                            shift = '夜'
                    sheet.write(row, 5 + day, shift, data_style)

            # 设置派遣三级部门列的列宽
            sheet.col(5).width = int(3.85 * 256)

            # 设置天数列的列宽（从第6列开始，即1号到31号）
            for day in range(1, days_in_month + 1):
                sheet.col(5 + day).width = int(3.38 * 256)

            # 保存文件
            default_filename = f'考勤表_{year}年{month:02d}月.xls'
            filepath = filedialog.asksaveasfilename(
                title="保存考勤表",
                defaultextension=".xls",
                initialfile=default_filename,
                filetypes=[("Excel文件", "*.xls"), ("所有文件", "*.*")]
            )

            if filepath:
                workbook.save(filepath)
                messagebox.showinfo("成功", f"考勤表已导出到:\n{filepath}\n\n已导出 {len(employee_list)} 名员工的考勤数据")
                self.update_status(f"已导出 {year}年{month}月 考勤表（{len(employee_list)}人）")

        except Exception as e:
            messagebox.showerror("错误", f"导出考勤表失败:\n{str(e)}")
            self.update_status(f"导出考勤表失败: {str(e)}")

    def holiday_calendar_on_year_selected(self):
        """年份选择事件处理"""
        self.update_holiday_calendar()

    def _bind_mousewheel(self, widget):
        """为组件绑定鼠标滚轮事件"""
        # 绑定到Canvas本身
        if isinstance(widget, tk.Canvas):
            # 为Canvas绑定滚轮事件
            def on_canvas_mousewheel(event):
                widget.yview_scroll(int(-1 * (event.delta / 120)), "units")

            def on_canvas_shift_mousewheel(event):
                widget.xview_scroll(int(-1 * (event.delta / 120)), "units")

            # 绑定到Canvas和它的所有子组件
            widget.bind("<MouseWheel>", on_canvas_mousewheel)
            widget.bind("<Shift-MouseWheel>", on_canvas_shift_mousewheel)

            # 为Canvas内部的Frame也绑定滚轮事件
            def bind_to_all_children(parent):
                for child in parent.winfo_children():
                    child.bind("<MouseWheel>", on_canvas_mousewheel)
                    child.bind("<Shift-MouseWheel>", on_canvas_shift_mousewheel)
                    # 递归绑定到子组件
                    if hasattr(child, 'winfo_children') and child.winfo_children():
                        bind_to_all_children(child)

            # 延迟绑定，确保所有组件都已创建
            self.root.after(200, lambda: bind_to_all_children(self.holiday_calendar_container))

        # 为Text组件绑定滚轮事件
        elif isinstance(widget, tk.Text):
            def on_text_mousewheel(event):
                widget.yview_scroll(int(-1 * (event.delta / 120)), "units")

            widget.bind("<MouseWheel>", on_text_mousewheel)
            # 也绑定到父框架以确保滚轮事件能被捕获
            widget.master.bind("<MouseWheel>", on_text_mousewheel)

    def _format_member_name(self, name):
        """格式化成员姓名，为2字姓名添加空格，调整3字姓名的对齐"""
        if not isinstance(name, str):
            return name

        name = name.strip()
        name_length = len(name)

        if name_length == 2:
            # 2字姓名：中间加2个空格，总长度达到4
            return f"{name[0]}  {name[1]}"
        elif name_length == 3:
            # 3字姓名：保持原样，但在显示时会减少与冒号间的空格
            return name
        else:
            # 其他长度：保持原样
            return name

    def _optimize_holiday_calendar_display(self):
        """优化当月休假日历显示到最佳尺寸"""
        # 使用统一的自动适应方法
        self._auto_fit_calendar_display()

    def _on_holiday_calendar_configure(self, event=None):
        """当日历容器内容变化时更新滚动区域"""
        # 更新Canvas的滚动区域以包含所有内容
        self.holiday_calendar_canvas.configure(
            scrollregion=self.holiday_calendar_canvas.bbox('all')
        )

    def _on_canvas_configure(self, event=None):
        """当Canvas尺寸变化时调整内部框架宽度"""
        # 获取Canvas的可见宽度
        canvas_width = event.width

        # 如果内容宽度小于Canvas宽度，则调整内容框架宽度
        if canvas_width > 1:  # 确保Canvas有有效宽度
            # 更新内部框架的宽度配置
            self.holiday_calendar_canvas.itemconfig(
                self.holiday_calendar_canvas_window,
                width=canvas_width
            )

    def holiday_calendar_on_month_selected(self):
        """月份选择事件处理"""
        self.update_holiday_calendar()

    def update_holiday_calendar(self):
        """更新当月休假日历显示"""
        try:
            # 获取当前选择的年月
            year = int(self.holiday_year_var.get())
            month = int(self.holiday_month_var.get())
        except ValueError:
            today = datetime.date.today()
            year = today.year
            month = today.month
            self.holiday_year_var.set(str(year))
            self.holiday_month_var.set(f"{month:02d}")

        # 获取筛选条件
        member_filter = self.holiday_member_var.get()
        leave_type_filter = self.holiday_leave_type_var.get()

        # 更新月份年份显示
        self.holiday_month_year_var.set(f"{year}年{month}月")

        # 获取日历数据
        calendar_data = self.get_holiday_calendar_data(year, month)

        # 应用筛选
        filtered_data = self._filter_calendar_data(calendar_data, member_filter, leave_type_filter)

        # 渲染日历
        self._render_holiday_calendar_grid(year, month, filtered_data)

        # 更新统计信息
        self._update_holiday_statistics(year, month, member_filter, leave_type_filter)

        # 渲染图例
        self._render_holiday_legend()

        # 更新状态栏
        self.update_status(f"当月休假日历已更新: {year}年{month}月")

    def _filter_calendar_data(self, calendar_data, member_filter, leave_type_filter):
        """筛选日历数据"""
        filtered_data = {}

        for day, records in calendar_data.items():
            filtered_records = []

            for record in records:
                # 成员筛选
                if member_filter != "全部成员" and record['name'] != member_filter:
                    continue

                # 请假类型筛选
                if leave_type_filter != "全部类型" and record['type'] != leave_type_filter:
                    continue

                filtered_records.append(record)

            if filtered_records:
                filtered_data[day] = filtered_records

        return filtered_data

    def _render_holiday_calendar_grid(self, year, month, calendar_data):
        """渲染网格化日历视图 - 现代清新风格"""
        # 清空现有日历
        for widget in self.holiday_calendar_container.winfo_children():
            widget.destroy()

        # 动态计算每个格子的尺寸以适应窗口
        try:
            # 获取可用宽度（考虑右侧统计面板）
            canvas_width = self.holiday_calendar_canvas.winfo_width()
            if canvas_width < 100:
                canvas_width = 770  # 默认宽度

            # 计算每列最小宽度
            cell_min_width = max(100, (canvas_width - 20) // 7)
        except Exception:
            cell_min_width = 110

        # 设置网格布局权重
        for col in range(7):
            self.holiday_calendar_container.columnconfigure(col, weight=1, minsize=cell_min_width)

        # 星期标题行高度设置为40px，日期行高度设置为95px（稍微增大以确保显示完整）
        self.holiday_calendar_container.rowconfigure(0, weight=0, minsize=40)  # 星期标题行
        for row in range(1, 7):  # 日期行 (第1-6行)
            self.holiday_calendar_container.rowconfigure(row, weight=1, minsize=95)

        # 星期标题 - 现代简约风格
        week_days = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        week_colors = [self.colors['text_secondary']] * 5 + [self.colors['danger'], self.colors['danger']]

        for col, (day_name, day_color) in enumerate(zip(week_days, week_colors)):
            header_frame = tk.Frame(self.holiday_calendar_container,
                                   bg=self.colors['calendar_header'],
                                   relief='flat', borderwidth=0)
            header_frame.grid(row=0, column=col, padx=1, pady=(0, 2), sticky='ew')

            label = tk.Label(header_frame, text=day_name,
                            font=('Microsoft YaHei UI', 10, 'bold'),
                            bg=self.colors['calendar_header'],
                            fg=day_color,
                            anchor='center')
            label.pack(fill=tk.BOTH, padx=4, pady=6)

        # 获取月份第一天和最后一天
        first_day = datetime.date(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]

        # 计算第一周的起始位置 (周一为0)
        start_weekday = first_day.weekday()

        # 获取颜色映射
        color_mapping = self.get_leave_types_color_mapping()

        # 渲染日期格子
        day_counter = 1
        today = datetime.date.today()

        for week in range(6):
            for weekday in range(7):
                row = week + 1  # 日期行从第1行开始（星期标题在第0行）
                col = weekday

                if week == 0 and weekday < start_weekday:
                    # 空格子 - 更柔和的样式
                    empty_frame = tk.Frame(self.holiday_calendar_container,
                                         bg=self.colors['bg_main'], relief='flat', borderwidth=0)
                    empty_frame.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                    continue

                if day_counter > days_in_month:
                    # 空格子
                    empty_frame = tk.Frame(self.holiday_calendar_container,
                                         bg=self.colors['bg_main'], relief='flat', borderwidth=0)
                    empty_frame.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                    continue

                # 创建日期格子
                self._create_holiday_calendar_cell(
                    self.holiday_calendar_container, row, col,
                    year, month, day_counter,
                    calendar_data.get(day_counter, []),
                    color_mapping, today
                )

                day_counter += 1

    def _create_holiday_calendar_cell(self, parent, row, col, year, month, day, leave_records, color_mapping, today):
        """创建单个日期格子 - 现代卡片风格"""
        # 判断是否为今天
        is_today = (today.year == year and today.month == month and today.day == day)

        # 判断是否为周末
        current_date = datetime.date(year, month, day)
        is_weekend = current_date.weekday() >= 5  # 周六、周日

        # 设置背景色和边框色
        if is_today:
            bg_color = self.colors['today_bg']
            border_color = self.colors['today_border']
            border_width = 2
        elif is_weekend:
            bg_color = self.colors['weekend_bg']
            border_color = self.colors['border_light']
            border_width = 1
        else:
            bg_color = self.colors['bg_card']
            border_color = self.colors['border_light']
            border_width = 1

        # 外层边框框架（用于实现边框效果）
        border_frame = tk.Frame(parent, bg=border_color)
        border_frame.grid(row=row, column=col, padx=2, pady=2, sticky='nsew')

        # 格子框架 - 卡片风格
        cell_frame = tk.Frame(border_frame, bg=bg_color, relief='flat', borderwidth=0)
        cell_frame.pack(fill=tk.BOTH, expand=True, padx=border_width, pady=border_width)

        # 配置格子内部权重
        cell_frame.rowconfigure(0, weight=0)  # 日期头部
        cell_frame.rowconfigure(1, weight=1)  # 请假信息
        cell_frame.columnconfigure(0, weight=1)

        # ============ 日期头部区域 ============
        date_header = tk.Frame(cell_frame, bg=bg_color)
        date_header.grid(row=0, column=0, padx=6, pady=(6, 2), sticky='ew')
        date_header.columnconfigure(0, weight=0)
        date_header.columnconfigure(1, weight=1)

        # 今日标签（如果是今天）- 更精致的样式
        if is_today:
            today_badge = tk.Frame(date_header, bg=self.colors['secondary'])
            today_badge.grid(row=0, column=0, sticky='w', padx=(0, 6))
            today_label = tk.Label(today_badge, text="今日",
                                  bg=self.colors['secondary'], fg=self.colors['white'],
                                  font=('Microsoft YaHei UI', 8, 'bold'),
                                  padx=6, pady=1)
            today_label.pack()

        # 日期数字标签 - 更大更醒目
        if is_today:
            date_color = self.colors['secondary']
            date_font = ('Microsoft YaHei UI', 14, 'bold')
        elif is_weekend:
            date_color = self.colors['danger']
            date_font = ('Microsoft YaHei UI', 12, 'normal')
        else:
            date_color = self.colors['text_primary']
            date_font = ('Microsoft YaHei UI', 12, 'normal')

        date_label = tk.Label(date_header, text=str(day),
                             bg=bg_color, fg=date_color,
                             font=date_font, anchor='e')
        date_label.grid(row=0, column=1, sticky='e')

        # ============ 请假信息区域 ============
        leave_frame = tk.Frame(cell_frame, bg=bg_color)
        leave_frame.grid(row=1, column=0, padx=4, pady=(0, 4), sticky='nsew')

        # 对请假记录按班次类型排序：白班在前，夜班在后，其他班次按字母顺序，无班次的排最后
        def sort_by_shift(record):
            shift_type = record.get('shift')
            if shift_type is None:
                return (3, '')  # 无班次排最后
            elif shift_type == '白班':
                return (0, shift_type)
            elif shift_type == '夜班':
                return (1, shift_type)
            else:
                return (2, shift_type)

        sorted_records = sorted(leave_records, key=sort_by_shift)

        # 显示请假记录（最多显示3条，超出显示...）
        display_records = sorted_records[:3]
        for i, record in enumerate(display_records):
            # 请假类型颜色
            leave_color = color_mapping.get(record['type'], self.colors['text_muted'])

            # 创建单条请假记录的容器
            record_frame = tk.Frame(leave_frame, bg=bg_color)
            record_frame.pack(fill=tk.X, pady=(2, 0))

            # 排班类型标签（如果有）
            shift_type = record.get('shift')
            if shift_type:
                shift_color = record.get('shift_color', '#E1E8ED')
                text_color = self._get_contrast_color(shift_color)

                shift_text = shift_type
                if len(shift_text) > 2:
                    shift_text = shift_text[:2]

                shift_badge = tk.Frame(record_frame, bg=shift_color)
                shift_badge.pack(side=tk.LEFT, padx=(0, 2))

                shift_label = tk.Label(shift_badge, text=shift_text,
                                      bg=shift_color, fg=text_color,
                                      font=('Microsoft YaHei UI', 7, 'bold'),
                                      padx=3, pady=0)
                shift_label.pack()

            # 人员名称
            name_text = record['name']
            if len(name_text) > 3:
                name_text = name_text[:2] + '…'

            # 人员标签
            name_label = tk.Label(record_frame, text=name_text,
                                 bg=bg_color, fg=self.colors['text_secondary'],
                                 font=('Microsoft YaHei UI', 8),
                                 anchor='w')
            name_label.pack(side=tk.LEFT)

        # 如果还有更多记录，显示省略号
        if len(sorted_records) > 3:
            more_frame = tk.Frame(leave_frame, bg=bg_color)
            more_frame.pack(fill=tk.X, pady=(2, 0))

            more_label = tk.Label(more_frame, text=f"+{len(sorted_records)-3} 更多",
                                 bg=bg_color, fg=self.colors['text_muted'],
                                 font=('Microsoft YaHei UI', 8),
                                 anchor='w')
            more_label.pack(side=tk.LEFT)

        # 添加 tooltip 显示详细信息
        if leave_records:
            tooltip_text = f"📅 {year}年{month}月{day}日\n"
            tooltip_text += "─" * 16 + "\n"
            for record in sorted_records:
                shift_info = f"[{record.get('shift', '无班次')}] " if record.get('shift') else ""
                tooltip_text += f"• {shift_info}{record['name']}: {record['type']}"
                if record.get('note'):
                    tooltip_text += f"\n  备注: {record['note']}"
                tooltip_text += "\n"
            _SimpleTooltip(cell_frame, tooltip_text.strip())

    def _update_holiday_statistics(self, year, month, member_filter, leave_type_filter):
        """更新休假统计信息"""
        # 清空统计文本
        self.holiday_stats_text.delete('1.0', tk.END)

        # 获取统计数据 - 现在包含月份筛选
        if member_filter == "全部成员":
            stats = self.get_holiday_statistics(year=year, month=month, leave_type=leave_type_filter)
        else:
            stats = self.get_holiday_statistics(member_name=member_filter, year=year, month=month, leave_type=leave_type_filter)

        total_days = stats['total_days']
        records_by_type = stats['records_by_type']
        records_by_member = stats['records_by_member']

        # 格式化统计信息
        stats_text = f"{'='*25}\n"
        stats_text += f"   {year}年{month}月 休假统计\n"
        stats_text += f"{'='*25}\n\n"

        # 总体统计
        filter_desc = []
        if member_filter != "全部成员":
            filter_desc.append(f"成员: {member_filter}")
        if leave_type_filter != "全部类型":
            filter_desc.append(f"类型: {leave_type_filter}")

        if filter_desc:
            stats_text += f"筛选条件: {', '.join(filter_desc)}\n"

        stats_text += f"\n总请假天数: {total_days} 天\n\n"

        # 按类型统计
        if records_by_type:
            stats_text += "【按类型统计】\n"
            stats_text += "-" * 20 + "\n"
            for leave_type, days in sorted(records_by_type.items(), key=lambda x: x[1], reverse=True):
                bar_length = min(days * 2, 15)  # 简单的文本条形图
                bar = '█' * bar_length
                stats_text += f"{leave_type:6s}: {days:2d}天 {bar}\n"
            stats_text += "\n"

        # 按成员统计
        if records_by_member and member_filter == "全部成员":
            stats_text += "【按成员统计】\n"
            stats_text += "-" * 20 + "\n"
            for member, days in sorted(records_by_member.items(), key=lambda x: x[1], reverse=True):
                # 格式化成员姓名
                formatted_member = self._format_member_name(member)
                bar_length = min(days * 2, 15)
                bar = '█' * bar_length

                # 根据原姓名长度决定与冒号之间的间距
                original_name = member.strip()
                if len(original_name) == 2:
                    # 2字姓名：格式化后是4个字符，正常间距
                    stats_text += f"{formatted_member} : {days:2d}天 {bar}\n"
                elif len(original_name) == 3:
                    # 3字姓名：减少与冒号之间的间距
                    stats_text += f"{formatted_member}: {days:2d}天 {bar}\n"
                else:
                    # 其他长度：正常间距
                    stats_text += f"{formatted_member}: {days:2d}天 {bar}\n"
            stats_text += "\n"

        # 如果没有数据
        if total_days == 0:
            stats_text += "本月暂无请假记录\n"

        stats_text += f"{'='*25}\n"

        # 显示统计信息
        self.holiday_stats_text.insert('1.0', stats_text)

    def _render_holiday_legend(self):
        """渲染请假类型图例"""
        # 清空现有图例
        for widget in self.holiday_legend_frame.winfo_children():
            widget.destroy()

        # 获取颜色映射
        color_mapping = self.get_leave_types_color_mapping()

        # 按类型名称排序
        sorted_types = sorted(color_mapping.items())

        # 创建图例项
        for leave_type, color in sorted_types:
            legend_item = tk.Frame(self.holiday_legend_frame)
            legend_item.pack(side=tk.LEFT, padx=(0, 15))

            # 颜色块
            color_canvas = tk.Canvas(legend_item, width=16, height=12,
                                   highlightthickness=1, highlightbackground="#CCCCCC")
            color_canvas.create_rectangle(1, 1, 15, 11, fill=color, outline="")
            color_canvas.pack(side=tk.LEFT, padx=(0, 5))

            # 类型名称
            type_label = ttk.Label(legend_item, text=leave_type,
                                 font=('Microsoft YaHei UI', 9))
            type_label.pack(side=tk.LEFT)

    def _render_legend(self):
        """渲染右下角图例。"""
        for w in self.legend_frame.winfo_children():
            w.destroy()
        wrap = ttk.Frame(self.legend_frame)
        wrap.pack(side=tk.LEFT)
        # 班次图例
        for name, info in self.shift_types.items():
            item = ttk.Frame(wrap)
            item.pack(side=tk.LEFT, padx=8)
            swatch = tk.Canvas(item, width=16, height=12, highlightthickness=1, highlightbackground="#888")
            swatch.create_rectangle(1, 1, 15, 11, fill=info.get("color", "#FFFFFF"), outline="")
            swatch.pack(side=tk.LEFT)
            ttk.Label(item, text=name).pack(side=tk.LEFT, padx=4)
        # 其他标识
        other = ttk.Frame(self.legend_frame)
        other.pack(side=tk.LEFT, padx=16)
        ttk.Label(other, text="节假日: 红字/红条").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(other, text="请假: 红底白字").pack(side=tk.LEFT)

    def update_year_options(self):
        """根据当前计划或默认范围更新年份下拉选项"""
        try:
            years = set()
            if self.current_schedule and isinstance(self.current_schedule.get("shifts"), dict):
                for date_str in self.current_schedule["shifts"].keys():
                    try:
                        y = int(date_str.split('-')[0])
                        years.add(y)
                    except Exception:
                        continue
            if not years:
                y = self.current_date.year
                years = set(range(y - 2, y + 6))  # 默认当前年-2 到 +5
            values = sorted(years)
            if hasattr(self, 'year_combo'):
                self.year_combo["values"] = values
                # 同步选择当前年
                cy = self.current_date.year
                if cy not in values and values:
                    cy = values[0]
                    self.current_date = datetime.date(cy, self.current_date.month, 1)
                self.year_var.set(str(cy))
        except Exception:
            pass

    def on_year_selected(self):
        """切换年份后刷新日历"""
        try:
            y = int(self.year_var.get())
        except Exception:
            return
        self.current_date = datetime.date(y, self.current_date.month, 1)
        self.update_calendar()
        self.sync_month_combo()

    def sync_year_combo(self):
        """在翻月或其他操作后同步年份下拉"""
        if not hasattr(self, 'year_combo'):
            return
        try:
            y = self.current_date.year
            values = list(self.year_combo["values"]) if self.year_combo["values"] else []
            values = [int(v) for v in values] if values else []
            if y not in values:
                self.update_year_options()
            else:
                self.year_var.set(str(y))
        except Exception:
            pass

    def update_month_options(self):
        """更新月份下拉为1-12并选中当前月"""
        if hasattr(self, 'month_combo'):
            months = [str(i) for i in range(1, 13)]
            self.month_combo["values"] = months
            self.month_var.set(str(self.current_date.month))

    def on_month_selected(self):
        """切换月份后刷新日历"""
        try:
            m = int(self.month_var.get())
        except Exception:
            return
        y = self.current_date.year
        self.current_date = datetime.date(y, m, 1)
        self.update_calendar()

    def sync_month_combo(self):
        """在翻月或其他操作后同步月份下拉"""
        if not hasattr(self, 'month_combo'):
            return
        try:
            self.month_var.set(str(self.current_date.month))
        except Exception:
            pass

    def setup_leave_tab(self):
        """请假管理标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="🔥 请假管理")

        # 左侧：请假类型
        left = ttk.LabelFrame(frame, text="请假类型", padding=10)
        left.grid(row=0, column=0, sticky=tk.NSEW, padx=10, pady=10)
        self.leave_type_list = tk.Listbox(left, height=8)
        self.leave_type_list.grid(row=0, column=0, columnspan=3, sticky=tk.NSEW)
        ttk.Button(left, text="添加类型", command=self.add_leave_type).grid(row=1, column=0, pady=6, sticky=tk.W)
        ttk.Button(left, text="删除类型", command=self.delete_leave_type).grid(row=1, column=1, pady=6, sticky=tk.W)
        self.update_leave_type_list()
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)

        # 右侧：请假记录
        right = ttk.LabelFrame(frame, text="请假记录", padding=10)
        right.grid(row=0, column=1, sticky=tk.NSEW, padx=10, pady=10)
        # 控件区
        ttk.Label(right, text="人员名称").grid(row=0, column=0, sticky=tk.W)
        plan_names = list(self.shift_schedules.keys())
        self.leave_plan_var = tk.StringVar(value=plan_names[0] if plan_names else "")
        self.leave_plan_combo = ttk.Combobox(right, textvariable=self.leave_plan_var, values=plan_names, state="readonly")
        self.leave_plan_combo.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(right, text="请假日期").grid(row=0, column=2, sticky=tk.W, padx=(50, 0))
        self.leave_date = DateEntry(right, date_pattern='yyyy-MM-dd')
        self.leave_date.set_date(datetime.date.today())
        self.leave_date.grid(row=0, column=2, sticky=tk.W, padx=(5, 0))

        ttk.Label(right, text="类型").grid(row=1, column=0, sticky=tk.W, pady=6)
        self.leave_type_var = tk.StringVar()
        self.leave_type_combo = ttk.Combobox(right, textvariable=self.leave_type_var, values=self.leave_types, state="readonly")
        if self.leave_types:
            self.leave_type_combo.current(0)
        self.leave_type_combo.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(right, text="备注").grid(row=1, column=2, sticky=tk.W, padx=(50, 0))
        self.leave_note_var = tk.StringVar()
        ttk.Entry(right, textvariable=self.leave_note_var, width=28).grid(row=1, column=2, sticky=tk.W, padx=(5, 0))

        ttk.Button(right, text="添加记录", command=self.add_leave_record, style="Small.TButton").grid(row=2, column=0, sticky=tk.W, pady=6)
        ttk.Button(right, text="删除选中记录", command=self.delete_leave_record, style="Small.TButton").grid(row=2, column=1, sticky=tk.W, pady=6)
        ttk.Button(right, text="编辑记录", command=self.edit_leave_record, style="Small.TButton").grid(row=2, column=1, sticky=tk.W, pady=6, padx=(80, 0))
        ttk.Button(right, text="查询记录", command=self.query_leave_records, style="Small.TButton").grid(row=2, column=1, sticky=tk.W, pady=6, padx=(160, 0))
        ttk.Button(right, text="查看全部", command=self.view_all_leave_records, style="Small.TButton").grid(row=2, column=1, sticky=tk.W, pady=6, padx=(240, 0))

        # 导入导出按钮区域
        import_export_frame = ttk.Frame(right)
        import_export_frame.grid(row=2, column=2, sticky=tk.W, pady=6)
        ttk.Button(import_export_frame, text="导入Excel", command=self.import_leave_records_from_excel, style="Small.TButton").pack(side=tk.LEFT, padx=(5, 2))
        ttk.Button(import_export_frame, text="导出Excel", command=self.export_leave_records_to_excel, style="Small.TButton").pack(side=tk.LEFT, padx=2)
        ttk.Button(import_export_frame, text="下载模板", command=self.download_import_template, style="Small.TButton").pack(side=tk.LEFT, padx=2)

        # 记录表格
        self.leave_tree = ttk.Treeview(right, columns=("plan", "date", "type", "note"), show="headings")
        self.leave_tree.heading("plan", text="人员名称")
        self.leave_tree.heading("date", text="请假日期")
        self.leave_tree.heading("type", text="类型")
        self.leave_tree.heading("note", text="备注")
        # 设置列宽度，确保合适的间距
        self.leave_tree.column("plan", width=100, minwidth=80)
        self.leave_tree.column("date", width=100, minwidth=80)
        self.leave_tree.column("type", width=80, minwidth=60)
        self.leave_tree.column("note", width=150, minwidth=100)
        self.leave_tree.grid(row=3, column=0, columnspan=4, sticky=tk.NSEW, pady=(6, 0))
        # 绑定双击事件
        self.leave_tree.bind("<Double-1>", lambda e: self.on_leave_tree_double_click(e))
        self.update_leave_tree()

        # 统计区域：每个计划的请假总天数
        stats = ttk.LabelFrame(frame, text="请假统计（各计划总请假天数与类型分布）", padding=10)
        stats.grid(row=1, column=0, columnspan=1, sticky=tk.NSEW, padx=10, pady=(0,10))

        # 工具栏：年份选择 - 放在树视图上方
        stats_toolbar = ttk.Frame(stats)
        stats_toolbar.pack(side=tk.TOP, fill=tk.X, pady=(0, 6))

        ttk.Label(stats_toolbar, text="统计年份:").pack(side=tk.LEFT, padx=(0, 5))

        # 创建年份变量
        self.leave_stats_year_var = tk.StringVar(value="全部")
        self.leave_stats_year_combo = ttk.Combobox(stats_toolbar, textvariable=self.leave_stats_year_var,
                                                     width=10, state="readonly")
        self.leave_stats_year_combo['values'] = ["全部", str(datetime.date.today().year)]  # 设置初始值
        self.leave_stats_year_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.leave_stats_year_combo.bind('<<ComboboxSelected>>', self.on_leave_stats_year_selected)

        # 初始化年份选项
        self.update_leave_stats_year_options()

        stats_tree_frame = ttk.Frame(stats)
        stats_tree_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.leave_stats_tree = ttk.Treeview(stats_tree_frame, columns=("plan", "days", "types"), show="headings")
        self.leave_stats_tree.heading("plan", text="人员名称")
        self.leave_stats_tree.heading("days", text="天数")
        self.leave_stats_tree.heading("types", text="类型分布")
        # 设置列宽度，确保合适的间距
        self.leave_stats_tree.column("plan", width=120, minwidth=100)
        self.leave_stats_tree.column("days", width=80, minwidth=60)
        self.leave_stats_tree.column("types", width=200, minwidth=150)
        self.leave_stats_tree.grid(row=0, column=0, sticky=tk.NSEW)
        stats_xscroll = ttk.Scrollbar(stats_tree_frame, orient=tk.HORIZONTAL, command=self.leave_stats_tree.xview)
        stats_xscroll.grid(row=1, column=0, sticky=tk.EW)
        self.leave_stats_tree.configure(xscrollcommand=stats_xscroll.set)
        stats_tree_frame.rowconfigure(0, weight=1)
        stats_tree_frame.columnconfigure(0, weight=1)
        self.update_leave_stats()

        # 右下角：年度配额与使用
        self.quota_label_frame = ttk.LabelFrame(frame, text="年度配额与使用", padding=10)
        self.quota_label_frame.grid(row=1, column=1, columnspan=1, sticky=tk.NSEW, padx=10, pady=(0,10))
        toolbar = ttk.Frame(self.quota_label_frame)
        toolbar.grid(row=0, column=0, sticky=tk.W, pady=(0,6))

        # 年份选择下拉框
        year_label = ttk.Label(toolbar, text="年份:")
        year_label.pack(side=tk.LEFT, padx=(0, 5))

        self.quota_year_combo = ttk.Combobox(toolbar, textvariable=self.quota_year_var,
                                           width=8, state="readonly")
        self.quota_year_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.quota_year_combo.bind('<<ComboboxSelected>>', self.on_quota_year_selected)

        # 初始化年份选项
        self.update_quota_year_options()

        # 醒目的当前年休假年度显示
        self.current_leave_year_label = ttk.Label(toolbar, text="",
                                                font=('Microsoft YaHei UI', 10, 'bold'),
                                                foreground=self.colors.get('accent', '#0078d4'))
        self.current_leave_year_label.pack(side=tk.LEFT, padx=(10, 0))
        self.update_current_leave_year_display()

        # 设置配额按钮
        ttk.Button(toolbar, text="设置配额", command=self.open_quota_setting).pack(side=tk.LEFT)

        # 年休假周期提示
        hint_label = ttk.Label(toolbar, text="提示: 年休假可延至次年3月，1-3月先扣上年余额",
                              font=('Microsoft YaHei UI', 8),
                              foreground=self.colors['text_muted'])
        hint_label.pack(side=tk.LEFT, padx=(15, 0))

        quota_tree_frame = ttk.Frame(self.quota_label_frame)
        quota_tree_frame.grid(row=1, column=0, sticky=tk.NSEW)
        self.leave_quota_tree = ttk.Treeview(quota_tree_frame, columns=("plan", "type", "year", "quota", "used", "remain"), show="headings")
        self.leave_quota_tree.heading("plan", text="人员名称")
        self.leave_quota_tree.heading("type", text="配额汇总")
        self.leave_quota_tree.heading("year", text="年份")
        self.leave_quota_tree.heading("quota", text="已用汇总")
        self.leave_quota_tree.heading("used", text="剩余汇总")
        self.leave_quota_tree.heading("remain", text="剩余总数")
        # 设置列宽度，确保合适的间距
        self.leave_quota_tree.column("plan", width=100, minwidth=80)
        self.leave_quota_tree.column("type", width=150, minwidth=120)
        self.leave_quota_tree.column("year", width=60, minwidth=50)
        self.leave_quota_tree.column("quota", width=120, minwidth=100)
        self.leave_quota_tree.column("used", width=120, minwidth=100)
        self.leave_quota_tree.column("remain", width=80, minwidth=60)
        self.leave_quota_tree.grid(row=0, column=0, sticky=tk.NSEW)
        # 添加底部横向滚动条
        quota_xscroll = ttk.Scrollbar(quota_tree_frame, orient=tk.HORIZONTAL, command=self.leave_quota_tree.xview)
        quota_xscroll.grid(row=1, column=0, sticky=tk.EW)
        self.leave_quota_tree.configure(xscrollcommand=quota_xscroll.set)
        # 绑定双击事件以编辑配额
        self.leave_quota_tree.bind("<Double-1>", self.on_quota_double_click)
        quota_tree_frame.rowconfigure(0, weight=1)
        quota_tree_frame.columnconfigure(0, weight=1)
        self.update_quota_summary()

        # 布局伸缩
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=3)
        frame.rowconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        right.columnconfigure(3, weight=1)
        right.rowconfigure(3, weight=1)
        stats.rowconfigure(0, weight=1)
        stats.columnconfigure(0, weight=1)
        self.quota_label_frame.rowconfigure(1, weight=1)
        self.quota_label_frame.columnconfigure(0, weight=1)

    def setup_holiday_tab(self):
        """节假日管理标签页"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="节假日管理")

        # 顶部：年份选择与操作按钮
        top = ttk.Frame(frame)
        top.grid(row=0, column=0, sticky=tk.EW, padx=10, pady=10)
        ttk.Label(top, text="年份").pack(side=tk.LEFT)
        self.holiday_year_var = tk.StringVar(value=str(self.current_date.year))
        self.holiday_year_combo = ttk.Combobox(top, textvariable=self.holiday_year_var, values=[str(y) for y in range(self.current_date.year-2, self.current_date.year+6)], state="readonly", width=8)
        self.holiday_year_combo.pack(side=tk.LEFT, padx=(6, 12))
        self.holiday_year_combo.bind('<<ComboboxSelected>>', lambda e: self.update_holiday_tree())

        ttk.Button(top, text="添加节日", command=self.add_holiday_dialog).pack(side=tk.LEFT)
        ttk.Button(top, text="编辑选中", command=self.edit_holiday_dialog).pack(side=tk.LEFT, padx=(6,0))
        ttk.Button(top, text="删除选中", command=self.delete_selected_holiday).pack(side=tk.LEFT, padx=(6,0))
        ttk.Button(top, text="填充法定假日", command=self.fill_national_holidays).pack(side=tk.LEFT, padx=(12,0))
        ttk.Button(top, text="添加春节", command=lambda: self.add_lunar_holiday("春节", 1, 1)).pack(side=tk.LEFT, padx=(6,0))
        ttk.Button(top, text="添加端午", command=lambda: self.add_lunar_holiday("端午节", 5, 5)).pack(side=tk.LEFT, padx=(6,0))
        ttk.Button(top, text="添加中秋", command=lambda: self.add_lunar_holiday("中秋节", 8, 15)).pack(side=tk.LEFT, padx=(6,0))

        # 列表
        self.holiday_tree = ttk.Treeview(frame, columns=("date", "name"), show="headings")
        self.holiday_tree.heading("date", text="日期(MM-DD)")
        self.holiday_tree.heading("name", text="节日名称")
        self.holiday_tree.grid(row=1, column=0, sticky=tk.NSEW, padx=10, pady=(0,10))
        self.update_holiday_tree()

        # 布局伸缩
        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)

    def update_holiday_tree(self):
        if not hasattr(self, 'holiday_tree'):
            return
        year = self.holiday_year_var.get() if hasattr(self, 'holiday_year_var') else str(self.current_date.year)
        self.holiday_tree.delete(*self.holiday_tree.get_children())
        items = []
        for md, name in self.holidays.get(year, {}).items():
            items.append((md, name))
        for md, name in sorted(items):
            self.holiday_tree.insert("", tk.END, values=(md, name))

    def add_holiday_dialog(self):
        top = tk.Toplevel(self.root)
        top.title("添加节日")
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(frm, text="年份").grid(row=0, column=0, sticky=tk.W, pady=4)
        year_var = tk.StringVar(value=self.holiday_year_var.get())
        year_combo = ttk.Combobox(frm, textvariable=year_var, values=self.holiday_year_combo["values"], state="readonly", width=10)
        year_combo.grid(row=0, column=1, sticky=tk.W)

        # 日期选择区域
        date_frame = ttk.LabelFrame(frm, text="日期选择", padding=5)
        date_frame.grid(row=1, column=0, columnspan=3, sticky=tk.NSEW, pady=4)

        # 左侧：日历选择器
        calendar_frame = ttk.Frame(date_frame)
        calendar_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 10))

        ttk.Label(calendar_frame, text="选择日期").pack(anchor=tk.W)
        self.holiday_calendar = Calendar(calendar_frame, date_pattern='yyyy-MM-dd',
                                       selectmode='day', font=('Arial', 10))
        self.holiday_calendar.pack(pady=5)

        # 添加日期按钮
        add_date_btn = ttk.Button(calendar_frame, text="添加到列表 →",
                                 command=lambda: self._add_selected_date())
        add_date_btn.pack(pady=5)

        # 右侧：已选日期列表
        selected_frame = ttk.Frame(date_frame)
        selected_frame.grid(row=0, column=1, sticky=tk.NSEW)

        ttk.Label(selected_frame, text="已选择的日期").pack(anchor=tk.W)

        # 日期列表显示
        list_frame = ttk.Frame(selected_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.selected_dates_list = tk.Listbox(list_frame, height=8,
                                            yscrollcommand=scrollbar.set,
                                            selectmode=tk.SINGLE)
        self.selected_dates_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.selected_dates_list.yview)

        # 移除日期按钮
        remove_btn = ttk.Button(selected_frame, text="移除选中日期",
                               command=lambda: self._remove_selected_date())
        remove_btn.pack(pady=5)

        # 手动输入日期选项
        manual_frame = ttk.LabelFrame(frm, text="手动输入日期", padding=5)
        manual_frame.grid(row=2, column=0, columnspan=3, sticky=tk.EW, pady=4)

        ttk.Label(manual_frame, text="日期(MM-DD)").grid(row=0, column=0, sticky=tk.W, pady=4)
        md_var = tk.StringVar()
        ttk.Entry(manual_frame, textvariable=md_var, width=12).grid(row=0, column=1, sticky=tk.W)
        ttk.Label(manual_frame, text="示例: 10-01").grid(row=0, column=2, sticky=tk.W)

        manual_add_btn = ttk.Button(manual_frame, text="手动添加",
                                   command=lambda: self._add_manual_date(md_var))
        manual_add_btn.grid(row=0, column=3, padx=5)

        ttk.Label(frm, text="节日名称").grid(row=3, column=0, sticky=tk.W, pady=4)
        name_var = tk.StringVar()
        ttk.Entry(frm, textvariable=name_var, width=25).grid(row=3, column=1, columnspan=2, sticky=tk.W)

        def _add_selected_date():
            """从日历添加选中的日期"""
            try:
                selected_date = self.holiday_calendar.get_date()
                # 检查返回的是字符串还是datetime对象
                if isinstance(selected_date, str):
                    # 如果是字符串格式，直接解析
                    if len(selected_date) >= 10:  # yyyy-MM-dd格式
                        month = selected_date[5:7]
                        day = selected_date[8:10]
                        md_str = f"{month}-{day}"
                    else:  # 其他格式，尝试分割
                        parts = selected_date.split('-')
                        if len(parts) >= 3:
                            md_str = f"{parts[1]}-{parts[2]}"
                        else:
                            raise ValueError("日期格式不正确")
                else:
                    # 如果是datetime对象，使用strftime
                    md_str = selected_date.strftime('%m-%d')

                # 检查是否已存在
                for i in range(self.selected_dates_list.size()):
                    if self.selected_dates_list.get(i) == md_str:
                        messagebox.showinfo("提示", f"日期 {md_str} 已在列表中")
                        return

                self.selected_dates_list.insert(tk.END, md_str)
            except Exception as e:
                messagebox.showwarning("警告", f"添加日期失败: {str(e)}")

        def _add_manual_date(md_var):
            """手动输入日期"""
            md = md_var.get().strip()
            if not md:
                return

            # 校验格式 MM-DD
            try:
                mm, dd = md.split('-')
                mm_i, dd_i = int(mm), int(dd)
                if mm_i < 1 or mm_i > 12 or dd_i < 1 or dd_i > 31:
                    raise ValueError()
            except Exception:
                messagebox.showwarning("警告", "日期格式应为 MM-DD，例如 10-01")
                return

            # 统一格式
            md_norm = f"{mm_i:02d}-{dd_i:02d}"

            # 检查是否已存在
            for i in range(self.selected_dates_list.size()):
                if self.selected_dates_list.get(i) == md_norm:
                    messagebox.showinfo("提示", f"日期 {md_norm} 已在列表中")
                    return

            self.selected_dates_list.insert(tk.END, md_norm)
            md_var.set("")  # 清空输入框

        def _remove_selected_date():
            """移除选中的日期"""
            selection = self.selected_dates_list.curselection()
            if selection:
                self.selected_dates_list.delete(selection[0])

        # 绑定方法到实例，供内部调用
        self._add_selected_date = _add_selected_date
        self._add_manual_date = _add_manual_date
        self._remove_selected_date = _remove_selected_date

        def on_save():
            y = year_var.get().strip()
            nm = name_var.get().strip()

            if not y or not nm:
                messagebox.showwarning("警告", "请完整填写年份与节日名称")
                return

            # 获取所有选择的日期
            selected_dates = []
            for i in range(self.selected_dates_list.size()):
                selected_dates.append(self.selected_dates_list.get(i))

            if not selected_dates:
                messagebox.showwarning("警告", "请至少选择一个日期")
                return

            # 添加所有选择的日期
            added_count = 0
            for md in selected_dates:
                # 日期格式已经在添加时验证过
                self.holidays.setdefault(y, {})[md] = nm
                added_count += 1

            self.save_data()
            self.update_holiday_tree()
            self.update_calendar()
            self.update_status(f"已添加节日: {nm} ({added_count}个日期)")
            top.destroy()

        action = ttk.Frame(top, padding=(0,10))
        action.grid(row=4, column=0, sticky=tk.E)
        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=6)
        ttk.Button(action, text="取消", command=top.destroy).pack(side=tk.RIGHT)

        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        frm.columnconfigure(2, weight=1)

    def edit_holiday_dialog(self, event=None):
        """编辑选中节日"""
        sel = self.holiday_tree.selection()
        if not sel:
            return

        item = sel[0]
        vals = self.holiday_tree.item(item).get("values", [])
        if len(vals) < 2:
            return

        old_md, old_name = vals[0], vals[1]
        year = self.holiday_year_var.get()

        top = tk.Toplevel(self.root)
        top.title("编辑节日")
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(frm, text="日期(MM-DD)").grid(row=0, column=0, sticky=tk.W, pady=4)
        md_var = tk.StringVar(value=old_md)
        ttk.Entry(frm, textvariable=md_var, width=12).grid(row=0, column=1, sticky=tk.W)

        ttk.Label(frm, text="节日名称").grid(row=1, column=0, sticky=tk.W, pady=4)
        name_var = tk.StringVar(value=old_name)
        ttk.Entry(frm, textvariable=name_var, width=18).grid(row=1, column=1, sticky=tk.W)

        def on_save():
            new_md = md_var.get().strip()
            new_name = name_var.get().strip()

            if not new_md or not new_name:
                messagebox.showwarning("警告", "请完整填写日期与名称")
                return

            # 检查日期格式
            try:
                mm, dd = new_md.split('-')
                mm_i, dd_i = int(mm), int(dd)
                if mm_i < 1 or mm_i > 12 or dd_i < 1 or dd_i > 31:
                    raise ValueError()
            except Exception:
                messagebox.showwarning("警告", "日期格式应为 MM-DD，例如 10-01")
                return

            # 如果日期有变更，先删除旧的
            if new_md != old_md:
                if self.holidays.get(year) and old_md in self.holidays[year]:
                    self.holidays[year].pop(old_md, None)

            # 添加或更新节日
            md_norm = f"{int(mm_i):02d}-{int(dd_i):02d}"
            self.holidays.setdefault(year, {})[md_norm] = new_name

            self.save_data()
            self.update_holiday_tree()
            self.update_calendar()
            self.update_status(f"已更新节日: {year}-{md_norm} {new_name}")
            top.destroy()

        action = ttk.Frame(top, padding=(0,10))
        action.grid(row=1, column=0, sticky=tk.E)
        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=6)
        ttk.Button(action, text="取消", command=top.destroy).pack(side=tk.RIGHT)

        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

    def delete_selected_holiday(self):
        """批量删除选中的多个节日记录，一次性确认后统一删除"""
        sel = self.holiday_tree.selection()
        if not sel:
            messagebox.showwarning("警告", "请先选择节日")
            return

        year = self.holiday_year_var.get()

        # 收集所有要删除的节日信息
        holidays_to_delete = []
        for item in sel:
            vals = self.holiday_tree.item(item).get("values", [])
            if len(vals) >= 1:
                md = vals[0]
                name = vals[1] if len(vals) > 1 else ""
                holidays_to_delete.append((md, name))

        if not holidays_to_delete:
            messagebox.showinfo("提示", "没有选中有效的节日")
            return

        # 一次性确认所有删除操作
        if len(holidays_to_delete) == 1:
            # 单条记录删除确认
            md, name = holidays_to_delete[0]
            holiday_info = f"{year}-{md} ({name})" if name else f"{year}-{md}"
            if not messagebox.askyesno("确认删除", f"确定删除节日：\n{holiday_info} 吗？"):
                return
        else:
            # 多条记录批量删除确认
            holiday_list = "\n".join([f"  • {md} ({name})" if name else f"  • {md}"
                                     for md, name in holidays_to_delete])
            confirm_msg = f"您选择了 {len(holidays_to_delete)} 个节日进行删除：\n\n{holiday_list}\n\n确定要删除这些节日吗？"

            if not messagebox.askyesno("批量删除确认", confirm_msg):
                return

        # 批量删除所有选中的节日
        deleted_count = 0
        failed_count = 0
        error_messages = []

        for md, name in holidays_to_delete:
            try:
                if self.holidays.get(year) and md in self.holidays[year]:
                    self.holidays[year].pop(md, None)
                    deleted_count += 1
                else:
                    failed_count += 1
                    error_messages.append(f"{md}: 节日不存在或已删除")
            except Exception as e:
                failed_count += 1
                error_messages.append(f"{md}: {str(e)}")

        # 保存数据并更新界面
        if deleted_count > 0:
            self.save_data()
            self.update_holiday_tree()
            self.update_calendar()

            # 显示删除结果
            if failed_count == 0:
                status_msg = f"成功删除 {deleted_count} 个节日"
                if deleted_count == 1:
                    messagebox.showinfo("删除成功", f"已成功删除选中的节日。")
                else:
                    messagebox.showinfo("批量删除成功", f"已成功删除 {deleted_count} 个节日。")
            else:
                status_msg = f"删除 {deleted_count} 个节日，失败 {failed_count} 个"
                error_detail = "\n".join(error_messages)
                messagebox.showwarning("删除完成", f"成功删除 {deleted_count} 个节日，\n失败 {failed_count} 个。\n\n失败详情：\n{error_detail}")

            self.update_status(status_msg)
        else:
            if failed_count > 0:
                messagebox.showerror("删除失败", "所有选中的节日删除失败。\n失败原因：\n" + "\n".join(error_messages))
            else:
                messagebox.showinfo("提示", "没有节日被删除")

    def copy_holiday_year(self):
        """复制当前选择年份的节假日到剪贴板。"""
        year = self.holiday_year_var.get()
        data = dict(self.holidays.get(year, {}))
        self._holidays_clipboard = {"year": year, "data": data}
        self.update_status(f"已复制 {year} 年节假日，共 {len(data)} 条")

    def paste_holiday_year_dialog(self):
        if not self._holidays_clipboard:
            messagebox.showwarning("警告", "剪贴板为空，请先复制年份")
            return
        top = tk.Toplevel(self.root)
        top.title("粘贴到年份")
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.grid(row=0, column=0, sticky=tk.NSEW)

        src_year = self._holidays_clipboard.get("year")
        ttk.Label(frm, text=f"来源年份: {src_year}").grid(row=0, column=0, columnspan=2, sticky=tk.W)

        ttk.Label(frm, text="目标年份").grid(row=1, column=0, sticky=tk.W, pady=6)
        target_var = tk.StringVar(value=self.holiday_year_var.get())
        target_entry = ttk.Entry(frm, textvariable=target_var, width=10)
        target_entry.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(frm, text="冲突策略").grid(row=2, column=0, sticky=tk.W, pady=6)
        mode_var = tk.StringVar(value="merge")
        ttk.Radiobutton(frm, text="合并(保留已存在)", variable=mode_var, value="merge").grid(row=2, column=1, sticky=tk.W)
        ttk.Radiobutton(frm, text="覆盖(替换已存在)", variable=mode_var, value="overwrite").grid(row=3, column=1, sticky=tk.W)

        def on_paste():
            ty = target_var.get().strip()
            if not ty.isdigit():
                messagebox.showwarning("警告", "目标年份需为数字")
                return
            copied = self._holidays_clipboard.get("data", {})
            if mode_var.get() == "overwrite":
                self.holidays[ty] = dict(copied)
            else:
                dest = self.holidays.setdefault(ty, {})
                for md, nm in copied.items():
                    if md not in dest:
                        dest[md] = nm
            self.save_data()
            # 若当前节假日页显示目标年，则刷新
            if hasattr(self, 'holiday_year_var') and self.holiday_year_var.get() == ty:
                self.update_holiday_tree()
            self.update_calendar()
            self.update_status(f"已粘贴到 {ty} 年")
            top.destroy()

        actions = ttk.Frame(top, padding=(0,10))
        actions.grid(row=1, column=0, sticky=tk.E)
        ttk.Button(actions, text="粘贴", command=on_paste).pack(side=tk.RIGHT, padx=6)
        ttk.Button(actions, text="取消", command=top.destroy).pack(side=tk.RIGHT)

        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

    def lunar_to_solar(self, lunar_year, lunar_month, lunar_day):
        """农历日期转公历"""
        lunar_date = Lunar(lunar_year, lunar_month, lunar_day)
        solar_date = Converter.Lunar2Solar(lunar_date)
        return solar_date

    def add_lunar_holiday(self, name, lunar_month, lunar_day):
        """添加农历节日"""
        year = self.holiday_year_var.get()
        try:
            year_num = int(year)
            solar_date = self.lunar_to_solar(year_num, lunar_month, lunar_day)
            md = f"{solar_date.month:02d}-{solar_date.day:02d}"
            self.holidays.setdefault(year, {})[md] = name
            self.save_data()
            self.update_holiday_tree()
            self.update_calendar()
            self.update_status(f"已添加农历节日: {year}年{lunar_month}月{lunar_day}日 ({md}) {name}")
        except ValueError:
            messagebox.showwarning("警告", "请输入有效的年份")

    def fill_national_holidays(self):
        """一键填充国家法定节假日（包含当前年份）。"""
        # 从当前年份向前后各扩展2年
        current_year = int(self.holiday_year_var.get())
        start_year, end_year = current_year - 2, current_year + 2
        for y in range(start_year, end_year + 1):
            ys = str(y)
            # 清空当前年份的节假日
            self.holidays[ys] = {}

            # 公历节日
            self.holidays[ys].setdefault("01-01", "元旦")
            self.holidays[ys].setdefault("05-01", "劳动节")
            for d in range(1, 4):  # 国庆节前3天
                self.holidays[ys].setdefault(f"10-{d:02d}", "国庆节")
            # 添加除夕（春节的前一天）
            try:
                # 首先计算春节（农历正月初一）
                spring_festival = self.lunar_to_solar(y, 1, 1)
                festival_date = datetime.date(spring_festival.year,
                                             spring_festival.month,
                                             spring_festival.day)
                # 除夕是春节的前一天
                eve_date = festival_date - datetime.timedelta(days=1)
                md_eve = f"{eve_date.month:02d}-{eve_date.day:02d}"
                self.holidays[ys][md_eve] = "除夕"
            except Exception:
                # 备用方法：尝试计算农历腊月三十或二十九
                try:
                    # 尝试农历腊月三十
                    lunar_new_year_eve = self.lunar_to_solar(y-1, 12, 30)
                    md_eve = f"{lunar_new_year_eve.month:02d}-{lunar_new_year_eve.day:02d}"
                    self.holidays[ys][md_eve] = "除夕"
                except Exception:
                    # 如果没有腊月三十，尝试腊月二十九
                    try:
                        lunar_new_year_eve = self.lunar_to_solar(y-1, 12, 29)
                        md_eve = f"{lunar_new_year_eve.month:02d}-{lunar_new_year_eve.day:02d}"
                        self.holidays[ys][md_eve] = "除夕"
                    except Exception:
                        pass

            # 农历节日（精确计算）
            # 春节（农历正月初一）
            try:
                solar_date = self.lunar_to_solar(y, 1, 1)
                spring_md = f"{solar_date.month:02d}-{solar_date.day:02d}"
                self.holidays[ys].setdefault(spring_md, "春节")
            except Exception:
                pass

            # 元宵节（农历正月十五）
            try:
                solar_date = self.lunar_to_solar(y, 1, 15)
                self.holidays[ys].setdefault(f"{solar_date.month:02d}-{solar_date.day:02d}", "元宵节")
            except Exception:
                pass

            # 清明节（公历4月4日或5日）
            self.holidays[ys].setdefault("04-04", "清明节")

            # 端午节（农历五月初五）
            try:
                solar_date = self.lunar_to_solar(y, 5, 5)
                self.holidays[ys].setdefault(f"{solar_date.month:02d}-{solar_date.day:02d}", "端午节")
            except Exception:
                pass

            # 中秋节（农历八月十五）
            try:
                solar_date = self.lunar_to_solar(y, 8, 15)
                self.holidays[ys].setdefault(f"{solar_date.month:02d}-{solar_date.day:02d}", "中秋节")
            except Exception:
                pass

            # 重阳节（农历九月初九）
            try:
                solar_date = self.lunar_to_solar(y, 9, 9)
                self.holidays[ys].setdefault(f"{solar_date.month:02d}-{solar_date.day:02d}", "重阳节")
            except Exception:
                pass

            # 冬至（公历12月21日或22日）
            self.holidays[ys].setdefault("12-21", "冬至")

        self.save_data()
        self.update_holiday_tree()
        self.update_calendar()
        self.update_status(f"已填充法定节假日（{start_year}-{end_year}），包含农历节日")



    def open_official_fetch_dialog(self):
        """根据权威数据源抓取节假日（需要网络）。"""
        top = tk.Toplevel(self.root)
        top.title("权威源更新节假日")
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(frm, text="起始年").grid(row=0, column=0, sticky=tk.W, pady=4)
        start_var = tk.StringVar(value=str(max(2025, self.current_date.year)))
        ttk.Entry(frm, textvariable=start_var, width=8).grid(row=0, column=1, sticky=tk.W)

        ttk.Label(frm, text="结束年").grid(row=1, column=0, sticky=tk.W, pady=4)
        end_var = tk.StringVar(value=str(max(2025, self.current_date.year)))
        ttk.Entry(frm, textvariable=end_var, width=8).grid(row=1, column=1, sticky=tk.W)

        ttk.Label(frm, text="API（如 Nager.Date 或工作日历API）").grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(8,4))
        api_var = tk.StringVar(value="https://date.nager.at/api/v3/PublicHolidays/{year}/CN")
        ttk.Entry(frm, textvariable=api_var, width=48).grid(row=3, column=0, columnspan=2, sticky=tk.W)

        def on_fetch():
            if requests is None:
                messagebox.showwarning("警告", "未安装requests库，无法联网获取。可先使用CSV导入或安装requests。")
                return
            try:
                s = int(start_var.get()); e = int(end_var.get())
            except Exception:
                messagebox.showwarning("警告", "年份需为整数")
                return
            if e < s:
                s, e = e, s
            total_added = 0
            for y in range(s, e+1):
                url = api_var.get().replace("{year}", str(y))
                try:
                    resp = requests.get(url, timeout=10)
                    resp.raise_for_status()
                    data = resp.json()
                except Exception as ex:
                    messagebox.showwarning("提示", f"获取 {y} 年失败: {ex}")
                    continue
                # 解析Nager.Date结构：[{date:"2025-01-01", localName:"元旦", ...}]
                for item in data:
                    date_iso = item.get("date", "")
                    name = item.get("localName") or item.get("name") or "节假日"
                    if len(date_iso) >= 10:
                        yyyy, mm, dd = date_iso[0:4], date_iso[5:7], date_iso[8:10]
                        if yyyy.isdigit():
                            md = f"{mm}-{dd}"
                            self.holidays.setdefault(yyyy, {})[md] = name
                            total_added += 1
            self.save_data()
            self.update_holiday_tree()
            self.update_calendar()
            self.update_status(f"已从权威源获取完成，新增/更新 {total_added} 条")
            top.destroy()

        actions = ttk.Frame(top, padding=(0,10))
        actions.grid(row=1, column=0, sticky=tk.E)
        ttk.Button(actions, text="获取", command=on_fetch).pack(side=tk.RIGHT, padx=6)
        ttk.Button(actions, text="取消", command=top.destroy).pack(side=tk.RIGHT)

        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

    def update_leave_type_list(self):
        self.leave_type_list.delete(0, tk.END)
        # 显示时按名称排序
        for t in sorted(self.leave_types):
            self.leave_type_list.insert(tk.END, t)

    def add_leave_type(self):
        def do_add():
            val = entry_var.get().strip()
            if not val:
                return
            if val in self.leave_types:
                messagebox.showwarning("警告", "类型已存在")
                return
            self.leave_types.append(val)
            self.update_leave_type_list()
            # 更新类型下拉
            self.leave_type_combo["values"] = self.leave_types
            if len(self.leave_types) == 1:
                self.leave_type_combo.current(0)
            # 若当前选中类型不在新列表中，回退到刚添加的类型
            if self.leave_type_var.get() not in self.leave_types:
                self.leave_type_var.set(val)
                try:
                    idx = self.leave_types.index(val)
                    self.leave_type_combo.current(idx)
                except Exception:
                    pass
            self.save_data()
            top.destroy()

        top = tk.Toplevel(self.root)
        top.title("添加请假类型")
        ttk.Label(top, text="类型名称").grid(row=0, column=0, padx=10, pady=10)
        entry_var = tk.StringVar()
        ttk.Entry(top, textvariable=entry_var, width=24).grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(top, text="确定", command=do_add).grid(row=1, column=0, columnspan=2, pady=10)
        top.transient(self.root)
        top.grab_set()

    def delete_leave_type(self):
        sel = self.leave_type_list.curselection()
        if not sel:
            messagebox.showwarning("警告", "请先选择类型")
            return
        val = self.leave_type_list.get(sel[0])
        if messagebox.askyesno("确认", f"确定删除类型 '{val}' 吗？"):
            try:
                self.leave_types.remove(val)
            except ValueError:
                pass
            self.update_leave_type_list()
            # 同步下拉
            self.leave_type_combo["values"] = self.leave_types
            # 若删除的是当前选项或当前不合法，则回退到第一个或空
            if self.leave_type_var.get() not in self.leave_types:
                if self.leave_types:
                    self.leave_type_var.set(self.leave_types[0])
                    try:
                        self.leave_type_combo.current(0)
                    except Exception:
                        pass
                else:
                    self.leave_type_var.set("")
            self.save_data()

    def update_leave_tree(self):
        if not hasattr(self, 'leave_tree'):
            return
        self.leave_tree.delete(*self.leave_tree.get_children())
        # 排序：人员名称 -> 日期 -> 类型
        def sort_key(r):
            return (
                r.get("plan_name", ""),
                r.get("date", ""),
                r.get("type", "")
            )
        for rec in sorted(self.leave_records, key=sort_key):
            self.leave_tree.insert("", tk.END, values=(rec.get("plan_name", ""), rec.get("date", ""), rec.get("type", ""), rec.get("note", "")))
        # 同步统计
        self.update_leave_stats()
        # 更新年份选项
        self.update_leave_stats_year_options()

    def update_leave_stats(self):
        """统计每个计划的总请假天数与类型分布。
        规则：同一计划、同一天、同一类型 计 1 天；总天数按去重的日期集合计算。
        现在按年休假年度区分统计。
        支持按年份过滤显示。
        """
        if not hasattr(self, 'leave_stats_tree'):
            return
        self.leave_stats_tree.delete(*self.leave_stats_tree.get_children())

        # 获取选择的年份
        selected_year_str = self.leave_stats_year_var.get() if hasattr(self, 'leave_stats_year_var') else "全部"
        selected_year = None
        if selected_year_str != "全部":
            try:
                selected_year = int(selected_year_str)
            except ValueError:
                selected_year = None

        # 聚合：plan -> leave_year -> type -> set(dates)
        plan_year_type_dates = {}
        current_date = datetime.date.today()

        # 确定当前年休假年度
        if current_date.month >= 4:
            current_leave_year = current_date.year
        else:
            current_leave_year = current_date.year - 1

        for rec in self.leave_records:
            plan = rec.get("plan_name", "")
            date_str = rec.get("date", "")
            ltype = rec.get("type", "")
            if not plan or not date_str or not ltype:
                continue

            try:
                parts = date_str.split('-')
                record_year = int(parts[0])
                record_month = int(parts[1])

                # 根据年休假规则确定配额年份
                if self._is_annual_leave(ltype):
                    # 年休假：4-12月属于当年配额，1-3月属于上年配额
                    if record_month >= 4:
                        leave_year = record_year
                    else:
                        leave_year = record_year - 1
                else:
                    # 其他假期类型使用自然年
                    leave_year = record_year

                # 如果选择了特定年份，只统计该年份的数据
                if selected_year is not None and leave_year != selected_year:
                    continue

                # 添加到对应的年份分组
                plan_year_type_dates.setdefault(plan, {}).setdefault(leave_year, {}).setdefault(ltype, set()).add(date_str)

            except Exception:
                continue

        # 按人员显示统计，区分不同年份
        for plan in sorted(plan_year_type_dates.keys()):
            for leave_year in sorted(plan_year_type_dates[plan].keys()):
                # 计算该年份的总天数（按日期去重，不区分类型）
                all_dates = set()
                type_parts = []
                for ltype in sorted(plan_year_type_dates[plan][leave_year].keys()):
                    dates = plan_year_type_dates[plan][leave_year][ltype]
                    all_dates |= dates
                    type_parts.append(f"{ltype}:{len(dates)}")
                days = len(all_dates)
                types_str = "，".join(type_parts)

                # 根据是否为当前年份显示不同标识
                if leave_year == current_leave_year:
                    year_display = f"{leave_year}年 (当前)"
                else:
                    year_display = f"{leave_year}年"

                # 如果选择了"全部"，显示年份信息；否则不显示年份（因为都是同一年）
                if selected_year_str == "全部":
                    plan_display = f"{plan} - {year_display}"
                else:
                    plan_display = plan

                self.leave_stats_tree.insert("", tk.END, values=(plan_display, days, types_str))

    def on_leave_stats_year_selected(self, event=None):
        """当请假统计年份选择变更时更新统计数据"""
        self.update_leave_stats()

    def update_leave_stats_year_options(self):
        """更新请假统计年份选择器的选项"""
        if not hasattr(self, 'leave_stats_year_combo'):
            return

        # 收集所有请假记录中的年份
        years = set()

        # 添加当前年份作为默认选项
        current_date = datetime.date.today()
        if current_date.month >= 4:
            current_leave_year = current_date.year
        else:
            current_leave_year = current_date.year - 1
        years.add(current_leave_year)

        for rec in self.leave_records:
            date_str = rec.get("date", "")
            if date_str:
                try:
                    parts = date_str.split('-')
                    record_year = int(parts[0])
                    record_month = int(parts[1])
                    ltype = rec.get("type", "")

                    # 根据年休假规则确定配额年份
                    if self._is_annual_leave(ltype):
                        # 年休假：4-12月属于当年配额，1-3月属于上年配额
                        if record_month >= 4:
                            leave_year = record_year
                        else:
                            leave_year = record_year - 1
                    else:
                        # 其他假期类型使用自然年
                        leave_year = record_year

                    years.add(leave_year)
                except Exception:
                    pass

        # 排序并添加"全部"选项
        year_options = ["全部"] + [str(y) for y in sorted(list(years), reverse=True)]
        self.leave_stats_year_combo['values'] = year_options

        # 如果当前选择的年份不在选项中，重置为"全部"
        current_year = self.leave_stats_year_var.get()
        if current_year not in year_options:
            self.leave_stats_year_var.set("全部")

    def _is_annual_leave(self, leave_type):
        """判断是否为年休假类型"""
        return leave_type in ["年休假", "年假"]

    def _calculate_annual_leave_usage(self, plan, year):
        """计算指定人员指定年份的年休假使用情况

        年休假特殊规则：
        - 当年年休假可延续到次年3月底使用
        - 1-3月请的年休假优先扣减上一年剩余配额
        - 上一年配额用完后才扣减当年配额

        Args:
            plan: 人员名称
            year: 配额年份

        Returns:
            int: 该年份配额被使用的天数
        """
        # 获取该年份的年休假配额
        quota = self.leave_quotas.get(plan, {}).get(str(year), {}).get("年休假", 0)
        if quota == 0:
            # 也检查"年假"这个名称
            quota = self.leave_quotas.get(plan, {}).get(str(year), {}).get("年假", 0)

        # 收集所有该人员的年休假记录
        annual_leave_records = []
        for rec in self.leave_records:
            if rec.get("plan_name") != plan:
                continue
            if not self._is_annual_leave(rec.get("type", "")):
                continue
            date_str = rec.get("date", "")
            if not date_str:
                continue
            annual_leave_records.append(date_str)

        # 按日期排序（去重）
        annual_leave_dates = sorted(set(annual_leave_records))

        # 分类：当年4-12月 和 次年1-3月
        current_year_dates = []  # 当年4-12月
        next_year_q1_dates = []  # 次年1-3月

        for date_str in annual_leave_dates:
            try:
                parts = date_str.split('-')
                d_year = int(parts[0])
                d_month = int(parts[1])

                if d_year == year and d_month >= 4:
                    # 当年4-12月
                    current_year_dates.append(date_str)
                elif d_year == year + 1 and d_month <= 3:
                    # 次年1-3月
                    next_year_q1_dates.append(date_str)
            except Exception:
                continue

        # 计算使用量
        # 1. 当年4-12月的全部计入当年配额
        used_from_current_year = len(current_year_dates)

        # 2. 次年1-3月的，先用当年剩余配额，用完后算下一年的
        remaining_quota = max(0, quota - used_from_current_year)
        used_from_next_year_q1 = min(len(next_year_q1_dates), remaining_quota)

        # 总使用量
        total_used = used_from_current_year + used_from_next_year_q1

        return total_used

    def _calculate_current_year_annual_leave_usage(self, plan, year):
        """计算当年1-3月从当年配额中扣除的年休假天数

        这个函数专门用来计算：当查看当年配额时，有多少天是从当年1-3月使用的
        （即上一年配额用完后，从当年配额扣除的部分）

        Args:
            plan: 人员名称
            year: 当年年份

        Returns:
            int: 从当年配额扣除的天数
        """
        # 获取上一年的配额
        last_year = year - 1
        last_year_quota = self.leave_quotas.get(plan, {}).get(str(last_year), {}).get("年休假", 0)
        if last_year_quota == 0:
            last_year_quota = self.leave_quotas.get(plan, {}).get(str(last_year), {}).get("年假", 0)

        # 计算上一年的使用情况（不包括当年1-3月）
        last_year_used = 0
        for rec in self.leave_records:
            if rec.get("plan_name") != plan:
                continue
            if not self._is_annual_leave(rec.get("type", "")):
                continue
            date_str = rec.get("date", "")
            try:
                parts = date_str.split('-')
                d_year = int(parts[0])
                d_month = int(parts[1])
                # 只统计上一年4-12月的使用
                if d_year == last_year and d_month >= 4:
                    last_year_used += 1
            except Exception:
                continue

        # 上一年的剩余配额
        last_year_remaining = max(0, last_year_quota - last_year_used)

        # 统计当年1-3月的年休假记录数
        current_year_q1_count = 0
        for rec in self.leave_records:
            if rec.get("plan_name") != plan:
                continue
            if not self._is_annual_leave(rec.get("type", "")):
                continue
            date_str = rec.get("date", "")
            try:
                parts = date_str.split('-')
                d_year = int(parts[0])
                d_month = int(parts[1])
                if d_year == year and d_month >= 1 and d_month <= 3:
                    current_year_q1_count += 1
            except Exception:
                continue

        # 当年1-3月从当年配额扣除的天数 = 总数 - 从上一年扣除的天数
        used_from_current_year = max(0, current_year_q1_count - last_year_remaining)

        return used_from_current_year

    def _get_annual_leave_date_range(self, year):
        """获取指定年份年休假的有效日期范围

        Args:
            year: 年休假归属年份

        Returns:
            tuple: (开始日期字符串, 结束日期字符串)
        """
        # 年休假周期：当年4月1日 至 次年3月31日
        start_date = f"{year}-04-01"
        end_date = f"{year + 1}-03-31"
        return start_date, end_date

    def _get_remaining_quota(self, plan, date_str, leave_type, exclude_record=None):
        """计算指定人员、日期和请假类型的剩余配额

        Args:
            plan: 人员名称
            date_str: 请假日期 (YYYY-MM-DD格式)
            leave_type: 请假类型
            exclude_record: 需要排除的记录（用于编辑时）

        Returns:
            int: 剩余配额天数
        """
        try:
            date_parts = date_str.split('-')
            year = int(date_parts[0])
            month = int(date_parts[1])
        except Exception:
            return 0

        # 根据请假类型确定配额年份
        if self._is_annual_leave(leave_type):
            # 年休假：4-12月用当年配额，1-3月优先用上一年配额
            if month >= 4:
                quota_year = year
            else:
                quota_year = year - 1
        else:
            # 其他类型：使用自然年
            quota_year = year

        # 跨年失效检查：非年休假在跨年后失效
        if not self._is_annual_leave(leave_type):
            current_year = self.current_date.year
            current_month = self.current_date.month
            # 当前时间是第二年1-3月，且请假日期的配额年份是上一年，则配额已失效
            if current_month >= 1 and current_month <= 3 and quota_year < current_year:
                return 0

        # 获取配额
        quota = self.leave_quotas.get(plan, {}).get(str(quota_year), {}).get(leave_type, 0)

        # 计算已用天数
        if self._is_annual_leave(leave_type):
            # 年休假使用特殊计算方法
            used_days = self._calculate_annual_leave_usage(plan, quota_year)

            # 年休假特殊处理：1-3月时，如果上一年配额不足，检查当年配额
            if month >= 1 and month <= 3:
                remaining_last_year = max(0, quota - used_days)
                if remaining_last_year == 0:
                    # 上一年配额已用完，检查当年配额
                    current_year_quota = self.leave_quotas.get(plan, {}).get(str(year), {}).get(leave_type, 0)
                    if current_year_quota == 0:
                        current_year_quota = self.leave_quotas.get(plan, {}).get(str(year), {}).get("年假", 0)

                    # 当年配额的已用天数（只计算当年1-3月的使用情况）
                    current_year_used = 0
                    for rec in self.leave_records:
                        if exclude_record and rec == exclude_record:
                            continue
                        if rec.get("plan_name") == plan and self._is_annual_leave(rec.get("type", "")):
                            rec_date_str = rec.get("date", "")
                            try:
                                rec_parts = rec_date_str.split('-')
                                rec_year = int(rec_parts[0])
                                rec_month = int(rec_parts[1])
                                # 只统计当年1-3月使用上一年配额后，继续使用当年配额的部分
                                if rec_year == year and rec_month >= 1 and rec_month <= 3:
                                    # 这部分需要减去上一年剩余配额后才是使用当年配额的
                                    pass  # 先不计算，使用简化逻辑
                            except Exception:
                                continue

                    # 返回当年配额的剩余（这里简化处理，假设当年1-3月没有用过当年配额）
                    return current_year_quota
                else:
                    return remaining_last_year
        else:
            # 其他类型统计自然年内的使用情况
            used_days = 0
            for rec in self.leave_records:
                # 如果需要排除某条记录（编辑时）
                if exclude_record and rec == exclude_record:
                    continue

                if rec.get("plan_name") == plan and rec.get("type") == leave_type:
                    rec_date_str = rec.get("date", "")
                    try:
                        rec_year = int(rec_date_str.split('-')[0])
                        if rec_year == quota_year:
                            used_days += 1
                    except Exception:
                        continue

        # 返回剩余配额（不允许负数）
        return max(0, quota - used_days)

    def _check_and_allocate_quota(self, plan, date_str, requested_type, exclude_record=None):
        """检查配额并自动顺延到可用的请假类型

        按照优先级顺序：婚假→育儿假→年休假→带薪病事假

        Args:
            plan: 人员名称
            date_str: 请假日期
            requested_type: 用户请求的请假类型
            exclude_record: 需要排除的记录（用于编辑时）

        Returns:
            dict: {
                'success': bool,  # 是否找到可用配额
                'allocated_type': str,  # 实际分配的请假类型
                'message': str,  # 提示信息
                'cascaded': bool  # 是否发生了自动顺延
            }
        """
        # 定义配额使用优先级顺序
        quota_priority = ["婚假", "育儿假", "年休假", "带薪病事假"]

        # 检查请求的类型是否有剩余配额
        remaining = self._get_remaining_quota(plan, date_str, requested_type, exclude_record)

        if remaining > 0:
            # 请求的类型有配额，直接使用
            return {
                'success': True,
                'allocated_type': requested_type,
                'message': f"使用 {requested_type} 配额，剩余 {remaining - 1} 天",
                'cascaded': False
            }

        # 请求的类型配额不足，尝试自动顺延
        # 从优先级列表中查找可用配额
        try:
            requested_index = quota_priority.index(requested_type)
        except ValueError:
            # 如果请求的类型不在优先级列表中，说明不支持自动顺延
            return {
                'success': False,
                'allocated_type': None,
                'message': f"{requested_type} 配额不足（剩余0天），且该类型不支持自动顺延",
                'cascaded': False
            }

        # 从下一个优先级开始查找
        cascade_message_parts = [f"{requested_type} 配额不足"]

        for fallback_type in quota_priority[requested_index + 1:]:
            fallback_remaining = self._get_remaining_quota(plan, date_str, fallback_type, exclude_record)

            if fallback_remaining > 0:
                # 找到可用的配额类型
                message = f"{requested_type} 配额不足，已自动使用 {fallback_type} 配额（剩余 {fallback_remaining - 1} 天）"
                return {
                    'success': True,
                    'allocated_type': fallback_type,
                    'message': message,
                    'cascaded': True
                }

        # 所有配额类型都不足
        return {
            'success': False,
            'allocated_type': None,
            'message': f"{requested_type} 及后续可用配额类型均不足，无法添加请假记录",
            'cascaded': False
        }

    def update_quota_summary(self, year=None):
        """汇总右下角年度配额与使用。支持指定年份。

        年休假特殊规则：
        - 年休假周期为4月1日至次年3月31日
        - 1-3月请的年休假优先扣减上一年剩余配额
        - 上一年配额用完后才扣减当年配额

        其他假期类型使用自然年（1月-12月）

        跨年规则：
        - 当前时间在第二年1-3月，查看上一年配额时：
          - 年休假：正常显示配额和使用情况
          - 其他假期：配额和已使用都清零（因为已过期）
        """
        if not hasattr(self, 'leave_quota_tree'):
            return
        self.leave_quota_tree.delete(*self.leave_quota_tree.get_children())

        # 如果没有指定年份，使用当前选择的年份
        if year is None:
            try:
                year = int(self.quota_year_var.get())
            except Exception:
                year = self.current_date.year

        # 获取当前日期的年份和月份
        current_year = self.current_date.year
        current_month = self.current_date.month

        # 判断是否需要清零非年休假配额
        # 条件：当前时间是第二年的1-3月，且查看的是上一年的配额
        should_clear_non_annual = (
            current_month >= 1 and current_month <= 3 and
            year < current_year
        )

        # 计算非年休假的已用天数（使用自然年）
        used_map_normal = {}
        for rec in self.leave_records:
            plan = rec.get("plan_name", "")
            date_str = rec.get("date", "")
            ltype = rec.get("type", "")
            if not plan or not date_str or not ltype:
                continue

            # 跳过年休假，年休假单独处理
            if self._is_annual_leave(ltype):
                continue

            # 其他假期类型使用自然年
            try:
                d_year = int(date_str.split('-')[0])
            except Exception:
                continue

            if d_year != year:
                continue
            used_map_normal.setdefault((plan, ltype), set()).add(date_str)

        # 计算每个人员的剩余总数
        plans = sorted(self.shift_schedules.keys())
        types = sorted(set(self.leave_types))

        # 请假类型简化映射
        type_short_names = {
            "带薪病事假": "带薪",
            "年休假": "年",
            "年假": "年",
            "育儿假": "育儿",
            "病假": "病",
            "事假": "事",
            "婚假": "婚",
            "产假": "产",
            "陪产假": "陪产",
            "丧假": "丧"
        }

        for plan in plans:
            # 计算该人员所有请假类型的剩余天数总和
            total_remain = 0
            quota_summary = []  # 用于显示配额汇总信息
            used_summary = []   # 用于显示已用汇总信息
            remain_summary = [] # 用于显示剩余汇总信息

            for ltype in types:
                is_annual = self._is_annual_leave(ltype)

                # 获取原始配额（总是显示实际设置的配额）
                quota = self.leave_quotas.get(plan, {}).get(str(year), {}).get(ltype, 0)

                # 计算已使用天数（总是显示实际使用情况）
                if is_annual:
                    # 年休假使用特殊计算方法
                    # 查看当年时：需要包含当年1-3月从当年配额扣除的部分
                    # 查看历史年时：使用原有逻辑
                    if year == current_year:
                        # 当前年份：包含1-3月从当年配额扣除的部分
                        used_days_from_q1 = self._calculate_current_year_annual_leave_usage(plan, year)
                        # 计算4-12月的使用（如果有的话）
                        used_days_from_rest = 0
                        for rec in self.leave_records:
                            if rec.get("plan_name") == plan and self._is_annual_leave(rec.get("type", "")):
                                rec_date_str = rec.get("date", "")
                                try:
                                    parts = rec_date_str.split('-')
                                    d_year = int(parts[0])
                                    d_month = int(parts[1])
                                    if d_year == year and d_month >= 4:
                                        used_days_from_rest += 1
                                except Exception:
                                    continue
                        used_days = used_days_from_q1 + used_days_from_rest
                    else:
                        # 历史年份：使用原有逻辑
                        used_days = self._calculate_annual_leave_usage(plan, year)
                else:
                    # 其他假期类型使用自然年统计
                    used_days = len(used_map_normal.get((plan, ltype), set()))

                # 计算剩余天数
                if should_clear_non_annual and not is_annual:
                    # 非年休假在跨年后：剩余清零（已失效），但已用仍显示实际值
                    remain = 0
                else:
                    # 年休假或非跨年情况：正常计算剩余
                    remain = max(quota - used_days, 0)

                # 剩余总数：只累加有效的剩余天数（跨年失效的不计入）
                if not (should_clear_non_annual and not is_annual):
                    total_remain += remain

                # 构建汇总信息（使用简化名称）
                # 只要配额大于0（即设置过配额），就显示该类型
                if quota > 0:
                    short_type = type_short_names.get(ltype, ltype)
                    quota_summary.append(f"{short_type}:{quota}")
                    used_summary.append(f"{short_type}:{used_days}")
                    remain_summary.append(f"{short_type}:{remain}")

            # 每个人员只显示一行，包含剩余总数
            quota_str = "，".join(quota_summary) if quota_summary else "无配额"
            used_str = "，".join(used_summary) if used_summary else "0"
            remain_str = "，".join(remain_summary) if remain_summary else "0"

            self.leave_quota_tree.insert("", tk.END, values=(
                plan,  # 人员名称
                quota_str,  # 配额汇总
                year,  # 年份
                used_str,  # 已用汇总
                remain_str,  # 剩余汇总
                total_remain  # 剩余总数
            ))

    def open_quota_setting(self):
        """设置年度配额的对话框（计划/类型/年份/配额）。"""
        # 获取当前选择的年份
        selected_year = self.quota_year_var.get() if hasattr(self, 'quota_year_var') else None
        # 调用edit_quota_dialog，预填当前选择的年份
        self.edit_quota_dialog(year=selected_year)

    def add_leave_record(self):
        plan = self.leave_plan_var.get().strip()
        if not plan:
            messagebox.showwarning("警告", "请先选择人员名称")
            return
        if plan not in self.shift_schedules:
            messagebox.showwarning("警告", "该计划不存在")
            return
        ltype = self.leave_type_var.get().strip()
        if not ltype:
            messagebox.showwarning("警告", "请选择请假类型")
            return
        date_str = self.leave_date.get_date().strftime('%Y-%m-%d')
        note = self.leave_note_var.get().strip()

        # 检查配额并实现自动顺延
        quota_result = self._check_and_allocate_quota(plan, date_str, ltype)

        if not quota_result['success']:
            # 配额不足，无法添加
            messagebox.showerror("配额不足", quota_result['message'])
            return

        # 获取实际分配的请假类型
        allocated_type = quota_result['allocated_type']

        # 如果发生了自动顺延，需要用户确认
        if quota_result['cascaded']:
            confirm = messagebox.askyesno(
                "配额自动顺延",
                f"{quota_result['message']}\n\n是否确认添加此请假记录？"
            )
            if not confirm:
                return

        # 使用实际分配的请假类型添加记录
        self.leave_records.append({"plan_name": plan, "date": date_str, "type": allocated_type, "note": note})
        self.update_leave_tree()
        self.update_leave_stats()
        self.update_quota_summary()
        self.update_quota_year_options()  # 更新年份选项
        self.update_current_leave_year_display()  # 更新年份显示
        self.save_data()
        self.update_calendar()
        self.update_status(f"已添加请假记录: {plan} {date_str} {allocated_type}")

        # 构建成功消息
        success_message = f"已成功添加请假记录:\n人员: {plan}\n日期: {date_str}\n类型: {allocated_type}\n备注: {note}"
        if quota_result['cascaded']:
            success_message += f"\n\n{quota_result['message']}"
        else:
            success_message += f"\n\n{quota_result['message']}"

        messagebox.showinfo("成功", success_message)

    def delete_leave_record(self):
        sel = self.leave_tree.selection()
        if not sel:
            messagebox.showwarning("警告", "请先选择记录")
            return

        # 收集所有要删除的记录信息
        records_to_delete = []
        for item_id in sel:
            item = self.leave_tree.item(item_id)
            vals = item.get("values", [])
            if len(vals) >= 3:
                plan, date_str, ltype = vals[0], vals[1], vals[2]
                records_to_delete.append({
                    "plan_name": plan,
                    "date": date_str,
                    "type": ltype
                })

        if not records_to_delete:
            messagebox.showwarning("警告", "未找到有效的记录")
            return

        # 确认删除
        count = len(records_to_delete)
        if count == 1:
            rec = records_to_delete[0]
            confirm_msg = f"确定删除记录: {rec['plan_name']} {rec['date']} {rec['type']} 吗？"
        else:
            confirm_msg = f"确定删除选中的 {count} 条记录吗？"

        if messagebox.askyesno("确认", confirm_msg):
            # 删除所有选中的记录
            for rec in records_to_delete:
                self.leave_records = [r for r in self.leave_records if not (
                    r.get("plan_name") == rec["plan_name"] and
                    r.get("date") == rec["date"] and
                    r.get("type") == rec["type"]
                )]

            # 更新界面
            self.update_leave_tree()
            self.update_leave_stats()
            self.update_quota_summary()
            self.update_quota_year_options()  # 更新年份选项
            self.update_current_leave_year_display()  # 更新年份显示
            self.save_data()
            self.update_calendar()

            if count == 1:
                self.update_status("已删除请假记录")
            else:
                self.update_status(f"已删除 {count} 条请假记录")

    def edit_leave_record(self, selected_item=None):
        """编辑请假记录
        Args:
            selected_item: 可选，如果传入则直接使用，否则从selection获取
        """
        if selected_item is None:
            sel = self.leave_tree.selection()
            if not sel:
                messagebox.showwarning("警告", "请先选择要编辑的记录")
                return
            selected_item = sel[0]

        item = self.leave_tree.item(selected_item)
        vals = item.get("values", [])
        if len(vals) < 3:
            messagebox.showwarning("警告", "记录数据不完整")
            return
        plan, date_str, ltype = vals[0], vals[1], vals[2]

        # 查找要编辑的记录
        target_record = None
        for record in self.leave_records:
            if (record.get("plan_name") == plan and
                record.get("date") == date_str and
                record.get("type") == ltype):
                target_record = record
                break

        if not target_record:
            messagebox.showerror("错误", "找不到要编辑的记录")
            return

        # 创建编辑对话框
        edit_window = tk.Toplevel(self.root)
        edit_window.title("编辑请假记录")
        edit_window.geometry("400x250")
        edit_window.resizable(False, False)
        edit_window.transient(self.root)
        edit_window.grab_set()

        # 计算对话框位置（居中）
        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (edit_window.winfo_width() // 2)
        y = (edit_window.winfo_screenheight() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")

        # 创建表单
        main_frame = ttk.Frame(edit_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 人员名称
        ttk.Label(main_frame, text="人员名称:").grid(row=0, column=0, sticky=tk.W, pady=5)
        plan_var = tk.StringVar(value=target_record.get("plan_name", ""))
        plan_combo = ttk.Combobox(main_frame, textvariable=plan_var, width=25, state="readonly")
        plan_combo['values'] = list(self.shift_schedules.keys())
        plan_combo.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 请假日期
        ttk.Label(main_frame, text="请假日期:").grid(row=1, column=0, sticky=tk.W, pady=5)
        date_var = tk.StringVar(value=target_record.get("date", ""))
        date_entry = ttk.Entry(main_frame, textvariable=date_var, width=27)
        date_entry.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 请假类型
        ttk.Label(main_frame, text="请假类型:").grid(row=2, column=0, sticky=tk.W, pady=5)
        type_var = tk.StringVar(value=target_record.get("type", ""))
        type_combo = ttk.Combobox(main_frame, textvariable=type_var, width=25, state="readonly")
        type_combo['values'] = ["带薪病事假", "年休假", "育儿假", "婚假", "丧假"]
        type_combo.grid(row=2, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 备注
        ttk.Label(main_frame, text="备注:").grid(row=3, column=0, sticky=tk.W, pady=5)
        note_var = tk.StringVar(value=target_record.get("note", ""))
        note_entry = ttk.Entry(main_frame, textvariable=note_var, width=28)
        note_entry.grid(row=3, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0))

        def save_changes():
            """保存修改"""
            new_plan = plan_var.get().strip()
            new_date = date_var.get().strip()
            new_type = type_var.get().strip()
            new_note = note_var.get().strip()

            if not new_plan:
                messagebox.showwarning("警告", "请选择人员名称")
                return

            if new_plan not in self.shift_schedules:
                messagebox.showwarning("警告", "该计划不存在")
                return

            if not new_type:
                messagebox.showwarning("警告", "请选择请假类型")
                return

            if not new_date:
                messagebox.showwarning("警告", "请输入请假日期")
                return

            try:
                # 验证日期格式
                from datetime import datetime
                datetime.strptime(new_date, '%Y-%m-%d')
            except ValueError:
                messagebox.showwarning("警告", "日期格式不正确，请使用 YYYY-MM-DD 格式")
                return

            # 检查配额并实现自动顺延（编辑时需要排除原记录）
            quota_result = self._check_and_allocate_quota(new_plan, new_date, new_type, exclude_record=target_record)

            if not quota_result['success']:
                # 配额不足，无法修改
                messagebox.showerror("配额不足", quota_result['message'])
                return

            # 获取实际分配的请假类型
            allocated_type = quota_result['allocated_type']

            # 如果发生了自动顺延，需要用户确认
            if quota_result['cascaded']:
                confirm = messagebox.askyesno(
                    "配额自动顺延",
                    f"{quota_result['message']}\n\n是否确认修改此请假记录？"
                )
                if not confirm:
                    return

            # 更新记录（使用实际分配的请假类型）
            target_record.update({
                "plan_name": new_plan,
                "date": new_date,
                "type": allocated_type,
                "note": new_note
            })

            # 刷新界面
            self.update_leave_tree()
            self.update_leave_stats()
            self.update_quota_summary()
            self.update_quota_year_options()  # 更新年份选项
            self.update_current_leave_year_display()  # 更新年份显示
            self.save_data()
            self.update_calendar()
            self.update_status(f"已修改请假记录: {new_plan} {new_date} {allocated_type}")

            edit_window.destroy()

        # 按钮
        ttk.Button(button_frame, text="保存", command=save_changes).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="取消", command=edit_window.destroy).pack(side=tk.LEFT)

        # 设置焦点到第一个输入框
        plan_combo.focus_set()

    def on_leave_tree_double_click(self, event):
        """处理请假记录表格的双击事件"""
        selected_item = self.leave_tree.selection()
        if selected_item:
            self.edit_leave_record(selected_item[0])

    def query_leave_records(self):
        """查询特定人员的请假记录"""
        # 创建查询对话框
        query_window = tk.Toplevel(self.root)
        query_window.title("查询请假记录")
        query_window.geometry("800x600")
        query_window.transient(self.root)
        query_window.grab_set()

        # 计算对话框位置（居中）
        query_window.update_idletasks()
        x = (query_window.winfo_screenwidth() // 2) - (query_window.winfo_width() // 2)
        y = (query_window.winfo_screenheight() // 2) - (query_window.winfo_height() // 2)
        query_window.geometry(f"+{x}+{y}")

        # 创建主框架
        main_frame = ttk.Frame(query_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 查询条件区域
        query_frame = ttk.LabelFrame(main_frame, text="查询条件", padding=10)
        query_frame.pack(fill=tk.X, pady=(0, 10))

        # 人员选择
        ttk.Label(query_frame, text="选择人员:").grid(row=0, column=0, sticky=tk.W, pady=5)
        person_var = tk.StringVar()
        person_combo = ttk.Combobox(query_frame, textvariable=person_var, width=25, state="readonly")
        person_combo['values'] = list(self.shift_schedules.keys())
        if person_combo['values']:
            person_combo.current(0)
        person_combo.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 年份选择（可选）
        ttk.Label(query_frame, text="年份（可选）:").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        year_var = tk.StringVar(value="全部")
        year_combo = ttk.Combobox(query_frame, textvariable=year_var, width=15, state="readonly")
        # 获取所有请假记录中的年份
        years = set()
        for rec in self.leave_records:
            date_str = rec.get('date', '')
            if date_str:
                try:
                    year = date_str.split('-')[0]
                    years.add(year)
                except:
                    pass
        year_combo['values'] = ["全部"] + sorted(list(years), reverse=True)
        year_combo.grid(row=0, column=3, sticky=tk.W, pady=5, padx=(5, 0))

        # 查询按钮
        def do_query():
            person = person_var.get().strip()
            if not person:
                messagebox.showwarning("警告", "请选择人员")
                return

            selected_year = year_var.get()

            # 清空结果树
            for item in result_tree.get_children():
                result_tree.delete(item)

            # 过滤记录
            filtered_records = []
            for rec in self.leave_records:
                if rec.get("plan_name") == person:
                    if selected_year == "全部":
                        filtered_records.append(rec)
                    else:
                        date_str = rec.get('date', '')
                        if date_str.startswith(selected_year):
                            filtered_records.append(rec)

            # 按日期排序（最新的在前）
            filtered_records.sort(key=lambda r: r.get('date', ''), reverse=True)

            # 显示结果
            if not filtered_records:
                messagebox.showinfo("查询结果", f"未找到 {person} 的请假记录")
                return

            for rec in filtered_records:
                result_tree.insert("", tk.END, values=(
                    rec.get("date", ""),
                    rec.get("type", ""),
                    rec.get("note", "")
                ))

            # 更新统计信息
            total_days = len(filtered_records)
            type_stats = {}
            for rec in filtered_records:
                ltype = rec.get("type", "")
                type_stats[ltype] = type_stats.get(ltype, 0) + 1

            stats_text = f"共 {total_days} 条记录"
            if type_stats:
                stats_text += "，类型分布: " + "，".join([f"{k}:{v}天" for k, v in type_stats.items()])
            stats_label.config(text=stats_text)

        ttk.Button(query_frame, text="查询", command=do_query).grid(row=0, column=4, sticky=tk.W, pady=5, padx=(20, 0))

        # 结果显示区域
        result_frame = ttk.LabelFrame(main_frame, text="查询结果", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True)

        # 创建结果树视图
        tree_frame = ttk.Frame(result_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        result_tree = ttk.Treeview(tree_frame, columns=("date", "type", "note"),
                                   show="headings", yscrollcommand=scrollbar.set)
        result_tree.heading("date", text="请假日期")
        result_tree.heading("type", text="类型")
        result_tree.heading("note", text="备注")

        result_tree.column("date", width=120)
        result_tree.column("type", width=120)
        result_tree.column("note", width=300)

        result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=result_tree.yview)

        # 统计信息标签
        stats_label = ttk.Label(result_frame, text="请选择人员并点击查询",
                               font=('Microsoft YaHei UI', 10))
        stats_label.pack(pady=(10, 0))

        # 关闭按钮
        ttk.Button(main_frame, text="关闭", command=query_window.destroy).pack(pady=(10, 0))

    def view_all_leave_records(self):
        """查看所有人员的请假记录"""
        # 创建查看对话框
        view_window = tk.Toplevel(self.root)
        view_window.title("所有人员请假记录")
        view_window.geometry("1000x700")
        view_window.transient(self.root)
        view_window.grab_set()

        # 计算对话框位置（居中）
        view_window.update_idletasks()
        x = (view_window.winfo_screenwidth() // 2) - (view_window.winfo_width() // 2)
        y = (view_window.winfo_screenheight() // 2) - (view_window.winfo_height() // 2)
        view_window.geometry(f"+{x}+{y}")

        # 创建主框架
        main_frame = ttk.Frame(view_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 筛选条件区域
        filter_frame = ttk.LabelFrame(main_frame, text="筛选条件", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        # 年份筛选
        ttk.Label(filter_frame, text="年份:").grid(row=0, column=0, sticky=tk.W, pady=5)
        year_var = tk.StringVar(value="全部")
        year_combo = ttk.Combobox(filter_frame, textvariable=year_var, width=15, state="readonly")
        # 获取所有请假记录中的年份
        years = set()
        for rec in self.leave_records:
            date_str = rec.get('date', '')
            if date_str:
                try:
                    year = date_str.split('-')[0]
                    years.add(year)
                except:
                    pass
        year_combo['values'] = ["全部"] + sorted(list(years), reverse=True)
        year_combo.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(5, 0))

        # 人员筛选
        ttk.Label(filter_frame, text="人员:").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        person_var = tk.StringVar(value="全部")
        person_combo = ttk.Combobox(filter_frame, textvariable=person_var, width=20, state="readonly")
        person_combo['values'] = ["全部"] + list(self.shift_schedules.keys())
        person_combo.grid(row=0, column=3, sticky=tk.W, pady=5, padx=(5, 0))

        # 类型筛选
        ttk.Label(filter_frame, text="类型:").grid(row=0, column=4, sticky=tk.W, pady=5, padx=(20, 0))
        type_var = tk.StringVar(value="全部")
        type_combo = ttk.Combobox(filter_frame, textvariable=type_var, width=15, state="readonly")
        type_combo['values'] = ["全部"] + self.leave_types
        type_combo.grid(row=0, column=5, sticky=tk.W, pady=5, padx=(5, 0))

        # 应用筛选按钮
        def apply_filter():
            selected_year = year_var.get()
            selected_person = person_var.get()
            selected_type = type_var.get()

            # 清空结果树
            for item in result_tree.get_children():
                result_tree.delete(item)

            # 过滤记录
            filtered_records = []
            for rec in self.leave_records:
                # 年份筛选
                if selected_year != "全部":
                    date_str = rec.get('date', '')
                    if not date_str.startswith(selected_year):
                        continue

                # 人员筛选
                if selected_person != "全部":
                    if rec.get("plan_name") != selected_person:
                        continue

                # 类型筛选
                if selected_type != "全部":
                    if rec.get("type") != selected_type:
                        continue

                filtered_records.append(rec)

            # 按日期排序（最新的在前）
            filtered_records.sort(key=lambda r: (r.get('date', ''), r.get('plan_name', '')), reverse=True)

            # 显示结果
            if not filtered_records:
                messagebox.showinfo("筛选结果", "未找到符合条件的请假记录")
                stats_label.config(text="未找到符合条件的记录")
                return

            for rec in filtered_records:
                result_tree.insert("", tk.END, values=(
                    rec.get("plan_name", ""),
                    rec.get("date", ""),
                    rec.get("type", ""),
                    rec.get("note", "")
                ))

            # 更新统计信息
            total_days = len(filtered_records)
            person_stats = {}
            type_stats = {}
            for rec in filtered_records:
                person = rec.get("plan_name", "")
                ltype = rec.get("type", "")
                person_stats[person] = person_stats.get(person, 0) + 1
                type_stats[ltype] = type_stats.get(ltype, 0) + 1

            stats_text = f"共 {total_days} 条记录"
            if person_stats:
                stats_text += f"，涉及 {len(person_stats)} 人"
            if type_stats:
                stats_text += "，类型分布: " + "，".join([f"{k}:{v}天" for k, v in type_stats.items()])
            stats_label.config(text=stats_text)

        ttk.Button(filter_frame, text="应用筛选", command=apply_filter).grid(row=0, column=6, sticky=tk.W, pady=5, padx=(20, 0))
        ttk.Button(filter_frame, text="重置", command=lambda: [year_var.set("全部"), person_var.set("全部"), type_var.set("全部"), apply_filter()]).grid(row=0, column=7, sticky=tk.W, pady=5, padx=(5, 0))

        # 结果显示区域
        result_frame = ttk.LabelFrame(main_frame, text="请假记录列表", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True)

        # 创建结果树视图
        tree_frame = ttk.Frame(result_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        result_tree = ttk.Treeview(tree_frame, columns=("person", "date", "type", "note"),
                                   show="headings",
                                   yscrollcommand=scrollbar_y.set,
                                   xscrollcommand=scrollbar_x.set)
        result_tree.heading("person", text="人员名称")
        result_tree.heading("date", text="请假日期")
        result_tree.heading("type", text="类型")
        result_tree.heading("note", text="备注")

        result_tree.column("person", width=120)
        result_tree.column("date", width=120)
        result_tree.column("type", width=120)
        result_tree.column("note", width=300)

        result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.config(command=result_tree.yview)
        scrollbar_x.config(command=result_tree.xview)

        # 统计信息标签
        stats_label = ttk.Label(result_frame, text="",
                               font=('Microsoft YaHei UI', 10))
        stats_label.pack(pady=(10, 0))

        # 关闭按钮
        ttk.Button(main_frame, text="关闭", command=view_window.destroy).pack(pady=(10, 0))

        # 初始加载所有记录
        apply_filter()

    # 班次类型管理方法
    def update_shift_type_tree(self):
        """更新班次类型树视图"""
        self.shift_type_tree.delete(*self.shift_type_tree.get_children())
        for name, info in self.shift_types.items():
            self.shift_type_tree.insert("", tk.END, values=(
                name, info["start_time"], info["end_time"], info["color"]))
    
    def add_shift_type(self):
        """添加班次类型"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加班次类型")
        dialog.transient(self.root)
        dialog.grab_set()

        form = ttk.Frame(dialog, padding=10)
        form.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(form, text="班次名称").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(form, textvariable=name_var, width=20)
        name_entry.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(form, text="开始时间 (HH:MM)").grid(row=1, column=0, sticky=tk.W, pady=5)
        start_var = tk.StringVar()
        start_entry = ttk.Entry(form, textvariable=start_var, width=10)
        start_entry.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(form, text="结束时间 (HH:MM)").grid(row=2, column=0, sticky=tk.W, pady=5)
        end_var = tk.StringVar()
        end_entry = ttk.Entry(form, textvariable=end_var, width=10)
        end_entry.grid(row=2, column=1, sticky=tk.W)

        ttk.Label(form, text="显示颜色 (#RRGGBB)").grid(row=3, column=0, sticky=tk.W, pady=5)
        color_var = tk.StringVar(value="#FFFFFF")
        color_entry = ttk.Entry(form, textvariable=color_var, width=10)
        color_entry.grid(row=3, column=1, sticky=tk.W)

        def on_save():
            name = name_var.get().strip()
            start_time = start_var.get().strip()
            end_time = end_var.get().strip()
            color = color_var.get().strip()

            # 收集所有验证错误
            errors = []

            # 验证班次名称
            is_valid, error_msg = DataValidator.validate_shift_name(name, self.shift_types.keys())
            if not is_valid:
                errors.append(error_msg)

            # 验证开始时间
            if start_time:
                is_valid, error_msg = DataValidator.validate_time_format(start_time)
                if not is_valid:
                    errors.append(f"开始时间: {error_msg}")
            else:
                start_time = "00:00"

            # 验证结束时间
            if end_time:
                is_valid, error_msg = DataValidator.validate_time_format(end_time)
                if not is_valid:
                    errors.append(f"结束时间: {error_msg}")
            else:
                end_time = "00:00"

            # 验证颜色格式
            if color:
                if not color.startswith('#'):
                    color = "#" + color
                is_valid, error_msg = DataValidator.validate_color_format(color)
                if not is_valid:
                    errors.append(f"颜色: {error_msg}")
            else:
                color = "#FFFFFF"

            # 处理验证错误
            if not self.error_handler.handle_validation_errors(errors):
                return

            # 所有验证通过，保存数据
            self.shift_types[name] = {
                "start_time": start_time,
                "end_time": end_time,
                "color": color
            }
            self.update_shift_type_tree()
            self.save_data()
            self.update_status(f"已添加班次: {name}")
            dialog.destroy()

        action = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        action.grid(row=1, column=0, sticky=tk.EW)

        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(action, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)

        # 布局伸缩
        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
    
    def edit_shift_type(self):
        """编辑班次类型"""
        if not (selected := self.shift_type_tree.selection()):
            messagebox.showwarning("警告", "请先选择班次类型")
            return

        item_id = selected[0]
        name = self.shift_type_tree.item(item_id)["values"][0]
        shift_info = self.shift_types.get(name)
        if not shift_info:
            messagebox.showerror("错误", "未找到所选班次类型")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("编辑班次类型")
        dialog.transient(self.root)
        dialog.grab_set()

        form = ttk.Frame(dialog, padding=10)
        form.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(form, text="班次名称").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar(value=name)
        name_entry = ttk.Entry(form, textvariable=name_var, width=20)
        name_entry.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(form, text="开始时间 (HH:MM)").grid(row=1, column=0, sticky=tk.W, pady=5)
        start_var = tk.StringVar(value=shift_info["start_time"])
        start_entry = ttk.Entry(form, textvariable=start_var, width=10)
        start_entry.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(form, text="结束时间 (HH:MM)").grid(row=2, column=0, sticky=tk.W, pady=5)
        end_var = tk.StringVar(value=shift_info["end_time"])
        end_entry = ttk.Entry(form, textvariable=end_var, width=10)
        end_entry.grid(row=2, column=1, sticky=tk.W)

        ttk.Label(form, text="显示颜色 (#RRGGBB)").grid(row=3, column=0, sticky=tk.W, pady=5)
        color_var = tk.StringVar(value=shift_info["color"])
        color_entry = ttk.Entry(form, textvariable=color_var, width=10)
        color_entry.grid(row=3, column=1, sticky=tk.W)

        def on_save():
            new_name = name_var.get().strip()
            start_time = start_var.get().strip()
            end_time = end_var.get().strip()
            color = color_var.get().strip()

            if not new_name:
                messagebox.showwarning("警告", "请输入班次名称")
                return

            if new_name != name and new_name in self.shift_types:
                messagebox.showwarning("警告", "班次名称已存在")
                return

            if new_name != name:
                # 重命名情况：先删除旧的，添加新的
                del self.shift_types[name]

            if not start_time:
                start_time = "00:00"
            if not end_time:
                end_time = "00:00"
            if not color.startswith('#'):
                color = "#" + color

            self.shift_types[new_name] = {
                "start_time": start_time,
                "end_time": end_time,
                "color": color
            }
            self.update_shift_type_tree()
            self.save_data()
            self.update_status(f"已更新班次: {name} -> {new_name}")
            dialog.destroy()

        action = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        action.grid(row=1, column=0, sticky=tk.EW)

        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(action, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)

        # 布局伸缩
        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
    
    def delete_shift_type(self):
        """删除班次类型"""
        if not (selected := self.shift_type_tree.selection()):
            messagebox.showwarning("警告", "请先选择班次类型")
            return
        
        name = self.shift_type_tree.item(selected[0])["values"][0]
        if messagebox.askyesno("确认", f"确定要删除班次 '{name}' 吗？"):
            del self.shift_types[name]
            self.update_shift_type_tree()
            self.save_data()
            self.update_status(f"已删除班次: {name}")
    
    # 排班计划管理方法
    def update_schedule_tree(self):
        """更新排班计划树视图"""
        self.schedule_tree.delete(*self.schedule_tree.get_children())
        for name, info in self.shift_schedules.items():
        # 使用人员名称作为iid，确保后续通过选择项可稳定取回名称
            self.schedule_tree.insert("", tk.END, iid=str(name), values=(
                name, 
                " / ".join(info["shift_pattern"]),  # 用斜杠显示更清晰
                info["start_date"]
            ))
        # 同步请假页计划下拉
        self.update_leave_plan_combo()

    
    def create_schedule(self):
        """创建排班计划"""
        dialog = tk.Toplevel(self.root)
        dialog.title("新建排班计划")
        dialog.transient(self.root)
        dialog.grab_set()

        # 表单区
        form = ttk.Frame(dialog, padding=10)
        form.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(form, text="人员名称").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(form, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(form, text="开始日期").grid(row=1, column=0, sticky=tk.W, pady=5)
        start_date = DateEntry(form, date_pattern='yyyy-MM-dd')
        start_date.set_date(datetime.date.today())
        start_date.grid(row=1, column=1, sticky=tk.W)

        # 轮班模式编辑
        ttk.Label(form, text="可选班次").grid(row=2, column=0, sticky=tk.W, pady=(10, 5))
        ttk.Label(form, text="轮班模式").grid(row=2, column=1, sticky=tk.W, pady=(10, 5))

        # 左侧：可选班次列表
        available_frame = ttk.Frame(form)
        available_frame.grid(row=3, column=0, sticky=tk.NSEW)
        available_list = tk.Listbox(available_frame, height=8, exportselection=False)
        available_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        for shift_name in self.shift_types.keys():
            available_list.insert(tk.END, shift_name)
        avail_scroll = ttk.Scrollbar(available_frame, command=available_list.yview)
        avail_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        available_list.config(yscrollcommand=avail_scroll.set)

        # 右侧：已选（轮班模式）列表
        pattern_frame = ttk.Frame(form)
        pattern_frame.grid(row=3, column=1, sticky=tk.NSEW)
        pattern_list = tk.Listbox(pattern_frame, height=8, exportselection=False)
        pattern_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pattern_scroll = ttk.Scrollbar(pattern_frame, command=pattern_list.yview)
        pattern_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        pattern_list.config(yscrollcommand=pattern_scroll.set)

        # 中间操作按钮
        btns = ttk.Frame(form)
        btns.grid(row=3, column=2, padx=10)

        def add_to_pattern():
            sel = available_list.curselection()
            if not sel:
                return
            value = available_list.get(sel[0])
            pattern_list.insert(tk.END, value)

        def remove_from_pattern():
            sel = pattern_list.curselection()
            if not sel:
                return
            pattern_list.delete(sel[0])

        def move_up():
            sel = pattern_list.curselection()
            if not sel or sel[0] == 0:
                return
            idx = sel[0]
            value = pattern_list.get(idx)
            pattern_list.delete(idx)
            pattern_list.insert(idx - 1, value)
            pattern_list.selection_set(idx - 1)

        def move_down():
            sel = pattern_list.curselection()
            if not sel or sel[0] == pattern_list.size() - 1:
                return
            idx = sel[0]
            value = pattern_list.get(idx)
            pattern_list.delete(idx)
            pattern_list.insert(idx + 1, value)
            pattern_list.selection_set(idx + 1)

        ttk.Button(btns, text=">>", command=add_to_pattern).grid(row=0, column=0, pady=2)
        ttk.Button(btns, text="<<", command=remove_from_pattern).grid(row=1, column=0, pady=2)
        ttk.Button(btns, text="上移", command=move_up).grid(row=2, column=0, pady=8)
        ttk.Button(btns, text="下移", command=move_down).grid(row=3, column=0)

        # 操作区：保存/取消
        action = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        action.grid(row=1, column=0, sticky=tk.EW)

        def on_save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("警告", "请输入人员名称")
                return
            if name in self.shift_schedules:
                messagebox.showwarning("警告", "该人员名称已存在")
                return
            pattern = [pattern_list.get(i) for i in range(pattern_list.size())]
            if not pattern:
                messagebox.showwarning("警告", "请至少添加一个班次到轮班模式")
                return
            info = {
                "shift_pattern": pattern,
                "start_date": start_date.get_date().strftime('%Y-%m-%d')
            }
            self.shift_schedules[name] = info
            self.update_schedule_tree()
            self.update_leave_plan_combo()
            self.save_data()
            self.update_status(f"已创建人员: {name}")
            dialog.destroy()

        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(action, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)

        # 布局伸缩
        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(0, weight=1)
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(2, weight=0)

        name_entry.focus_set()
    
    def edit_schedule(self):
        """编辑排班计划"""
        if not (selected := self.schedule_tree.selection()):
            messagebox.showwarning("警告", "请先选择排班计划")
            return
        
        name = selected[0]  # iid即人员名称
        info = self.shift_schedules.get(name)
        if not info:
            messagebox.showerror("错误", "未找到所选人员")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("编辑排班计划")
        dialog.transient(self.root)
        dialog.grab_set()

        form = ttk.Frame(dialog, padding=10)
        form.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(form, text="人员名称").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar(value=name)
        name_entry = ttk.Entry(form, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(form, text="开始日期").grid(row=1, column=0, sticky=tk.W, pady=5)
        start_date = DateEntry(form, date_pattern='yyyy-MM-dd')
        try:
            start_date.set_date(datetime.datetime.strptime(info.get("start_date", ''), '%Y-%m-%d').date())
        except Exception:
            start_date.set_date(datetime.date.today())
        start_date.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(form, text="可选班次").grid(row=2, column=0, sticky=tk.W, pady=(10, 5))
        ttk.Label(form, text="轮班模式").grid(row=2, column=1, sticky=tk.W, pady=(10, 5))

        available_frame = ttk.Frame(form)
        available_frame.grid(row=3, column=0, sticky=tk.NSEW)
        available_list = tk.Listbox(available_frame, height=8, exportselection=False)
        available_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        for shift_name in self.shift_types.keys():
            available_list.insert(tk.END, shift_name)
        avail_scroll = ttk.Scrollbar(available_frame, command=available_list.yview)
        avail_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        available_list.config(yscrollcommand=avail_scroll.set)

        pattern_frame = ttk.Frame(form)
        pattern_frame.grid(row=3, column=1, sticky=tk.NSEW)
        pattern_list = tk.Listbox(pattern_frame, height=8, exportselection=False)
        pattern_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pattern_scroll = ttk.Scrollbar(pattern_frame, command=pattern_list.yview)
        pattern_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        pattern_list.config(yscrollcommand=pattern_scroll.set)
        for s in info.get("shift_pattern", []):
            pattern_list.insert(tk.END, s)

        btns = ttk.Frame(form)
        btns.grid(row=3, column=2, padx=10)

        def add_to_pattern():
            sel = available_list.curselection()
            if not sel:
                return
            value = available_list.get(sel[0])
            pattern_list.insert(tk.END, value)

        def remove_from_pattern():
            sel = pattern_list.curselection()
            if not sel:
                return
            pattern_list.delete(sel[0])

        def move_up():
            sel = pattern_list.curselection()
            if not sel or sel[0] == 0:
                return
            idx = sel[0]
            value = pattern_list.get(idx)
            pattern_list.delete(idx)
            pattern_list.insert(idx - 1, value)
            pattern_list.selection_set(idx - 1)

        def move_down():
            sel = pattern_list.curselection()
            if not sel or sel[0] == pattern_list.size() - 1:
                return
            idx = sel[0]
            value = pattern_list.get(idx)
            pattern_list.delete(idx)
            pattern_list.insert(idx + 1, value)
            pattern_list.selection_set(idx + 1)

        ttk.Button(btns, text=">>", command=add_to_pattern).grid(row=0, column=0, pady=2)
        ttk.Button(btns, text="<<", command=remove_from_pattern).grid(row=1, column=0, pady=2)
        ttk.Button(btns, text="上移", command=move_up).grid(row=2, column=0, pady=8)
        ttk.Button(btns, text="下移", command=move_down).grid(row=3, column=0)

        action = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        action.grid(row=1, column=0, sticky=tk.EW)

        def on_save():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showwarning("警告", "请输入人员名称")
                return
            if new_name != name and new_name in self.shift_schedules:
                messagebox.showwarning("警告", "该人员名称已存在")
                return
            new_pattern = [pattern_list.get(i) for i in range(pattern_list.size())]
            if not new_pattern:
                messagebox.showwarning("警告", "请至少添加一个班次到轮班模式")
                return

            new_start = start_date.get_date().strftime('%Y-%m-%d')
            changed_core = (new_name != name) or (new_pattern != info.get("shift_pattern")) or (new_start != info.get("start_date"))

            # 应用修改（处理重命名）
            updated = {
                "shift_pattern": new_pattern,
                "start_date": new_start
            }
            # 如果原有shifts在核心要素变化后需重算
            regenerate = False
            if changed_core and info.get("shifts"):
                if messagebox.askyesno("是否重新生成", "检测到名称/开始日期/模式变更，是否立即重新生成一年排班？"):
                    regenerate = True

            if regenerate:
                # 直接复用生成逻辑
                shifts_map = {}
                try:
                    sd = datetime.datetime.strptime(new_start, "%Y-%m-%d").date()
                except Exception:
                    sd = datetime.date.today()
                total_days = 365
                for i in range(total_days):
                    d = sd + datetime.timedelta(days=i)
                    pattern_idx = i % len(new_pattern)
                    shifts_map[d.strftime("%Y-%m-%d")] = new_pattern[pattern_idx]
                updated["shifts"] = shifts_map
            else:
                # 不重算则保留原有shifts（若核心变更则清空，避免误差）
                if changed_core:
                    updated["shifts"] = {}
                else:
                    updated["shifts"] = info.get("shifts", {})

            # 执行字典更新（考虑名称变更）
            if new_name != name:
                self.shift_schedules.pop(name, None)
                self.shift_schedules[new_name] = updated
            else:
                self.shift_schedules[name] = updated

            # 同步当前计划
            if self.current_schedule is info or (self.current_schedule and self.current_schedule == info):
                self.current_schedule = updated
                self.current_plan_name = new_name

            self.update_schedule_tree()
            self.update_leave_plan_combo()
            self.save_data()
            self.update_calendar()
            self.update_status(f"已保存人员: {new_name}")
            dialog.destroy()

        ttk.Button(action, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(action, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)

        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(0, weight=1)
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(2, weight=0)

        name_entry.focus_set()
    
    def delete_schedule(self):
        """删除排班计划"""
        if not (selected := self.schedule_tree.selection()):
            messagebox.showwarning("警告", "请先选择排班计划")
            return
        
        name = selected[0]  # iid即人员名称
        if messagebox.askyesno("确认", f"确定要删除排班计划 '{name}' 吗？"):
            if name not in self.shift_schedules:
                messagebox.showerror("错误", "未在数据中找到该人员，列表将被刷新")
                self.update_schedule_tree()
                return
            del self.shift_schedules[name]
            if self.current_schedule and self.current_schedule is not None:
                # 如果当前计划被删除，清空当前计划
                selected_name = name
                cur_pattern = self.current_schedule.get("shift_pattern")
                cur_start = self.current_schedule.get("start_date")
                # 保守重置：直接置空
                self.current_schedule = None
                if self.current_plan_name == name:
                    self.current_plan_name = None
            self.update_schedule_tree()
            self.update_leave_plan_combo()
            self.save_data()
            self.update_status(f"已删除人员: {name}")
    
    def generate_schedule(self):
        """生成排班"""
        if not (selected := self.schedule_tree.selection()):
            messagebox.showwarning("警告", "请先选择排班计划")
            return
        
        name = selected[0]  # iid即人员名称
        info = self.shift_schedules.get(name)
        if not info:
            messagebox.showerror("错误", "未找到所选人员")
            return

        pattern = info.get("shift_pattern", [])
        start_date_str = info.get("start_date")
        if not pattern or not start_date_str:
            messagebox.showwarning("警告", "人员计划不完整，缺少轮班模式或开始日期")
            return

        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except Exception:
            messagebox.showerror("错误", "开始日期格式错误，应为 YYYY-MM-DD")
            return

        # 生成从开始日期起一年的排班映射
        shifts_map = {}
        total_days = 1825
        for i in range(total_days):
            d = start_date + datetime.timedelta(days=i)
            pattern_idx = i % len(pattern)
            shifts_map[d.strftime("%Y-%m-%d")] = pattern[pattern_idx]

        # 保存到计划并设为当前计划
        info["shifts"] = shifts_map
        self.shift_schedules[name] = info
        self.current_schedule = info
        self.current_plan_name = name
        self.save_data()
        self.update_calendar()
        self.update_year_options()
        self.sync_year_combo()
        self.update_month_options()
        self.sync_month_combo()
              # 切换到日历视图标签页
        calendar_tab_index = 4  # 日历视图是第5个标签页（索引从0开始）
        self.notebook.select(calendar_tab_index)
        self.update_status(f"已生成排班并设置当前人员: {name}")
    
    def select_current_person(self, event):
        """通过双击选择当前人员"""
        if not (selected := self.schedule_tree.selection()):
            return
        
        name = selected[0]  # iid即人员名称
        info = self.shift_schedules.get(name)
        if not info:
            messagebox.showerror("错误", "未找到所选人员")
            return
        
        # 设置为当前人员
        self.current_schedule = info
        self.current_plan_name = name
        
        # 刷新日历视图以显示新的人员信息
        self.update_calendar()
        self.update_status(f"已选择当前人员: {name}")
    
    def update_leave_plan_combo(self):
        """同步请假页计划下拉选项"""
        if hasattr(self, "leave_plan_combo"):
            plan_names = list(self.shift_schedules.keys())
            self.leave_plan_combo["values"] = plan_names
            # 若当前值不在新列表中，则回退到第一个或空
            current = self.leave_plan_var.get() if hasattr(self, "leave_plan_var") else ""
            if current not in plan_names:
                self.leave_plan_var.set(plan_names[0] if plan_names else "")
    
    # 日历操作方法
    def update_calendar(self):
        """更新日历显示"""
        # 检查calendar_container是否存在
        if not hasattr(self, 'calendar_container'):
            return

        # 使用临时禁用更新机制，减少GUI更新次数
        for widget in self.calendar_container.winfo_children():
            widget.destroy()
        
        year, month = self.current_date.year, self.current_date.month
        self.month_year_var.set(f"{year}年{month}月")
        
        # 获取今天的日期用于高亮显示
        today = datetime.date.today()
        
        # 在年月标题下方显示当前人员名称
        person_info_frame = ttk.Frame(self.calendar_container)
        person_info_frame.grid(row=0, column=0, columnspan=7, sticky=tk.EW, pady=(0, 3))  # 减少下边距

        if self.current_plan_name:
            person_label = ttk.Label(person_info_frame, text=f"当前人员: {self.current_plan_name}",
                                   font=("Arial", 11, "bold"), foreground="#0066CC")  # 稍微减小字体
            person_label.pack()
        else:
            no_person_label = ttk.Label(person_info_frame, text="未选择人员",
                                     font=("Arial", 11, "bold"), foreground="#FF6666")  # 稍微减小字体
            no_person_label.pack()

        # 星期标题
        weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        for col, day in enumerate(weekdays):
            ttk.Label(self.calendar_container, text=day).grid(row=1, column=col, sticky=tk.NSEW, padx=1, pady=0)  # 减少垂直间距
        
        # 日期格子
        first_day, num_days = calendar.monthrange(year, month)
        day_num = 1
        for row in range(2, 8):  # 从第2行开始，因为第0行是人员信息，第1行是星期标题
            for col in range(7):
                if (row == 2 and col < first_day) or day_num > num_days:
                    continue
                
                date_str = f"{year}-{month:02d}-{day_num:02d}"
                date_key = f"{month:02d}-{day_num:02d}"
                current_date = datetime.date(year, month, day_num)
                
                # 判断是否为今天
                is_today = current_date == today

                # 周末底色（周六日轻微灰蓝，以增强可读性）
                is_weekend = current_date.weekday() >= 5
                cell_bg = "#F2F6FC" if is_weekend else "#FFFFFF"
                border_color = "#0066CC" if is_today else "#CCCCCC"
                
                # 统一使用 tk.Frame 以便自定义背景色
                frame = tk.Frame(self.calendar_container,
                                 relief=tk.RIDGE,
                                 borderwidth=2 if is_today else 1,
                                 bg="#E6F3FF" if is_today else cell_bg,
                                 highlightbackground=border_color,
                                 highlightthickness=2 if is_today else 1)
                
                frame.grid(row=row, column=col, sticky=tk.NSEW, padx=1, pady=1)  # 减少间距
                
                # 显示日期 - 今天使用特殊样式
                if is_today:
                    # 今天的日期使用粗体、蓝色文字和特殊背景
                    label = tk.Label(frame, text=str(day_num), font=("Arial", 10, "bold"), 
                                   fg="#0066CC", bg="#E6F3FF")
                    # 添加"今日"标识
                    today_label = tk.Label(frame, text="今日", font=("Arial", 8, "bold"), 
                                         fg="#FFFFFF", bg="#0066CC")
                    today_label.pack(anchor=tk.NE, padx=2, pady=1)
                else:
                    # 普通日期的标准样式
                    label = tk.Label(frame, text=str(day_num), bg=cell_bg)
                
                label.pack(anchor=tk.NW, padx=4, pady=2)
                
                # 预先计算可能用到的数据 - 优化性能
                holiday = self.holidays.get(str(year), {}).get(date_key)
                
                # 标记节假日（重要节假日突出显示）
                if self.show_holidays.get() and holiday:
                    if is_today:
                        label.config(foreground="#CC0000", font=("Arial", 10, "bold"))
                        holiday_label = tk.Label(frame, text=f"今日·{holiday}", 
                                              font=("Arial", 8, "bold"), 
                                              fg="#FFFFFF", bg="#FF4444")
                        holiday_label.pack(fill=tk.X, padx=2, pady=(2, 2))
                        _SimpleTooltip(holiday_label, f"节假日: {holiday}")
                    else:
                        label.config(foreground="red")
                        h_lbl = tk.Label(frame, text=holiday, bg="#FF6666", fg="#FFFFFF")
                        h_lbl.pack(fill=tk.X, padx=2, pady=(2, 2))
                        _SimpleTooltip(h_lbl, f"节假日: {holiday}")
                
                # 显示排班
                shift = None
                if self.current_schedule:
                    shift = self.current_schedule["shifts"].get(date_str)
                
                if shift and (shift_info := self.shift_types.get(shift)):
                    # 今天的排班使用更醒目的样式
                    if is_today:
                        shift_label = tk.Label(frame, text=shift, 
                                             background=shift_info["color"], 
                                             font=("Arial", 9, "bold"),
                                             fg="#000000" if shift_info["color"] != "#000000" else "#FFFFFF")
                    else:
                        shift_label = tk.Label(frame, text=shift, background=shift_info["color"])
                    shift_label.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
                    # 悬浮提示显示班次时间
                    st = shift_info.get("start_time", "")
                    et = shift_info.get("end_time", "")
                    tip = f"{shift}  时间: {st} - {et}".strip()
                    _SimpleTooltip(shift_label, tip)

                # 减少查找请假记录的次数 - 使用预计算的字典
                if self.show_leaves.get() and self.current_plan_name:
                    # 构建当天请假记录的快速查找
                    daily_leave = None
                    for rec in self.leave_records:
                        if (rec.get("plan_name") == self.current_plan_name and 
                            rec.get("date") == date_str):
                            daily_leave = rec.get("type", "请假")
                            break

                    if daily_leave:
                        # 今天的请假使用更醒目的样式
                        if is_today:
                            lbl = tk.Label(frame, text=f"今日·{daily_leave}", 
                                        bg="#FF0000", fg="#FFFFFF", 
                                        font=("Arial", 8, "bold"))
                        else:
                            lbl = tk.Label(frame, text=daily_leave, bg="#FF0000", fg="#FFFFFF")
                        lbl.pack(fill=tk.X, padx=4, pady=(0, 4))
                        _SimpleTooltip(lbl, f"请假: {daily_leave}")
                
                day_num += 1
    
    def prev_month(self):
        """显示上个月"""
        year = self.current_date.year
        month = self.current_date.month
        self.current_date = datetime.date(year - (month==1), 12 if month==1 else month-1, 1)
        self.update_calendar()
        self.sync_year_combo()
        self.sync_month_combo()
    
    def next_month(self):
        """显示下个月"""
        year = self.current_date.year
        month = self.current_date.month
        self.current_date = datetime.date(year + (month==12), 1 if month==12 else month+1, 1)
        self.update_calendar()
        self.sync_year_combo()
        self.sync_month_combo()
    
    def show_current_month(self):
        """显示当前月份"""
        self.current_date = datetime.date.today()
        self.update_calendar()
        self.sync_year_combo()
        self.sync_month_combo()
    
    def go_to_today(self):
        """跳转到今日并高亮显示"""
        today = datetime.date.today()
        self.current_date = today
        self.update_calendar()
        self.sync_year_combo()
        self.sync_month_combo()
        self.update_status(f"已跳转到今日: {today.strftime('%Y年%m月%d日')}")
    
    # 数据持久化方法
    def load_data(self):
        """加载保存的数据"""
        try:
            if os.path.exists("shift_data.json"):
                with open("shift_data.json", "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.shift_types = data.get("shift_types", self.shift_types)
                    self.shift_schedules = data.get("schedules", {})
                    self.swap_records = data.get("swap_records", {})  # 加载调换班记录
                    self.leave_types = data.get("leave_types", self.leave_types)
                    self.leave_records = data.get("leave_records", self.leave_records)
                    self.leave_quotas = data.get("leave_quotas", self.leave_quotas)
                    self.holidays.update(data.get("holidays", {}))
                    # 加载字体设置
                    self.font_family.set(data.get("font_family", "Microsoft YaHei UI"))
                    self.font_size.set(data.get("font_size", 10))
                    self.multi_calendar_font_size.set(data.get("multi_calendar_font_size", 9))
                    # 加载主题设置
                    self.theme_var.set(data.get("theme", "light"))
                    # 加载托盘设置
                    self.minimize_to_tray.set(data.get("minimize_to_tray", False))
            # 加载后刷新界面列表
            if hasattr(self, "shift_type_tree"):
                self.update_shift_type_tree()
            if hasattr(self, "schedule_tree"):
                self.update_schedule_tree()
            if hasattr(self, "leave_type_list"):
                self.update_leave_type_list()
            if hasattr(self, "leave_tree"):
                self.update_leave_tree()
            self.update_leave_plan_combo()
            # 同步"请假记录-类型"下拉框
            if hasattr(self, "leave_type_combo"):
                self.leave_type_combo["values"] = self.leave_types
                current = self.leave_type_var.get() if hasattr(self, "leave_type_var") else ""
                if current not in self.leave_types:
                    if self.leave_types:
                        self.leave_type_var.set(self.leave_types[0])
                        try:
                            self.leave_type_combo.current(0)
                        except Exception:
                            pass
                    else:
                        self.leave_type_var.set("")
            # 刷新统计
            if hasattr(self, "leave_stats_tree"):
                self.update_leave_stats()
            if hasattr(self, "leave_quota_tree"):
                self.update_quota_summary()
            # 刷新当月休假日历的成员和请假类型下拉菜单
            if hasattr(self, "holiday_member_combo"):
                self._init_holiday_calendar_controls()
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败: {str(e)}")

    def save_data(self):
        """保存数据到文件"""
        data = {
            "shift_types": self.shift_types,
            "schedules": self.shift_schedules,
            "swap_records": self.swap_records,  # 保存调换班记录
            "leave_types": self.leave_types,
            "leave_records": self.leave_records,
            "leave_quotas": self.leave_quotas,
            "holidays": self.holidays,
            "font_family": self.font_family.get(),
            "font_size": self.font_size.get(),
            "multi_calendar_font_size": self.multi_calendar_font_size.get(),
            "theme": self.theme_var.get(),
            "minimize_to_tray": self.minimize_to_tray.get()
        }
        with open("shift_data.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def update_status(self, message):
        """更新状态栏信息"""
        self.status_var.set(message)

    # ==================== 当月休假日历数据访问方法 ====================

    def get_holiday_calendar_data(self, year, month):
        """获取指定年月的当月休假日历数据

        Args:
            year (int): 年份
            month (int): 月份

        Returns:
            dict: 日历数据，格式为 {day: [leave_records]}
        """
        calendar_data = {}

        # 获取指定月份的天数
        days_in_month = calendar.monthrange(year, month)[1]

        # 初始化每一天的数据
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            calendar_data[day] = []

        # 填入请假记录
        for record in self.leave_records:
            try:
                record_date = datetime.datetime.strptime(record['date'], "%Y-%m-%d")
                if record_date.year == year and record_date.month == month:
                    day = record_date.day
                    member_name = record['plan_name']
                    date_str = record['date']

                    # 查询该人员当天的排班类型
                    shift_type = None
                    shift_color = "#E1E8ED"  # 默认颜色
                    if member_name in self.shift_schedules:
                        member_shifts = self.shift_schedules[member_name].get('shifts', {})
                        if date_str in member_shifts:
                            shift_type = member_shifts[date_str]
                            # 获取班次颜色
                            if shift_type in self.shift_types:
                                shift_color = self.shift_types[shift_type].get('color', "#E1E8ED")

                    calendar_data[day].append({
                        'name': member_name,
                        'type': record['type'],
                        'note': record.get('note', ''),
                        'date': record['date'],
                        'shift': shift_type,  # 排班类型
                        'shift_color': shift_color  # 排班颜色
                    })
            except (ValueError, KeyError):
                continue

        return calendar_data

    def get_all_members_for_holiday_calendar(self):
        """获取所有成员列表，用于当月休假日历显示

        Returns:
            list: 成员名称列表
        """
        members = set()

        # 从排班计划中获取成员
        members.update(self.shift_schedules.keys())

        # 从请假记录中获取成员
        for record in self.leave_records:
            if 'plan_name' in record:
                members.add(record['plan_name'])

        # 返回排序后的成员列表
        return sorted(list(members))

    def get_leave_types_for_holiday_calendar(self):
        """获取所有请假类型，用于当月休假日历筛选

        Returns:
            list: 请假类型列表
        """
        leave_types = set()

        # 从配置的请假类型中获取
        leave_types.update(self.leave_types)

        # 从请假记录中获取实际使用的类型
        for record in self.leave_records:
            if 'type' in record:
                leave_types.add(record['type'])

        # 返回排序后的请假类型列表
        return sorted(list(leave_types))

    def get_holiday_statistics(self, member_name=None, year=None, month=None, leave_type=None):
        """获取休假统计数据

        Args:
            member_name (str, optional): 成员名称，None表示统计所有成员
            year (int, optional): 年份，None表示统计所有年份
            month (int, optional): 月份，None表示统计全年
            leave_type (str, optional): 请假类型，None表示统计所有类型

        Returns:
            dict: 统计数据
        """
        stats = {
            'total_days': 0,
            'records_by_type': {},
            'records_by_month': {},
            'records_by_member': {}
        }

        for record in self.leave_records:
            try:
                # 筛选条件 - 确保处理None值和"全部"值
                if member_name and member_name != "全部成员" and record.get('plan_name') != member_name:
                    continue
                if leave_type and leave_type != "全部类型" and record.get('type') != leave_type:
                    continue

                record_date = datetime.datetime.strptime(record['date'], "%Y-%m-%d")
                if year and record_date.year != year:
                    continue
                if month and record_date.month != month:
                    continue

                # 如果通过所有筛选条件，则添加到统计中
                stats['total_days'] += 1

                # 按类型统计
                record_type = record.get('type', '未知')
                stats['records_by_type'][record_type] = stats['records_by_type'].get(record_type, 0) + 1

                # 按月份统计
                month_key = f"{record_date.year}-{record_date.month:02d}"
                stats['records_by_month'][month_key] = stats['records_by_month'].get(month_key, 0) + 1

                # 按成员统计
                record_member = record.get('plan_name', '未知')
                stats['records_by_member'][record_member] = stats['records_by_member'].get(record_member, 0) + 1

            except (ValueError, KeyError) as e:
                continue

        return stats

    def get_leave_types_color_mapping(self):
        """获取请假类型的颜色映射，用于日历显示

        Returns:
            dict: 请假类型到颜色的映射
        """
        # 默认颜色方案
        default_colors = {
            '年假': '#4CAF50',      # 绿色
            '事假': '#FF9800',      # 橙色
            '病假': '#F44336',      # 红色
            '育儿假': '#9C27B0',    # 紫色
            '婚假': '#E91E63',      # 粉色
            '丧假': '#607D8B',      # 蓝灰色
            '调休': '#00BCD4',      # 青色
            '其他': '#9E9E9E'       # 灰色
        }

        # 获取所有实际使用的请假类型
        all_types = self.get_leave_types_for_holiday_calendar()

        # 为每种类型分配颜色
        color_mapping = {}
        color_index = 0
        available_colors = [
            '#4CAF50', '#FF9800', '#F44336', '#9C27B0', '#E91E63',
            '#607D8B', '#00BCD4', '#795548', '#FF5722', '#3F51B5',
            '#009688', '#CDDC39', '#8BC34A', '#FFC107', '#FFEB3B'
        ]

        for leave_type in all_types:
            if leave_type in default_colors:
                color_mapping[leave_type] = default_colors[leave_type]
            else:
                color_mapping[leave_type] = available_colors[color_index % len(available_colors)]
                color_index += 1

        return color_mapping

    def update_quota_year_options(self):
        """更新配额年份选择下拉框的选项"""
        try:
            years = set()

            # 获取请假记录中的所有年份，并按照年休假规则转换
            for rec in self.leave_records:
                date_str = rec.get("date", "")
                leave_type = rec.get("type", "")
                if not date_str:
                    continue

                try:
                    parts = date_str.split('-')
                    record_year = int(parts[0])
                    record_month = int(parts[1])

                    # 根据年休假规则确定配额年份
                    if self._is_annual_leave(leave_type):
                        # 年休假：4-12月属于当年配额，1-3月属于上年配额
                        if record_month >= 4:
                            quota_year = record_year
                        else:
                            quota_year = record_year - 1
                    else:
                        # 其他假期类型使用自然年
                        quota_year = record_year

                    years.add(quota_year)
                except Exception:
                    continue

            # 获取配额数据中的所有年份
            for plan, plan_quotas in self.leave_quotas.items():
                for year_str in plan_quotas.keys():
                    try:
                        years.add(int(year_str))
                    except Exception:
                        continue

            # 添加当前年份前后的年份作为选项
            current_date = datetime.date.today()
            if current_date.month >= 4:
                default_leave_year = current_date.year
            else:
                default_leave_year = current_date.year - 1

            # 添加当前年份前后3年的选项
            for y in range(default_leave_year - 3, default_leave_year + 4):
                years.add(y)

            # 如果没有找到年份，使用默认范围
            if not years:
                y = self.current_date.year
                years = set(range(y - 2, y + 6))  # 默认当前年-2 到 +5

            values = sorted(years)
            if hasattr(self, 'quota_year_combo'):
                self.quota_year_combo["values"] = values
                # 确保当前值在选项中
                current_val = self.quota_year_var.get()
                if current_val not in [str(v) for v in values]:
                    if values:
                        self.quota_year_var.set(str(default_leave_year))
                        self.update_quota_summary()
                else:
                    # 触发更新
                    self.update_quota_summary()
        except Exception as e:
            print(f"更新配额年份选项失败: {e}")

    def on_quota_year_selected(self, event=None):
        """年份选择改变事件处理"""
        try:
            year = self.quota_year_var.get()
            if year:
                self.update_quota_summary()
                self.update_current_leave_year_display()
                self.update_status(f"已切换到 {year} 年配额视图")
        except Exception as e:
            print(f"年份选择处理失败: {e}")

    def update_current_leave_year_display(self):
        """更新当前年休假年度的醒目显示"""
        try:
            # 确保必要的属性存在
            if not hasattr(self, 'quota_year_var') or not hasattr(self, 'current_leave_year_label'):
                return

            current_date = datetime.date.today()

            # 确定当前年休假年度
            if current_date.month >= 4:
                # 4-12月：属于当前年份的年休假年度
                leave_year = current_date.year
                period = f"{leave_year}年4月 - {leave_year+1}年3月"
            else:
                # 1-3月：属于上一年的年休假年度
                leave_year = current_date.year - 1
                period = f"{leave_year}年4月 - {leave_year+1}年3月"

            # 获取选择的年份
            selected_year = self.quota_year_var.get()
            if selected_year:
                try:
                    selected_year = int(selected_year)
                except ValueError:
                    selected_year = None

                if selected_year is not None:
                    # 如果选择的是当年年休假年度，显示特殊提示
                    if selected_year == leave_year:
                        display_text = f"📅 当前年休假年度: {period}"
                        color = self.colors.get('success', '#107c10')
                    else:
                        display_text = f"📅 年休假年度: {selected_year}年4月 - {selected_year+1}年3月"
                        color = self.colors.get('accent', '#0078d4')
                else:
                    display_text = f"📅 当前年休假年度: {period}"
                    color = self.colors.get('success', '#107c10')
            else:
                display_text = f"📅 当前年休假年度: {period}"
                color = self.colors.get('success', '#107c10')

            self.current_leave_year_label.config(text=display_text, foreground=color)

        except Exception as e:
            print(f"更新年休假年度显示失败: {e}")
            if hasattr(self, 'current_leave_year_label'):
                try:
                    self.current_leave_year_label.config(text="")
                except Exception:
                    pass

    def on_quota_double_click(self, event):
        """年度配额树视图双击事件处理 - 编辑选中配额"""
        sel = self.leave_quota_tree.selection()
        if not sel:
            return

        item = self.leave_quota_tree.item(sel[0])
        vals = item.get("values", [])
        if len(vals) < 6:
            messagebox.showwarning("警告", "数据不完整")
            return

        plan = vals[0]  # 人员名称
        year = vals[2]  # 年份

        # 打开编辑对话框，预填人员和年份
        self.edit_quota_dialog(plan, year)

    def edit_quota_dialog(self, plan=None, year=None):
        """编辑配额的对话框（可预填人员和年份）"""
        top = tk.Toplevel(self.root)
        top.title("编辑年度配额")
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.grid(row=0, column=0, sticky=tk.NSEW)

        ttk.Label(frm, text="人员名称").grid(row=0, column=0, sticky=tk.W, pady=4)
        plan_var = tk.StringVar(value=plan if plan else (list(self.shift_schedules.keys())[0] if self.shift_schedules else ""))
        plan_combo = ttk.Combobox(frm, textvariable=plan_var, values=list(self.shift_schedules.keys()), state="readonly", width=18)
        plan_combo.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(frm, text="类型").grid(row=1, column=0, sticky=tk.W, pady=4)
        type_var = tk.StringVar(value=(self.leave_types[0] if self.leave_types else ""))
        type_combo = ttk.Combobox(frm, textvariable=type_var, values=self.leave_types, state="readonly", width=18)
        type_combo.grid(row=1, column=1, sticky=tk.W)

        ttk.Label(frm, text="年份").grid(row=2, column=0, sticky=tk.W, pady=4)
        year_var = tk.StringVar(value=year if year else str(self.current_date.year))
        year_combo = ttk.Combobox(frm, textvariable=year_var, values=[str(y) for y in range(self.current_date.year-2, self.current_date.year+6)], state="readonly", width=10)
        year_combo.grid(row=2, column=1, sticky=tk.W)

        ttk.Label(frm, text="配额(天)").grid(row=3, column=0, sticky=tk.W, pady=4)
        quota_var = tk.StringVar(value="0")
        quota_entry = ttk.Entry(frm, textvariable=quota_var, width=10)
        quota_entry.grid(row=3, column=1, sticky=tk.W)

        def update_quota_display():
            """更新配额显示（当选择改变时调用）"""
            sel_plan = plan_var.get().strip()
            sel_type = type_var.get().strip()
            sel_year = year_var.get().strip()

            if not sel_plan or not sel_type or not sel_year:
                return

            # 从数据中加载该人员、该年份、该类型的配额
            current_quota = self.leave_quotas.get(sel_plan, {}).get(sel_year, {}).get(sel_type, 0)
            quota_var.set(str(current_quota))

        def copy_last_year_quota():
            """一键复制上一年配额到当前年份"""
            sel_year = year_var.get().strip()
            if not sel_year:
                messagebox.showwarning("警告", "请先选择年份")
                return

            try:
                current_year = int(sel_year)
                last_year = str(current_year - 1)
            except ValueError:
                messagebox.showwarning("警告", "年份格式错误")
                return

            # 检查上一年是否有配额数据
            has_last_year_data = False
            for plan_name in self.shift_schedules.keys():
                if last_year in self.leave_quotas.get(plan_name, {}):
                    has_last_year_data = True
                    break

            if not has_last_year_data:
                messagebox.showinfo("提示", f"{last_year}年没有配额数据可复制")
                return

            # 确认操作
            if not messagebox.askyesno("确认", f"确定要将{last_year}年的配额复制到{sel_year}年吗？\n这将覆盖{sel_year}年已有的配额设置。"):
                return

            # 复制上一年配额到当前年份
            copied_count = 0
            for plan_name in self.shift_schedules.keys():
                last_year_quotas = self.leave_quotas.get(plan_name, {}).get(last_year, {})
                if last_year_quotas:
                    self.leave_quotas.setdefault(plan_name, {}).setdefault(sel_year, {}).update(last_year_quotas.copy())
                    copied_count += 1

            if copied_count > 0:
                self.save_data()
                self.update_quota_summary()
                # 更新当前显示的配额
                update_quota_display()
                messagebox.showinfo("成功", f"已将{last_year}年的配额复制到{sel_year}年\n共更新{copied_count}个人员的配额")
                self.update_status(f"已复制{last_year}年配额到{sel_year}年")
            else:
                messagebox.showinfo("提示", "没有找到可复制的配额数据")

        # 为下拉框绑定事件，当选择改变时更新配额显示
        plan_combo.bind('<<ComboboxSelected>>', lambda e: update_quota_display())
        type_combo.bind('<<ComboboxSelected>>', lambda e: update_quota_display())
        year_combo.bind('<<ComboboxSelected>>', lambda e: update_quota_display())

        # 初始化配额显示
        top.after(100, update_quota_display)
        quota_entry.focus()

        def on_save_quota():
            sel_plan = plan_var.get().strip()
            sel_type = type_var.get().strip()
            sel_year = year_var.get().strip()
            try:
                q = int(quota_var.get().strip())
            except Exception:
                messagebox.showwarning("警告", "配额需为整数")
                return
            if not sel_plan or not sel_type or not sel_year:
                messagebox.showwarning("警告", "请完整选择计划、类型与年份")
                return
            self.leave_quotas.setdefault(sel_plan, {}).setdefault(sel_year, {})[sel_type] = max(q, 0)
            self.save_data()
            self.update_quota_summary()
            self.update_status(f"已更新配额: {sel_plan} {sel_year} {sel_type} = {q}")
            top.destroy()

        action = ttk.Frame(top, padding=(0, 10))
        action.grid(row=4, column=0, columnspan=2, sticky=tk.EW)
        ttk.Button(action, text="复制上年配额", command=copy_last_year_quota).pack(side=tk.LEFT, padx=6)
        ttk.Button(action, text="取消", command=top.destroy).pack(side=tk.RIGHT)
        ttk.Button(action, text="保存", command=on_save_quota).pack(side=tk.RIGHT, padx=6)

        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)


    def import_leave_records_from_excel(self):
        """从Excel文件导入请假记录"""
        if pd is None:
            messagebox.showerror("错误", "缺少pandas库，请安装：pip install pandas")
            return

        # 选择文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)

            # 验证列结构
            required_columns = ['员工姓名', '请假日期', '请假类型', '备注']
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("错误", f"文件格式错误，需要包含以下列：{', '.join(required_columns)}")
                return

            # 验证数据完整性
            if df.isnull().any().any():
                messagebox.showerror("错误", "存在空值，请确保所有字段都已填写")
                return

            # 验证日期格式
            date_pattern = r'^\d{4}-\d{2}-\d{2}$'
            invalid_dates = []
            for idx, date_str in enumerate(df['请假日期']):
                if not re.match(date_pattern, str(date_str)):
                    invalid_dates.append(f"第{idx+2}行: {date_str}")

            if invalid_dates:
                messagebox.showerror("错误", f"日期格式错误，应为YYYY-MM-DD：\n" + "\n".join(invalid_dates))
                return

            # 验证员工姓名和请假类型
            invalid_employees = []
            invalid_types = []
            for idx, row in df.iterrows():
                employee = str(row['员工姓名']).strip()
                leave_type = str(row['请假类型']).strip()

                if employee not in self.shift_schedules:
                    invalid_employees.append(f"第{idx+2}行: {employee}")

                if leave_type not in self.leave_types:
                    invalid_types.append(f"第{idx+2}行: {leave_type}")

            if invalid_employees:
                messagebox.showerror("错误", f"以下员工不存在：\n" + "\n".join(invalid_employees))
                return

            if invalid_types:
                messagebox.showerror("错误", f"以下请假类型不存在：\n" + "\n".join(invalid_types))
                return

            # 检查重复记录
            duplicates = []
            for idx, row in df.iterrows():
                employee = str(row['员工姓名']).strip()
                date_str = str(row['请假日期']).strip()
                leave_type = str(row['请假类型']).strip()

                # 检查是否已存在相同记录
                for record in self.leave_records:
                    if (record.get("plan_name") == employee and
                        record.get("date") == date_str and
                        record.get("type") == leave_type):
                        duplicates.append(f"第{idx+2}行: {employee} {date_str} {leave_type}")
                        break

            if duplicates:
                result = messagebox.askyesno("重复记录",
                    f"发现以下重复记录：\n" + "\n".join(duplicates[:5]) +
                    ("\n..." if len(duplicates) > 5 else "") +
                    "\n\n是否跳过重复记录继续导入？")
                if not result:
                    return

            # 导入记录
            imported_count = 0
            skipped_count = 0
            for idx, row in df.iterrows():
                employee = str(row['员工姓名']).strip()
                date_str = str(row['请假日期']).strip()
                leave_type = str(row['请假类型']).strip()
                note = str(row['备注']).strip()

                # 检查是否重复
                is_duplicate = False
                for record in self.leave_records:
                    if (record.get("plan_name") == employee and
                        record.get("date") == date_str and
                        record.get("type") == leave_type):
                        is_duplicate = True
                        break

                if is_duplicate:
                    skipped_count += 1
                    continue

                # 添加记录
                self.leave_records.append({
                    "plan_name": employee,
                    "date": date_str,
                    "type": leave_type,
                    "note": note
                })
                imported_count += 1

            # 保存数据和更新界面
            self.save_data()
            self.update_leave_tree()
            self.update_leave_stats()
            self.update_calendar()

            messagebox.showinfo("导入完成",
                f"成功导入 {imported_count} 条记录\n" +
                f"跳过重复记录 {skipped_count} 条")
            self.update_status(f"已从Excel导入 {imported_count} 条请假记录")

        except Exception as e:
            messagebox.showerror("导入失败", f"导入过程中发生错误：\n{str(e)}")

    def export_leave_records_to_excel(self):
        """导出请假记录到Excel文件"""
        if pd is None or Workbook is None:
            messagebox.showerror("错误", "缺少必要的库，请安装：pip install pandas openpyxl")
            return

        if not self.leave_records:
            messagebox.showwarning("警告", "没有请假记录可以导出")
            return

        # 选择保存路径
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            # 准备数据
            data = []
            for record in self.leave_records:
                data.append({
                    '员工姓名': record.get('plan_name', ''),
                    '请假日期': record.get('date', ''),
                    '请假类型': record.get('type', ''),
                    '备注': record.get('note', '')
                })

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "请假记录"

            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 写入表头
            headers = ['员工姓名', '请假日期', '请假类型', '备注']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 写入数据
            data_font = Font(color="000000")
            data_alignment = Alignment(horizontal="left", vertical="center")

            for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment

            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 员工姓名
            ws.column_dimensions['B'].width = 12  # 请假日期
            ws.column_dimensions['C'].width = 12  # 请假类型
            ws.column_dimensions['D'].width = 40  # 备注

            # 保存文件
            wb.save(file_path)

            messagebox.showinfo("导出完成", f"成功导出 {len(self.leave_records)} 条记录到：\n{file_path}")
            self.update_status(f"已导出 {len(self.leave_records)} 条请假记录到Excel")

        except Exception as e:
            messagebox.showerror("导出失败", f"导出过程中发生错误：\n{str(e)}")

    def download_import_template(self):
        """下载导入模板"""
        if pd is None or Workbook is None:
            messagebox.showerror("错误", "缺少必要的库，请安装：pip install pandas openpyxl")
            return

        # 选择保存路径
        file_path = filedialog.asksaveasfilename(
            title="保存导入模板",
            defaultextension=".xlsx",
            initialfile="请假记录导入模板.xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            # 创建模板数据
            template_data = [
                {
                    '员工姓名': '张三',
                    '请假日期': '2024-01-15',
                    '请假类型': '病假',
                    '备注': '感冒发烧，需要就医'
                },
                {
                    '员工姓名': '李四',
                    '请假日期': '2024-01-16',
                    '请假类型': '事假',
                    '备注': '家中有事，急需处理'
                },
                {
                    '员工姓名': '王五',
                    '请假日期': '2024-01-17',
                    '请假类型': '年假',
                    '备注': '年度体检，请假1天'
                }
            ]

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "请假记录"

            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 写入表头
            headers = ['员工姓名', '请假日期', '请假类型', '备注']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 写入示例数据
            data_font = Font(color="000000")
            data_alignment = Alignment(horizontal="left", vertical="center")

            for row_idx, row_data in enumerate(template_data, 2):
                for col_idx, (key, value) in enumerate(row_data.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment

            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 员工姓名
            ws.column_dimensions['B'].width = 12  # 请假日期
            ws.column_dimensions['C'].width = 12  # 请假类型
            ws.column_dimensions['D'].width = 40  # 备注

            # 添加说明工作表
            if len(wb.sheetnames) == 1:
                ws_instructions = wb.create_sheet("使用说明")
                instructions = [
                    "请假记录导入模板使用说明",
                    "",
                    "1. 数据格式要求：",
                    "   - 员工姓名：必须是在系统中存在的人员名称",
                    "   - 请假日期：格式为 YYYY-MM-DD，例如：2024-01-15",
                    "   - 请假类型：必须是系统中已定义的请假类型",
                    "   - 备注：请假原因的详细说明",
                    "",
                    "2. 注意事项：",
                    "   - 所有字段均为必填项",
                    "   - 日期格式必须正确",
                    "   - 员工姓名和请假类型必须在系统中存在",
                    "   - 系统会自动检测并提示重复记录",
                    "",
                    "3. 导入流程：",
                    "   - 填写完成后保存为 .xlsx 格式",
                    "   - 在系统中的请假管理页面点击'导入Excel'",
                    "   - 选择文件并按照提示完成导入",
                    "",
                    "4. 支持的请假类型：",
                    "   " + "、".join(self.leave_types) if self.leave_types else "   请在系统中先定义请假类型"
                ]

                for row_idx, instruction in enumerate(instructions, 1):
                    cell = ws_instructions.cell(row=row_idx, column=1, value=instruction)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                ws_instructions.column_dimensions['A'].width = 80

            # 保存模板文件
            wb.save(file_path)

            messagebox.showinfo("模板下载完成", f"导入模板已保存到：\n{file_path}")
            self.update_status("已下载请假记录导入模板")

        except Exception as e:
            messagebox.showerror("模板下载失败", f"下载模板过程中发生错误：\n{str(e)}")

    # ==================== 调换班功能 ====================

    def _bind_context_menu_recursive(self, widget, member, date_str):
        """递归地为控件及其所有子控件绑定右键菜单

        Args:
            widget: 要绑定的控件
            member: 人员姓名
            date_str: 日期字符串 YYYY-MM-DD
        """
        widget.bind("<Button-3>", lambda e, m=member, d=date_str: self._show_multi_calendar_context_menu(e, m, d))
        for child in widget.winfo_children():
            self._bind_context_menu_recursive(child, member, date_str)

    def _show_multi_calendar_context_menu(self, event, member, date_str):
        """显示多人日历格子的右键菜单

        Args:
            event: 鼠标事件
            member: 人员姓名
            date_str: 日期字符串 YYYY-MM-DD
        """
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"调换班 ({member})", command=lambda: self.show_swap_shift_dialog(member, date_str))

        # 如果有调换记录，添加还原选项
        if self.check_swap_record(member, date_str):
            menu.add_separator()
            menu.add_command(label=f"还原调换班 ({member})", command=lambda: self._do_restore_swap(member, date_str))

        menu.post(event.x_root, event.y_root)

    def _do_restore_swap(self, member, date_str):
        """执行还原调换班操作并刷新界面

        Args:
            member: 人员姓名
            date_str: 日期字符串 YYYY-MM-DD
        """
        success, message = self.restore_swap(member, date_str)

        if success:
            # 先刷新日历显示
            if hasattr(self, "multi_calendar_container"):
                self.update_multi_calendar()
            # 再显示成功消息
            messagebox.showinfo("成功", message)
        else:
            messagebox.showerror("错误", message)

    def check_swap_record(self, person, date_str, shift_type=None):
        """检查某人在某日期的某班次是否有调换班记录

        Args:
            person: 人员姓名
            date_str: 日期字符串 YYYY-MM-DD
            shift_type: 班次类型（可选），如果提供则检查该班次是否被调换

        Returns:
            bool: 是否有调换班记录
        """
        if date_str not in self.swap_records:
            return False

        for record in self.swap_records[date_str]:
            person_a = record.get("person_a")
            person_b = record.get("person_b")
            date_a = record.get("date_a")
            date_b = record.get("date_b")
            shift_a_original = record.get("shift_a_original")
            shift_b_original = record.get("shift_b_original")

            # 如果没有指定班次类型，只检查人员和日期
            if shift_type is None:
                if person_a == person or person_b == person:
                    return True
            else:
                # 检查该班次是否是调换来的
                # person_a在date_b得到了shift_b_original
                if person == person_a and date_str == date_b and shift_type == shift_b_original:
                    return True
                # person_b在date_a得到了shift_a_original
                if person == person_b and date_str == date_a and shift_type == shift_a_original:
                    return True

        return False

    def _add_shift(self, person, date, shift):
        """添加班次到指定日期（支持同一天多个班次）"""
        if "shifts" not in self.shift_schedules[person]:
            self.shift_schedules[person]["shifts"] = {}

        current = self.shift_schedules[person]["shifts"].get(date)
        if current is None:
            # 没有班次，直接设置
            self.shift_schedules[person]["shifts"][date] = shift
        elif isinstance(current, list):
            # 已经是列表，添加到列表
            if shift not in current:
                current.append(shift)
        else:
            # 是单个班次，转换为列表
            if current != shift:
                self.shift_schedules[person]["shifts"][date] = [current, shift]

    def _remove_shift(self, person, date, shift):
        """从指定日期删除班次"""
        if "shifts" not in self.shift_schedules[person]:
            return

        current = self.shift_schedules[person]["shifts"].get(date)
        if current is None:
            return
        elif isinstance(current, list):
            # 是列表，删除指定班次
            if shift in current:
                current.remove(shift)
                # 如果列表只剩一个元素，转换回字符串
                if len(current) == 1:
                    self.shift_schedules[person]["shifts"][date] = current[0]
                elif len(current) == 0:
                    del self.shift_schedules[person]["shifts"][date]
        else:
            # 是单个班次，直接删除
            if current == shift:
                del self.shift_schedules[person]["shifts"][date]

    def swap_shifts(self, person_a, person_b, date_a, date_b):
        """执行调换班操作（支持跨日期调换）

        Args:
            person_a: 人员A姓名
            person_b: 人员B姓名
            date_a: 人员A的日期字符串 YYYY-MM-DD
            date_b: 人员B的日期字符串 YYYY-MM-DD

        Returns:
            tuple: (success: bool, message: str)
        """
        # 验证输入
        if not person_a or not person_b:
            return False, "人员姓名不能为空"

        if person_a == person_b and date_a == date_b:
            return False, "不能与自己在同一天调换班"

        if person_a not in self.shift_schedules:
            return False, f"人员 {person_a} 不存在"

        if person_b not in self.shift_schedules:
            return False, f"人员 {person_b} 不存在"

        # 获取两人各自日期的班次
        shift_a = self.shift_schedules[person_a].get("shifts", {}).get(date_a)
        shift_b = self.shift_schedules[person_b].get("shifts", {}).get(date_b)

        if not shift_a:
            return False, f"{person_a} 在 {date_a} 没有排班"

        if not shift_b:
            return False, f"{person_b} 在 {date_b} 没有排班"

        # 执行交换
        if date_a == date_b:
            # 同一天调换：只交换班次类型
            self.shift_schedules[person_a]["shifts"][date_a] = shift_b
            self.shift_schedules[person_b]["shifts"][date_b] = shift_a
        else:
            # 跨日期调换：交换工作日期
            # 删除原有排班
            if "shifts" in self.shift_schedules[person_a]:
                self._remove_shift(person_a, date_a, shift_a)
            if "shifts" in self.shift_schedules[person_b]:
                self._remove_shift(person_b, date_b, shift_b)

            # 添加新排班（支持同一天多个班次）
            self._add_shift(person_a, date_b, shift_b)
            self._add_shift(person_b, date_a, shift_a)

        # 记录调换（保存原始班次信息以便还原）
        # 使用唯一ID来标识这次调换
        swap_id = f"{person_a}_{date_a}_{person_b}_{date_b}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"

        # 在两个日期都记录调换信息
        for date_str in [date_a, date_b]:
            if date_str not in self.swap_records:
                self.swap_records[date_str] = []

            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.swap_records[date_str].append({
                "swap_id": swap_id,
                "person_a": person_a,
                "person_b": person_b,
                "date_a": date_a,
                "date_b": date_b,
                "shift_a_original": shift_a,
                "shift_b_original": shift_b,
                "timestamp": timestamp
            })

        # 保存数据
        self.save_data()

        # 强制清除所有缓存，确保多人日历能显示最新数据
        if hasattr(self, '_calendar_data_cache'):
            self._calendar_data_cache.clear()
        if hasattr(self, '_multi_calendar_cell_cache'):
            self._multi_calendar_cell_cache.clear()

        if date_a == date_b:
            return True, f"成功调换班次:\n{person_a}: {shift_a} → {shift_b}\n{person_b}: {shift_b} → {shift_a}"
        else:
            return True, f"成功调换班次:\n{person_a} ({date_a}): {shift_a} → {shift_b}\n{person_b} ({date_b}): {shift_b} → {shift_a}"

    def restore_swap(self, person, date_str):
        """还原调换班操作

        Args:
            person: 人员姓名
            date_str: 日期字符串 YYYY-MM-DD

        Returns:
            tuple: (success: bool, message: str)
        """
        if date_str not in self.swap_records:
            return False, "该日期没有调换记录"

        # 查找该人员的调换记录
        swap_record = None
        for record in self.swap_records[date_str]:
            if record.get("person_a") == person or record.get("person_b") == person:
                swap_record = record
                break

        if not swap_record:
            return False, f"{person} 在 {date_str} 没有调换记录"

        # 获取调换信息
        swap_id = swap_record.get("swap_id")
        person_a = swap_record.get("person_a")
        person_b = swap_record.get("person_b")
        date_a = swap_record.get("date_a")
        date_b = swap_record.get("date_b")
        shift_a_original = swap_record.get("shift_a_original")
        shift_b_original = swap_record.get("shift_b_original")

        if not all([person_a, person_b, date_a, date_b, shift_a_original, shift_b_original]):
            return False, "调换记录数据不完整，无法还原"

        # 还原班次
        if date_a == date_b:
            # 同一天调换的还原：直接恢复班次类型
            self.shift_schedules[person_a]["shifts"][date_a] = shift_a_original
            self.shift_schedules[person_b]["shifts"][date_b] = shift_b_original
        else:
            # 跨日期调换的还原：恢复原有日期的排班
            # 删除调换后的排班
            if "shifts" in self.shift_schedules[person_a]:
                self._remove_shift(person_a, date_b, shift_b_original)
            if "shifts" in self.shift_schedules[person_b]:
                self._remove_shift(person_b, date_a, shift_a_original)

            # 恢复原有排班
            self._add_shift(person_a, date_a, shift_a_original)
            self._add_shift(person_b, date_b, shift_b_original)

        # 删除两个日期的调换记录（使用 swap_id 匹配）
        for date in [date_a, date_b]:
            if date in self.swap_records:
                self.swap_records[date] = [r for r in self.swap_records[date] if r.get("swap_id") != swap_id]
                if not self.swap_records[date]:
                    del self.swap_records[date]

        # 保存数据
        self.save_data()

        # 强制清除所有缓存，确保多人日历能显示最新数据
        if hasattr(self, '_calendar_data_cache'):
            self._calendar_data_cache.clear()
        if hasattr(self, '_multi_calendar_cell_cache'):
            self._multi_calendar_cell_cache.clear()

        if date_a == date_b:
            return True, f"成功还原班次:\n{person_a}: {shift_a_original}\n{person_b}: {shift_b_original}"
        else:
            return True, f"成功还原班次:\n{person_a} ({date_a}): {shift_a_original}\n{person_b} ({date_b}): {shift_b_original}"

    def show_swap_shift_dialog(self, person, date_str):
        """显示调换班对话框（支持跨日期调换）

        Args:
            person: 当前选中的人员姓名
            date_str: 日期字符串 YYYY-MM-DD
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("调换班")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # 主框架
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = tk.Label(main_frame, text="调换班",
                              font=("Microsoft YaHei UI", 12, "bold"))
        title_label.pack(pady=(0, 15))

        # 当前人员信息
        info_frame = ttk.LabelFrame(main_frame, text="当前人员", padding=10)
        info_frame.pack(fill=tk.X, pady=(0, 15))

        tk.Label(info_frame, text=f"姓名: {person}").pack(anchor=tk.W)
        tk.Label(info_frame, text=f"日期: {date_str}").pack(anchor=tk.W)
        current_shift = self.shift_schedules.get(person, {}).get("shifts", {}).get(date_str, "无")
        tk.Label(info_frame, text=f"班次: {current_shift}").pack(anchor=tk.W)

        # 选择调换对象
        select_frame = ttk.LabelFrame(main_frame, text="选择调换对象", padding=10)
        select_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # 选择人员
        tk.Label(select_frame, text="选择人员:").pack(anchor=tk.W, pady=(0, 5))

        # 获取所有人员（排除当前人员）
        all_persons = [p for p in self.shift_schedules.keys() if p != person]

        if not all_persons:
            tk.Label(select_frame, text="没有其他人员可以调换",
                    fg="red").pack(pady=10)
            ttk.Button(main_frame, text="关闭",
                      command=dialog.destroy).pack()
            return

        person_var = tk.StringVar(value=all_persons[0])
        person_combo = ttk.Combobox(select_frame, textvariable=person_var,
                                   values=all_persons, state="readonly")
        person_combo.pack(fill=tk.X, pady=(0, 10))

        # 选择目标日期
        tk.Label(select_frame, text="选择该人员的日期:").pack(anchor=tk.W, pady=(0, 5))

        # 解析初始日期
        year, month, day = map(int, date_str.split('-'))
        target_date_entry = DateEntry(select_frame, width=20, background='darkblue',
                                      foreground='white', borderwidth=2,
                                      year=year, month=month, day=day,
                                      date_pattern='yyyy-mm-dd')
        target_date_entry.pack(fill=tk.X, pady=(0, 10))

        # 显示目标人员在目标日期的班次
        target_info_label = tk.Label(select_frame, text="", fg="blue")
        target_info_label.pack(anchor=tk.W, pady=(0, 5))

        target_shift_label = tk.Label(select_frame, text="")
        target_shift_label.pack(anchor=tk.W)

        no_shift_label = tk.Label(select_frame, text="", fg="red")
        no_shift_label.pack(anchor=tk.W)

        def update_target_info(*args):
            """更新目标人员的班次信息"""
            target_person = person_var.get()
            target_date = target_date_entry.get_date().strftime("%Y-%m-%d")

            target_shift = self.shift_schedules.get(target_person, {}).get("shifts", {}).get(target_date)

            if target_shift:
                target_info_label.config(text=f"{target_person} 在 {target_date}")
                target_shift_label.config(text=f"班次: {target_shift}")
                no_shift_label.config(text="")
            else:
                target_info_label.config(text=f"{target_person} 在 {target_date}")
                target_shift_label.config(text="")
                no_shift_label.config(text="该日期没有排班，无法调换")

        # 绑定事件
        person_var.trace("w", update_target_info)
        target_date_entry.bind("<<DateEntrySelected>>", update_target_info)

        # 初始化显示
        update_target_info()

        # 按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        def do_swap():
            target_person = person_var.get()
            target_date = target_date_entry.get_date().strftime("%Y-%m-%d")

            # 检查目标人员在目标日期是否有排班
            target_shift = self.shift_schedules.get(target_person, {}).get("shifts", {}).get(target_date)
            if not target_shift:
                messagebox.showerror("错误", f"{target_person} 在 {target_date} 没有排班，无法调换")
                return

            success, message = self.swap_shifts(person, target_person, date_str, target_date)

            if success:
                dialog.destroy()
                # 刷新日历显示
                if hasattr(self, "multi_calendar_container"):
                    self.update_multi_calendar()
                # 最后显示成功消息
                messagebox.showinfo("成功", message)
            else:
                messagebox.showerror("错误", message)

        ttk.Button(button_frame, text="确认调换",
                  command=do_swap).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="取消",
                  command=dialog.destroy).pack(side=tk.LEFT)

    def refresh_swap_list(self):
        """刷新调班记录列表"""
        # 清空现有数据
        for item in self.swap_tree.get_children():
            self.swap_tree.delete(item)

        # 收集所有唯一的调班记录
        seen_swap_ids = set()
        swap_records_list = []

        for date_str, records in self.swap_records.items():
            for record in records:
                swap_id = record.get("swap_id")
                if swap_id and swap_id not in seen_swap_ids:
                    seen_swap_ids.add(swap_id)
                    swap_records_list.append(record)

        # 按时间戳排序（最新的在前）
        swap_records_list.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

        # 插入数据
        for record in swap_records_list:
            self.swap_tree.insert("", tk.END, values=(
                record.get("swap_id", ""),
                record.get("person_a", ""),
                record.get("date_a", ""),
                record.get("shift_a_original", ""),
                record.get("person_b", ""),
                record.get("date_b", ""),
                record.get("shift_b_original", ""),
                record.get("timestamp", "")
            ))

    def add_swap_record(self):
        """新增调班记录"""
        dialog = tk.Toplevel(self.root)
        dialog.title("新增调班")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="新增调班", font=("Microsoft YaHei UI", 12, "bold")).pack(pady=(0, 15))

        # 人员A
        frame_a = ttk.LabelFrame(main_frame, text="人员A", padding=10)
        frame_a.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(frame_a, text="选择人员:").pack(anchor=tk.W, pady=(0, 5))
        person_a_var = tk.StringVar()
        person_a_combo = ttk.Combobox(frame_a, textvariable=person_a_var,
                                     values=list(self.shift_schedules.keys()), state="readonly")
        person_a_combo.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(frame_a, text="选择日期:").pack(anchor=tk.W, pady=(0, 5))
        date_a_entry = DateEntry(frame_a, width=20, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        date_a_entry.pack(fill=tk.X, pady=(0, 10))

        shift_a_label = ttk.Label(frame_a, text="", foreground="blue")
        shift_a_label.pack(anchor=tk.W)

        # 人员B
        frame_b = ttk.LabelFrame(main_frame, text="人员B", padding=10)
        frame_b.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(frame_b, text="选择人员:").pack(anchor=tk.W, pady=(0, 5))
        person_b_var = tk.StringVar()
        person_b_combo = ttk.Combobox(frame_b, textvariable=person_b_var,
                                     values=list(self.shift_schedules.keys()), state="readonly")
        person_b_combo.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(frame_b, text="选择日期:").pack(anchor=tk.W, pady=(0, 5))
        date_b_entry = DateEntry(frame_b, width=20, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        date_b_entry.pack(fill=tk.X, pady=(0, 10))

        shift_b_label = ttk.Label(frame_b, text="", foreground="blue")
        shift_b_label.pack(anchor=tk.W)

        def update_shift_info(*args):
            if person_a_var.get():
                date_a = date_a_entry.get_date().strftime("%Y-%m-%d")
                shift_a = self.shift_schedules.get(person_a_var.get(), {}).get("shifts", {}).get(date_a, "无排班")
                shift_a_label.config(text=f"班次: {shift_a}")

            if person_b_var.get():
                date_b = date_b_entry.get_date().strftime("%Y-%m-%d")
                shift_b = self.shift_schedules.get(person_b_var.get(), {}).get("shifts", {}).get(date_b, "无排班")
                shift_b_label.config(text=f"班次: {shift_b}")

        person_a_var.trace("w", update_shift_info)
        person_b_var.trace("w", update_shift_info)
        date_a_entry.bind("<<DateEntrySelected>>", update_shift_info)
        date_b_entry.bind("<<DateEntrySelected>>", update_shift_info)

        # 按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        def do_add():
            person_a = person_a_var.get()
            person_b = person_b_var.get()
            date_a = date_a_entry.get_date().strftime("%Y-%m-%d")
            date_b = date_b_entry.get_date().strftime("%Y-%m-%d")

            if not person_a or not person_b:
                messagebox.showerror("错误", "请选择人员")
                return

            success, message = self.swap_shifts(person_a, person_b, date_a, date_b)
            if success:
                dialog.destroy()
                self.refresh_swap_list()
                if hasattr(self, "multi_calendar_container"):
                    self.update_multi_calendar()
                messagebox.showinfo("成功", message)
            else:
                messagebox.showerror("错误", message)

        ttk.Button(button_frame, text="确认", command=do_add, style='Success.TButton').pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT)

    def show_swap_context_menu(self, event):
        """显示调班记录右键菜单"""
        item = self.swap_tree.identify_row(event.y)
        if not item:
            return

        self.swap_tree.selection_set(item)
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="还原调班", command=lambda: self.restore_swap_from_list(item))
        menu.add_command(label="删除记录", command=lambda: self.delete_swap_record(item))
        menu.post(event.x_root, event.y_root)

    def restore_swap_from_list(self, item):
        """从列表还原调班"""
        values = self.swap_tree.item(item, "values")
        if not values:
            return

        person_a = values[1]
        date_a = values[2]

        result = messagebox.askyesno("确认", f"确定要还原 {person_a} 在 {date_a} 的调班吗？")
        if result:
            success, message = self.restore_swap(person_a, date_a)
            if success:
                self.refresh_swap_list()
                if hasattr(self, "multi_calendar_container"):
                    self.update_multi_calendar()
                messagebox.showinfo("成功", message)
            else:
                messagebox.showerror("错误", message)

    def delete_swap_record(self, item):
        """删除调班记录（仅删除记录，不还原班次）"""
        values = self.swap_tree.item(item, "values")
        if not values:
            return

        swap_id = values[0]
        result = messagebox.askyesno("确认", "确定要删除这条调班记录吗？\n注意：这不会还原班次，只是删除记录。")
        if result:
            # 从所有日期中删除该swap_id的记录
            for date_str in list(self.swap_records.keys()):
                self.swap_records[date_str] = [r for r in self.swap_records[date_str] if r.get("swap_id") != swap_id]
                if not self.swap_records[date_str]:
                    del self.swap_records[date_str]

            self.save_data()
            self.refresh_swap_list()
            messagebox.showinfo("成功", "调班记录已删除")

    def run(self):
        """运行主循环"""
        self.root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    app = ShiftScheduler(root)
    app.run()
